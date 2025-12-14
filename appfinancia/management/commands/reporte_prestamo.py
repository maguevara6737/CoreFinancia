# appfinancia/management/commands/reporte_prestamo.py

import os
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from appfinancia.models import (
    Prestamos, Historia_Prestamos, Pagos, Detalle_Aplicacion_Pago, Fechas_Sistema, Desembolsos
)

class Command(BaseCommand):
    help = 'Genera reporte cronológico del historial de un préstamo con desembolso y saldo pendiente'

    def add_arguments(self, parser):
        parser.add_argument('prestamo_id', type=int, help='ID del préstamo')

    def handle(self, *args, **options):
        prestamo_id = options['prestamo_id']

        try:
            prestamo = Prestamos.objects.get(prestamo_id=prestamo_id)
        except Prestamos.DoesNotExist:
            raise CommandError(f"Préstamo {prestamo_id} no existe.")

        # Obtener fecha de corte del sistema
        fecha_sistema = Fechas_Sistema.objects.first()
        if not fecha_sistema:
            raise CommandError("No hay registro en Fechas_Sistema.")
        fecha_corte = fecha_sistema.fecha_proceso_actual

        # === 1. Obtener datos del desembolso ===
        try:
            desembolso = Desembolsos.objects.get(prestamo_id=prestamo_id)
            monto_desembolso = float(desembolso.valor)
            fecha_desembolso = desembolso.fecha_desembolso
        except Desembolsos.DoesNotExist:
            monto_desembolso = 0.0
            fecha_desembolso = "No encontrado"

        # === 2. Obtener y agrupar cuotas ===
        cuotas = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo_id,
            fecha_vencimiento__lte=fecha_corte
        ).exclude(
            concepto_id__concepto_id="CAUSAC"
        ).order_by('fecha_vencimiento', 'numero_cuota')

        from collections import defaultdict
        cuotas_agrupadas = defaultdict(lambda: {
            'registro': None,
            'capital_prog': Decimal('0'),
            'interes_prog': Decimal('0'),
            'seguro_prog': Decimal('0'),
            'fee_prog': Decimal('0'),
            'capital_pag': Decimal('0'),
            'interes_pag': Decimal('0'),
            'seguro_pag': Decimal('0'),
            'fee_pag': Decimal('0'),
            'fecha_pago': None,
            'ref_bancaria': set(),
            'estado': 'PENDIENTE'
        })

        concepto_a_tipo = {
            'PLANCAP': 'capital',
            'PLANINT': 'interes',
            'PLANSEG': 'seguro',
            'PLANGTO': 'fee',
        }

        for reg in cuotas:
            num_cuota = reg.numero_cuota or 0
            tipo = concepto_a_tipo.get(reg.concepto_id.concepto_id)
            
            if tipo:
                cuotas_agrupadas[num_cuota][f'{tipo}_prog'] += reg.monto_transaccion

                detalles = Detalle_Aplicacion_Pago.objects.filter(historia_prestamo=reg)
                total_pagado = sum(d.monto_aplicado for d in detalles)
                cuotas_agrupadas[num_cuota][f'{tipo}_pag'] += total_pagado

                if detalles.exists():
                    fecha_pago_max = max(d.fecha_aplicacion.date() for d in detalles)
                    if cuotas_agrupadas[num_cuota]['fecha_pago'] is None or fecha_pago_max > cuotas_agrupadas[num_cuota]['fecha_pago']:
                        cuotas_agrupadas[num_cuota]['fecha_pago'] = fecha_pago_max
                    refs = {d.pago.ref_bancaria for d in detalles if d.pago.ref_bancaria}
                    cuotas_agrupadas[num_cuota]['ref_bancaria'].update(refs)

                if total_pagado < reg.monto_transaccion:
                    cuotas_agrupadas[num_cuota]['estado'] = 'PENDIENTE'
                # Si todo está pagado, el estado se mantendrá como "PAGADA" (por defecto)

        # === 3. Preparar datos para Excel ===
        datos_excel = []

        # Fila de desembolso
        datos_excel.append({
            'tipo': 'desembolso',
            'fecha_vencimiento': fecha_desembolso,
            'total_prog': monto_desembolso,
            'total_pag': monto_desembolso,  # asumimos que se desembolsó completo
            'estado': 'DESEMBOLSADO'
        })

        # Filas de cuotas
        total_pendiente_general = Decimal('0')
        for num_cuota in sorted(cuotas_agrupadas.keys()):
            c = cuotas_agrupadas[num_cuota]
            total_prog = c['capital_prog'] + c['interes_prog'] + c['seguro_prog'] + c['fee_prog']
            total_pag = c['capital_pag'] + c['interes_pag'] + c['seguro_pag'] + c['fee_pag']
            total_pend = total_prog - total_pag
            total_pendiente_general += total_pend

            estado_final = "PAGADA" if total_pend <= Decimal('0.01') else c['estado']

            datos_excel.append({
                'tipo': 'cuota',
                'cuota': num_cuota,
                'fecha_vencimiento': c['registro'].fecha_vencimiento if c['registro'] else '',
                'capital_prog': float(c['capital_prog']),
                'interes_prog': float(c['interes_prog']),
                'seguro_prog': float(c['seguro_prog']),
                'fee_prog': float(c['fee_prog']),
                'total_prog': float(total_prog),
                'capital_pag': float(c['capital_pag']),
                'interes_pag': float(c['interes_pag']),
                'seguro_pag': float(c['seguro_pag']),
                'fee_pag': float(c['fee_pag']),
                'total_pag': float(total_pag),
                'total_pend': float(total_pend),
                'estado': estado_final,
                'fecha_pago': c['fecha_pago'],
                'ref_bancaria': "; ".join(c['ref_bancaria']) if c['ref_bancaria'] else "N/A"
            })

        # === 4. Crear Excel ===
        wb = Workbook()
        ws = wb.active
        ws.title = f"Préstamo {prestamo_id}"

        # Encabezados
        headers = [
            "Tipo", "Cuota", "Fecha Vencimiento",
            "Capital Programado", "Interés Programado", "Seguro Programado", "Fee Programado", "Total Programado",
            "Capital Pagado", "Interés Pagado", "Seguro Pagado", "Fee Pagado", "Total Pagado", "Total Pendiente",
            "Estado", "Fecha Pago", "Ref. Bancaria"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E2F3")

        # Agregar datos
        for fila in datos_excel:
            if fila['tipo'] == 'desembolso':
                ws.append([
                    "Desembolso", "", fila['fecha_vencimiento'],
                    "", "", "", "", fila['total_prog'],
                    "", "", "", "", fila['total_pag'], 0,
                    fila['estado'], "", ""
                ])
            else:
                ws.append([
                    "Cuota", fila['cuota'], fila['fecha_vencimiento'],
                    fila['capital_prog'], fila['interes_prog'], fila['seguro_prog'], fila['fee_prog'], fila['total_prog'],
                    fila['capital_pag'], fila['interes_pag'], fila['seguro_pag'], fila['fee_pag'], fila['total_pag'], fila['total_pend'],
                    fila['estado'], fila['fecha_pago'], fila['ref_bancaria']
                ])

        # Formato de números
        number_format = '#,##0.00'
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=14):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format

        # === 5. Agregar resumen final ===
        ws.append([])  # fila vacía
        resumen_row = ws.max_row + 1
        ws.cell(row=resumen_row, column=1, value="SALDO TOTAL PENDIENTE:")
        ws.cell(row=resumen_row, column=2, value=float(total_pendiente_general))
        ws.cell(resumen_row, column=2).number_format = number_format
        ws.cell(resumen_row, column=1).font = Font(bold=True)
        ws.cell(resumen_row, column=2).font = Font(bold=True, color="FF0000")

        # Guardar archivo
        output_dir = getattr(settings, 'REPORTS_OUTPUT_DIR', '/tmp')
        os.makedirs(output_dir, exist_ok=True)
        filename = f"reporte_prestamo_{prestamo_id}.xlsx"
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)

        self.stdout.write(
            self.style.SUCCESS(f"✅ Reporte completo generado:")
        )
        self.stdout.write(f"   📁 Ruta: {filepath}")
        self.stdout.write(f"   💰 Saldo pendiente total: {total_pendiente_general:.2f}")