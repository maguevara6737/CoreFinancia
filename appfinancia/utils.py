# ======================================================
# 1. Futuro (debe ir primero)
# ======================================================
from __future__ import annotations

# ======================================================
# 2. Librerías estándar de Python
# ======================================================
import datetime
import io
import os
import re
from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP, getcontext
from typing import Literal, Union

# ======================================================
# 3. Librerías de terceros
# ======================================================
import pdfplumber
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ======================================================
# 4. Django Core
# ======================================================
from django.conf import settings
from django.core.cache import cache
from django.core.exceptions import ObjectDoesNotExist
from django.db import models, transaction
from django.db.models import Sum
from django.utils import timezone

# ======================================================
# 5. Modelos locales
# ======================================================
from .models import (
    Desembolsos,
    Detalle_Aplicacion_Pago,
    Fechas_Sistema,
    Historia_Prestamos,
    Prestamos,
)

# ======================================================
# 6. Configuración adicional (opcional pero recomendada al inicio)
# ======================================================
getcontext().prec = 28

CACHE_TIMEOUT = 3600  # 1 hora

# ======================================================
# 1. Imports de compatibilidad futura (DEBEN ir primero)
# ======================================================
from __future__ import annotations

# ======================================================
# 2. Librerías estándar de Python
# ======================================================
import io
import os
import re
from typing import Literal, Optional
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP, getcontext

# ======================================================
# 3. Librerías de terceros (Instaladas vía pip)
# ======================================================
import pdfplumber
from openpyxl import Workbook
from typing import Literal, Union
from openpyxl.cell.cell import MergedCell
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Alignment, Font, PatternFill


# ======================================================
# 4. Django Core
# ======================================================
from django.conf import settings
from django.db.models import Sum
from django.utils import timezone
from django.core.cache import cache
from django.db import models, transaction
from django.core.exceptions import ObjectDoesNotExist
from django.core.exceptions import ValidationError

# ======================================================
# 5. Constantes Globales
# ======================================================
CACHE_TIMEOUT = 3600  # 1 hora

# ======================================================
# 6. Importaciones Locales (Modelos)
# Se recomienda agruparlos para evitar redundancia y facilitar el mantenimiento
# ======================================================
from .models import (
    Clientes,
    Conceptos_Transacciones,
    Desembolsos,
    Detalle_Aplicacion_Pago,
    Fechas_Sistema,
    Historia_Prestamos,
    Prestamos
)

# ======================================================
# 7. Configuración opcional de precisión decimal si tu sistema lo requiere
# ======================================================
getcontext().prec = 28
ROUND = ROUND_HALF_UP






#Obtener las Políticas de crédito
def get_politicas():
    from .models import Politicas
    """
    Retorna el objeto único de Politicas.
    Usa caché para evitar múltiples consultas a la BD.
    Lanza excepción si no está configurado.
    """
    CACHE_KEY = 'politicas_globales'
    politicas = cache.get(CACHE_KEY)
    
    if politicas is None:
        try:
            politicas = Politicas.objects.get()
            # Cachear por 1 hora (3600 segundos)
            cache.set(CACHE_KEY, politicas, timeout=3600)
        except Politicas.DoesNotExist:
            raise RuntimeError("❌ Las políticas globales no han sido configuradas. "
                             "Por favor, configúrelas en el panel de administración.")
    
    return politicas
    
#CACHE_TIMEOUT = 3600  # 1 hora

def _increment_field(field_name: str) -> int:
    from .models import Numeradores
    """
    Incrementa atómicamente un campo y devuelve el nuevo valor.
    """
    with transaction.atomic():
        # Bloquear la fila única
        numerador = Numeradores.objects.select_for_update().first()
        if numerador is None:
            numerador, _ = Numeradores.objects.get_or_create()

        # Obtener y actualizar
        current_value = getattr(numerador, field_name)
        new_value = current_value + 1
        setattr(numerador, field_name, new_value)
        numerador.save(update_fields=[field_name])

        # Opcional: actualizar caché
        cache_key = f"numerador_{field_name}"
        cache.set(cache_key, new_value, timeout=CACHE_TIMEOUT)

        return new_value

#===========================   DESEMBOLSO BACKEND ==================================
# appfinancia/utils.py


# Ajuste de precisión decimal para cálculos financieros
getcontext().prec = 28
ROUND = ROUND_HALF_UP



def obtener_user_name_from_django_user(user_obj):
    """
    Recibe request.user y retorna su username o str(user_obj) si no tiene atributo.
    """
    try:
        return user_obj.username
    except Exception:
        return str(user_obj)



#---------------------- desembolso func



# utils.py
def create_prestamo(desembolso):
    from .models import Prestamos

    print(f"DEBUG: create_prestamo recibió desembolso = {desembolso}, tipo = {type(desembolso)}")
    if desembolso is None or not hasattr(desembolso, 'pk'):
        raise ValueError(f"Desembolso inválido: {desembolso}")
    
    # ✅ Desempaquetar la tupla: (objeto, created)
    prestamo_obj, created = Prestamos.objects.update_or_create(
        prestamo_id=desembolso,  # Esto es la clave única para buscar
        defaults={
            'cliente_id': desembolso.cliente_id,
            'asesor': desembolso.asesor_id, # Usar _id si el modelo espera el objeto
            'aseguradora': desembolso.aseguradora_id,
            'vendedor': desembolso.vendedor_id,
            'tipo_tasa': desembolso.tipo_tasa,
            'tasa_mes': desembolso.tasa_mes or 0,
            'tasa': desembolso.tasa or 0,
            'valor': desembolso.valor or 0,
            'valor_cuota_1': desembolso.valor_cuota_1 or 0,
            'valor_cuota_mensual': desembolso.valor_cuota_mensual or 0,
            'valor_seguro_mes': desembolso.valor_seguro_mes or 0,
            'tiene_fee': desembolso.tiene_fee or 'NO',
            'dia_cobro': desembolso.dia_cobro or 1,
            'plazo_en_meses': desembolso.plazo_en_meses or 0,
            'fecha_desembolso': desembolso.fecha_desembolso,
            'fecha_vencimiento': desembolso.fecha_vencimiento,
            'ofrece_cuota_inicial': desembolso.ofrece_cuota_inicial or 'NO',
            'valor_cuota_inicial': desembolso.valor_cuota_inicial or 0,
            'tiene_oneroso': desembolso.tiene_oneroso or 'NO',
            'entidad_onerosa': desembolso.entidad_onerosa,
            'suspender_causacion': 'NO',
            'revocatoria': 'NO',
            # 'estado': 'ACTIVO', # Podrías activarlo tras el migrate   <------- ACTIVAR ESTA LINEA
        }
    )
    print(f"DEBUG: Prestamo {'creado' if created else 'actualizado'} con PK={prestamo_obj.pk}")
    return prestamo_obj  # ← Devolver solo el objeto, no la tupla



def create_movimiento(desembolso):
    from .models import Desembolsos, Movimientos 
    """
    Creates a Movimiento record based on a Desembolso instance.
    """
    movimiento = Movimientos.objects.create(
        cliente_id=desembolso.cliente_id,
        asesor_id=desembolso.asesor_id,
        valor_movimiento=desembolso.valor,
        fecha_valor_mvto=desembolso.fecha_desembolso,
    )
    return movimiento
#----------------------------------------


def create_loan_payments(prestamo, desembolso, plan_pagos, user_name):
    """
    Guarda el plan de pagos generado por calculate_loan_schedule
    como registros en Historia_Prestamos, usando los conceptos adecuados.
    """
    from .models import Fechas_Sistema, Conceptos_Transacciones, Historia_Prestamos  # ✅ Corregido

    # Obtener el concepto "DES"
    concepto_des = Conceptos_Transacciones.objects.get(concepto_id="DES")
    fechas_sistema = Fechas_Sistema.load()
    # Registrar el desembolso como primer histórico
    Historia_Prestamos.objects.create(
        prestamo_id=prestamo,
        fecha_efectiva=desembolso.fecha_desembolso,
        fecha_proceso=fechas_sistema.fecha_proceso_actual,  # o timezone.now().date()
        ordinal_interno=1,
        numero_cuota=0,
        numero_operacion=0,
        concepto_id=concepto_des,
        fecha_vencimiento=desembolso.fecha_vencimiento,
        tasa=desembolso.tasa,
        monto_transaccion=desembolso.valor,
        estado="TRANSACCION ",  # ¡incluye el espacio como en tu ESTADO_CHOICES!
        usuario=user_name,
        # abono_capital, intrs_ctes, seguro, fee se dejan en 0 por defecto
    )
    concepto_capital = Conceptos_Transacciones.objects.get(concepto_id="PLANCAP")
    concepto_interes = Conceptos_Transacciones.objects.get(concepto_id="PLANINT")
    concepto_seguro = Conceptos_Transacciones.objects.get(concepto_id="PLANSEG")
    concepto_gastos = Conceptos_Transacciones.objects.get(concepto_id="PLANGTO")
    #concepto_capital = "PLANCAP"
    #concepto_interes = "PLANINT"
    #concepto_seguro =  "PLANSEG"
    #concepto_gastos =  "PLANGTO"

    created_count = 0

    for cuota in plan_pagos:
        numero_cuota = cuota['numero_cuota']
        fecha_vencimiento = cuota['fecha_vencimiento']

        # Crear registro para capital
        if cuota['capital'] > 0:
            Historia_Prestamos.objects.create(
                prestamo_id=prestamo,
                fecha_efectiva=None,
                fecha_proceso=fechas_sistema.fecha_proceso_actual,
                ordinal_interno=300+created_count,
                numero_operacion=created_count,
                concepto_id=concepto_capital,
                monto_transaccion=float(cuota['capital']),
                fecha_vencimiento=fecha_vencimiento,
                estado="PENDIENTE",
                numero_cuota=numero_cuota,
                usuario=user_name,
            )
            created_count += 1

        # Crear registro para intereses
        if cuota['intereses'] > 0:
            Historia_Prestamos.objects.create(
                prestamo_id=prestamo,
                fecha_efectiva=None,
                fecha_proceso=fechas_sistema.fecha_proceso_actual,
                ordinal_interno=300+created_count,
                numero_operacion=created_count,
                concepto_id=concepto_interes,
                monto_transaccion=float(cuota['intereses']),
                fecha_vencimiento=fecha_vencimiento,
                estado="PENDIENTE",
                numero_cuota=numero_cuota,
                usuario=user_name,
            )
            created_count += 1

        # Crear registro para seguro
        if cuota['seguro'] > 0:
            Historia_Prestamos.objects.create(
                prestamo_id=prestamo,
                fecha_efectiva=None,
                fecha_proceso=fechas_sistema.fecha_proceso_actual,
                ordinal_interno=300+created_count,
                numero_operacion=created_count,
                concepto_id=concepto_seguro,
                monto_transaccion=float(cuota['seguro']),
                fecha_vencimiento=fecha_vencimiento,
                estado="PENDIENTE",
                numero_cuota=numero_cuota,
                usuario=user_name,
            )
            created_count += 1

        # Crear registro para gastos
        if cuota['gastos'] > 0:
            Historia_Prestamos.objects.create(
                prestamo_id=prestamo,
                fecha_efectiva=None,
                fecha_proceso=fechas_sistema.fecha_proceso_actual,
                ordinal_interno=300+created_count,
                numero_operacion=created_count,
                concepto_id=concepto_gastos,
                monto_transaccion=float(cuota['gastos']),
                fecha_vencimiento=fecha_vencimiento,
                estado="PENDIENTE",
                numero_cuota=numero_cuota,
                usuario=user_name,
            )
            created_count += 1

    return created_count
	
#----------------------------------------------------------------------------


def calculate_loan_schedule(desembolso):
    """
    Calculates the loan payment schedule based on the Desembolso object.
    CASO 1: valor_cuota_inicial > 0 (Cuota 0 = Capital + Seguro, Int=0).
    CASO 2: valor_cuota_inicial = 0 (Cuota 1 = Capital + Seguro, Int=0, hoy).
    """
    from .models import Desembolsos  

    results = []
    valor_total = Decimal(str(desembolso.valor or 0))
    v_seguro = Decimal(str(desembolso.valor_seguro_mes or 0))
    
    # --- LÓGICA DE CASOS (Directriz 1, 2 y 3) ---
    v_inicial = Decimal(str(desembolso.valor_cuota_inicial or 0))

    if v_inicial > 0:
        # CASO 1: Cuota inicial (Cuota 0) se compone de Capital + Seguro
        # El capital que amortiza es: inicial - seguro
        cap_amortiza_hoy = (v_inicial - v_seguro).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        monto_a_financiar = (valor_total - cap_amortiza_hoy).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        results.append({
            'numero_cuota': 0,
            'capital': cap_amortiza_hoy,
            'intereses': Decimal('0.00'),
            'fecha_vencimiento': desembolso.fecha_desembolso,
            'seguro': v_seguro, # Caso 1: Cuota 0 incluye seguro
            'gastos': Decimal('0.00'),
            'saldo_capital': Decimal('0.00'), 'saldo_intereses': Decimal('0.00'),
            'saldo_seguro': Decimal('0.00'), 'saldo_gastos': Decimal('0.00'),
        })
        fecha_base = desembolso.fecha_desembolso
        offset_idx = 0 # La siguiente es cuota 1
    else:
        # CASO 2: Sin inicial, la Cuota 1 es hoy (Capital + Seguro)
        v_cuota_1 = Decimal(str(desembolso.valor_cuota_1 or 0))
        # El capital que amortiza es: cuota_1 - seguro
        cap_amortiza_hoy = (v_cuota_1 - v_seguro).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        monto_a_financiar = (valor_total - cap_amortiza_hoy).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        
        results.append({
            'numero_cuota': 1,
            'capital': cap_amortiza_hoy,
            'intereses': Decimal('0.00'),
            'fecha_vencimiento': desembolso.fecha_desembolso,
            'seguro': v_seguro, # Caso 2: Cuota 1 incluye seguro
            'gastos': Decimal('0.00'),
            'saldo_capital': Decimal('0.00'), 'saldo_intereses': Decimal('0.00'),
            'saldo_seguro': Decimal('0.00'), 'saldo_gastos': Decimal('0.00'),
        })
        fecha_base = desembolso.fecha_desembolso
        offset_idx = 1 # La siguiente es cuota 2

    # --- CÁLCULO DE ANUALIDAD (CUOTAS NIVELADAS RESTANTES) ---
    plazo_total = int(desembolso.plazo_en_meses or 0)
    tasa_pct = Decimal(str(desembolso.tasa or 0))
    tasa_mensual = (tasa_pct / Decimal('100')) * (Decimal('30') / Decimal('360'))
    
    # Restamos 1 cuota porque ya procesamos la cuota "hoy" (ya sea la 0 o la 1)
    cuotas_restantes = max(plazo_total - 1, 0)
    
    if cuotas_restantes > 0 and monto_a_financiar > 0:
        r = tasa_mensual
        n = cuotas_restantes
        
        if r == 0:
            pago_mensual = (monto_a_financiar / n).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        else:
            one_plus_r = (Decimal('1') + r)
            denom = (Decimal('1') - (one_plus_r ** (Decimal(-n))))
            pago_mensual = (monto_a_financiar * r / denom).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        saldo_vivo = monto_a_financiar
        for i in range(1, cuotas_restantes + 1):
            # i=1 + offset (0 o 1) + (1 si es Caso 1 para compensar la cuota 0)
            if v_inicial > 0:
                numero = i  # Si hubo cuota 0, i=1 es Cuota 1
            else:
                numero = i + 1 # Si empezó en cuota 1, i=1 es Cuota 2
            
            intereses = (saldo_vivo * r).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            capital = (pago_mensual - intereses).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

            # Ajuste final por redondeo
            if i == cuotas_restantes:
                capital = saldo_vivo
                intereses = max(pago_mensual - capital, Decimal('0.00'))

            results.append({
                'numero_cuota': numero,
                'capital': capital,
                'intereses': intereses,
                'fecha_vencimiento': fecha_base + relativedelta(months=i),
                'seguro': v_seguro,
                'gastos': Decimal('0.00'),
                'saldo_capital': Decimal('0.00'),
                'saldo_intereses': Decimal('0.00'),
                'saldo_seguro': Decimal('0.00'), 
                'saldo_gastos': Decimal('0.00'),
            })
            saldo_vivo -= capital

    return results
#-----------
#2025-11-25 6:52am traslado funcion pasar_a_desembolsado a utils.py
""" def pasar_a_desembolsado(self, request, queryset):
    updated = queryset.filter(estado='A_DESEMBOLSAR').update(estado='DESEMBOLSADO')
    self.message_user(request, f"{updated} desembolso(s) pasado(s) a DESEMBOLSADO.")
pasar_a_desembolsado.short_description = "Pasar a DESEMBOLSADO" 
 """

#----------------------------------------------------------------2025/11/26
class FechasSistemaHelper:  
    """
    Helper para acceder de forma segura y eficiente a la configuración global de fechas del sistema.
    Basado en el patrón Singleton (`Fechas_Sistema.load()`).
    """
    
    @staticmethod
    def get_fecha_proceso_actual():
        from .models import Fechas_Sistema
        return Fechas_Sistema.load().fecha_proceso_actual

    @staticmethod
    def get_fecha_proceso_anterior():
        from .models import Fechas_Sistema
        return Fechas_Sistema.load().fecha_proceso_anterior

    @staticmethod
    def get_fecha_proximo_proceso():
        from .models import Fechas_Sistema
        return Fechas_Sistema.load().fecha_proximo_proceso

    @staticmethod
    def get_estado_sistema():
        from .models import Fechas_Sistema
        return Fechas_Sistema.load().estado_sistema

    @staticmethod
    def get_modo_fecha_sistema():
        from .models import Fechas_Sistema
        return Fechas_Sistema.load().modo_fecha_sistema

    @staticmethod
    def sistema_esta_abierto():
        from .models import Fechas_Sistema
        return FechasSistemaHelper.get_estado_sistema() == 'ABIERTO'

    @staticmethod
    def fecha_proceso_es_hoy():
        from .models import Fechas_Sistema
        return FechasSistemaHelper.get_fecha_proceso_actual() == timezone.now().date()

    #Acceso grupal (para evitar múltiples llamadas a `load()`)
    @staticmethod
    def get_all():
        """
        Devuelve un dict con todos los valores actuales.
        Útil para optimizar cuando se necesitan varios campos.
        """
        
        from .models import Fechas_Sistema
        fs = Fechas_Sistema.load()
        return {
            'fecha_proceso_actual': fs.fecha_proceso_actual,
            'fecha_proceso_anterior': fs.fecha_proceso_anterior,
            'fecha_proximo_proceso': fs.fecha_proximo_proceso,
            'estado_sistema': fs.estado_sistema,
            'modo_fecha_sistema': fs.modo_fecha_sistema,
        }

#Fin Fechas del Sistema


#=========================================================================================
# InBox Pagos
#=========================================================================================


def InBox_Pagos(archivo_pagos, usuario, cabezal):
    #identificar automáticamente el formato del archvo
    formato=f_identificar_formato(cabezal.nombre_archivo_id)
    
    if formato == "1-FORMATO PSE":
        return inbox_formato_pse(
            xls_file=archivo_pagos,
            nombre_archivo_id=cabezal.nombre_archivo_id,
            user=usuario,
            cabezal=cabezal
        )
        return ok, msg

    elif formato == "2-FORMATO ESTANDAR":
        return False, "El Formato Estándar aún no está implementado."

    elif formato == "3-FORMATO EXTRACTO BANCOLOMBIA":
        return inbox_formato_bancolombia(
            pdf_file=archivo_pagos,
            nombre_archivo_id=cabezal.nombre_archivo_id,
            user=usuario,
            cabezal=cabezal
        )

    else:
        return False, "Formato desconocido."


#-----------------------------

def safe(txt, maxlen):
    """Evita errores de overflow truncando cadenas."""
    if txt is None:
        return ""
    # Aseguramos que la entrada sea string antes de truncar y limpiar
    return str(txt).strip()[:maxlen]

#================
#Limpiar el valor
#================

def clean_money(valor_str):
    if not valor_str:
        return None        
    #limpio = str(valor_str).replace('$', '').replace(',', '').strip()
    limpio=valor_str[valor_str.rfind(" "):].replace(" ", "").replace(",", "")
    
    try:
        valor_numerico = float(limpio)
        
        return valor_numerico       
    except ValueError as e:
        # Maneja la excepción si la cadena limpia todavía no es un número válido.
        # print(f"Error de conversión en clean_money: {e} para cadena limpia '{limpio}'")
        return None


#================
#Limpiar descripción
#================
def clean_descripcion(reg: str) -> str:
    """
    Extrae la subcadena desde el inicio (pos 0) hasta justo antes 
    del primer número (0-9) o el signo de resta (-).
    """
    reg=reg[11:70]
    
    # 1. Definir el patrón de búsqueda: [0-9-] busca cualquier dígito o un guion.
    #patron = r'[0-9-]'
    patron = r'\d{2}|-'

    # 2. Buscar la primera coincidencia en la cadena 'reg'.
    match = re.search(patron, reg)

    if match:
        # Si hay una coincidencia, match.start() devuelve el índice de ese primer carácter.
        indice_fin = match.start()
        
        # 3. Recortar la subcadena desde el inicio hasta ese índice.
        # Se usa .rstrip() para quitar cualquier espacio sobrante al final de la descripción.
        return reg[:indice_fin].rstrip()
    else:
        # Si no se encuentra ningún número o signo, se devuelve la cadena completa.
        return reg.rstrip()

#================
#Excluir registros
#================
#from typing import Literal
def f_estado_pago(tipo_mov: str) -> Literal['EXCLUIDO', 'RECIBIDO']:
    if tipo_mov and tipo_mov.upper() == 'EXCLUIDO':
        return 'EXCLUIDO'
    else:
        return 'RECIBIDO'

#===================
#Estado conciliación
#===================
        


def f_estado_conciliacion(p_clas_mov: str) -> Union[str, Literal['NO', 'SI']]:
    clasificacion = p_clas_mov.upper() if p_clas_mov else ""
    if clasificacion == 'ABONO_PSE' or clasificacion == 'PAGO_PSE':
        return 'NO'
    elif clasificacion == 'PAGO_BANCOL':
        return 'SI'
    # 3. Caso por defecto
    else:
        return ""

# ============================================================
# PROCESOS DE ESTADO
# ============================================================



#===============
# ANULAR ARCHIVO
# ==============
def f_anular_archivo(cabezal: InBox_PagosCabezal, usuario):
    from .models import InBox_PagosCabezal, InBox_PagosDetalle
    try:
        with transaction.atomic():

            # 1. Validar que el cabezal esté en estado que permita anulación
            if cabezal.estado_proceso_archivo != "RECIBIDO":
                return False, "Solo se pueden ANULAR archivos en estado RECIBIDO."

            # 2. Actualizar CABEZAL
            cabezal.estado_proceso_archivo = "ANULADO"
            cabezal.save()

            # 3. Actualizar DETALLE pero **solo los que estén RECIBIDO**
            registros_actualizados = InBox_PagosDetalle.objects.filter(
                nombre_archivo_id=cabezal.nombre_archivo_id,
                estado_pago="RECIBIDO"          # ← condición solicitada
            ).update(
                estado_pago="ANULADO"
            )

        return True, (
            f"Archivo {cabezal.nombre_archivo_id} anulado correctamente. "
            f"({registros_actualizados} registros cambiados a ANULADO)"
        )

    except Exception as e:
        return False, f"Error al anular archivo: {str(e)}"


# ================
# PROCESAR ARCHIVO
# ================
def f_procesar_archivo(cabezal: InBox_PagosCabezal, usuario):
    from .models import InBox_PagosCabezal, InBox_PagosDetalle
    try:
        with transaction.atomic():

            # 1. Validar que el cabezal esté en estado que permita anulación
            if cabezal.estado_proceso_archivo != "RECIBIDO":
                return False, "Solo se pueden PROCESAR archivos en estado RECIBIDO."

            # 2. Actualizar CABEZAL
            cabezal.estado_proceso_archivo = "A_PROCESAR"
            cabezal.save()

            # 3. Actualizar DETALLE pero **solo los que estén RECIBIDO**
            registros_actualizados = InBox_PagosDetalle.objects.filter(
                nombre_archivo_id=cabezal.nombre_archivo_id,
                estado_pago="RECIBIDO"          # ← condición solicitada
            ).update(
                estado_pago="A_PROCESAR"
            )

        return True, (
            f"Archivo {cabezal.nombre_archivo_id} enviados a PROCESAR correctamente. "
            f"({registros_actualizados} registros enviados a PROCESAR)"
        )

    except Exception as e:
        return False, f"Error al envio a proceso archivo: {str(e)}"

#===============================
#Identificar formato del archivo
#===============================
def f_identificar_formato(nombre_archivo: str) -> Literal['1-FORMATO PSE', '3-FORMATO EXTRACTO BANCOLOMBIA', 'Error: Formato no soportado']:
    _, extension = os.path.splitext(nombre_archivo)
    
    # Normalizar la extensión a minúsculas para una comparación robusta
    extension = extension.lower()

    # 2. Lógica de clasificación
    if extension == ".xlsx":
        return '1-FORMATO PSE'

    elif extension == ".pdf":
        return '3-FORMATO EXTRACTO BANCOLOMBIA'

    # 3. Caso por defecto (error)
    else:
        return f'Error: La extensión "{extension}" no es válida. Tipos soportados: .xlsx, .pdf'


# ============================================================
# Obtener la clase de movimiento. aplica para el Extracto Bancolombia
# ============================================================
from decimal import Decimal
from typing import Optional # Para clarificar que el valor puede ser un Decimal o None

def clase_movimiento(reg: str, valor: Optional[Decimal]) -> str:
    
    # 1.Si el valor no existe (es None) o es negativo.
    if valor is None or valor < 0:
        return "EXCLUIDO"
        # Se sale de la función inmediatamente (Early exit)
    
    # a) PAGO VIRTUAL PSE
    if "PAGO VIRTUAL PSE" in reg:
        return "ABONO_PSE"

    # b) ABONO INTERESES AHORROS
    elif "ABONO INTERESES AHORROS" in reg:
        return "EXCLUIDO"
        # Se sale de la función inmediatamente

    # c) IMPTO GOBIERNO 4X1000
    elif "IMPTO GOBIERNO 4X1000" in reg:
        return "EXCLUIDO"
        # Se sale de la función inmediatamente
    
    # Si pasa todas las comprobaciones anteriores, se asigna la clasificación por defecto.
    return "PAGO_BANCOL"
    
#Fin cargar archivos InBox----------------------------------------------------------------

#_______________________________________________________
def obtener_tasa_prestamo(prestamo_id):
    """
    Obtiene la tasa anual del préstamo (como Decimal).
    La tasa se almacena en el campo 'tasa' del modelo Prestamos.
    
    Parámetros:
        prestamo_id: ID del préstamo (coincide con el ID del desembolso)
    
    Retorna:
        Decimal: tasa anual en formato decimal (ej: 0.25 para 25%)
    """
    try:
        prestamo = Prestamos.objects.get(prestamo_id=prestamo_id)
        # Asegurarse de que sea Decimal y no float
        return Decimal(str(prestamo.tasa))
    except Prestamos.DoesNotExist:
        raise ValueError(f"Préstamo con ID {prestamo_id} no encontrado.")
# En cualquier parte de tu código
#tasa = obtener_tasa_prestamo(38)  # préstamo ID 38
#print(tasa)  # Ej: Decimal('0.2500') si la tasa es 25%
#------------------------------------------------------------------------

# utils.py
# utils.py - Actualiza SOLO las funciones de consulta
# appfinancia/utils.py

from datetime import date

#_____________________________________________
# utils.py
from decimal import Decimal, ROUND_HALF_UP
from datetime import timedelta, date
from django.db.models import Sum
from .models import Historia_Prestamos, Conceptos_Transacciones, Prestamos, Fechas_Sistema


def total_intereses_por_periodo(fecha_inicio, fecha_fin):
    """
    Calcula intereses causados y ajustes por periodo.
    - interes_causado: solo intereses normales.
    - ajuste_intrs_causacion: suma de ajustes (AJUINT) en el periodo.
    """
    #print(f"\n[total_intereses_por_periodo] INICIO | desde={fecha_inicio} hasta={fecha_fin}")
    # === Validaciones y ajuste de fecha_fin ===
    if fecha_inicio < date(2020, 12, 31):
        raise ValueError("La fecha de inicio no puede ser anterior al 2020-12-31.")
    
    fechas_sistema = Fechas_Sistema.objects.first()
    if not fechas_sistema:
        raise ValueError("No se encontró la configuración de fechas del sistema.")
    fecha_proceso_actual = fechas_sistema.fecha_proceso_actual

    if fecha_fin > fecha_proceso_actual:
        raise ValueError(f"La fecha de fin no puede ser posterior a la fecha de proceso del sistema ({fecha_proceso_actual}).")

    if fecha_fin == fecha_proceso_actual:
        fecha_fin = fecha_proceso_actual - timedelta(days=1)
        if fecha_fin < fecha_inicio:
            return {
                'total_intereses': 0.0,
                'total_ajustes': 0.0,
                'detalle_por_prestamo': {}
            }

    # === Configuración de conceptos ===
    concepto_des = Conceptos_Transacciones.objects.get(concepto_id="DES")
    concepto_causac = Conceptos_Transacciones.objects.get(concepto_id="CAUSAC")
    concepto_ajuint = Conceptos_Transacciones.objects.get(concepto_id="AJUINT")

    prestamos_activos = Prestamos.objects.filter(revocatoria='NO')
    total_intereses_general = Decimal('0.00')
    total_ajustes_general = Decimal('0.00')
    detalle_por_prestamo = {}

    for prestamo in prestamos_activos:
        desembolsos = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id=concepto_des,
            fecha_efectiva__isnull=False,
            fecha_efectiva__lte=fecha_fin
        )
        total_desembolsado = sum((d.monto_transaccion for d in desembolsos), Decimal('0.00'))
        if total_desembolsado <= 0:
            continue

        primer_desembolso = desembolsos.order_by('fecha_efectiva').first()
        if not primer_desembolso:
            continue
        fecha_inicio_prestamo = primer_desembolso.fecha_efectiva
        if fecha_inicio_prestamo > fecha_fin:
            continue

        # --- Eventos CAUSAC y AJUINT ---
        eventos_causac_ajuint = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id__in=[concepto_causac, concepto_ajuint],
            fecha_proceso__lte= fecha_fin
        ).order_by('fecha_proceso', 'ordinal_interno', 'id')

        eventos = []
        for ev in eventos_causac_ajuint:
            eventos.append({
                'fecha': ev.fecha_proceso,
                'tipo': 'CAUSAC' if ev.concepto_id == concepto_causac else 'AJUINT',
                'capital_aplicado': ev.capital_aplicado_periodo if ev.concepto_id == concepto_causac else Decimal('0.00'),
                'monto_transaccion': ev.monto_transaccion,
                'tasa': ev.tasa
            })
        eventos.sort(key=lambda x: x['fecha'])

        # --- Inicialización ---
        saldo = total_desembolsado
        tasa = prestamo.tasa
        periodos = []
        total_intereses_prestamo = Decimal('0.00')
        total_ajustes_prestamo = Decimal('0.00')

        # --- Eventos antes de fecha_inicio ---
        for ev in eventos:
            if ev['fecha'] < fecha_inicio:
                if ev['tipo'] == 'CAUSAC':
                    saldo -= ev['capital_aplicado']
                    tasa = ev['tasa']
            else:
                break

        # --- Procesar rango ---
        fecha_actual = max(fecha_inicio, fecha_inicio_prestamo)
        eventos_en_rango = [ev for ev in eventos if fecha_inicio <= ev['fecha'] <= fecha_fin]

        for ev in eventos_en_rango:
            fecha_evento = ev['fecha']

            # Periodo normal (interés diario)
            if fecha_actual <= fecha_evento - timedelta(days=1):
                periodo_fin = fecha_evento - timedelta(days=1)
                dias = (periodo_fin - fecha_actual).days + 1
                interes_diario = (saldo * tasa / Decimal('100')) / Decimal('360')
                interes = (interes_diario * dias).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                periodos.append({
                    'periodo_inicio': fecha_actual,
                    'periodo_fin': periodo_fin,
                    'dias': dias,
                    'saldo_inicial': float(saldo),
                    'tasa': float(tasa),
                    'interes_causado': float(interes),
                    'ajuste_intrs_causacion': 0.0,
                    'tipo_evento': 'CAUSACION'
                })
                total_intereses_prestamo += interes
                total_intereses_general += interes

            # Evento en la fecha exacta
            if ev['tipo'] == 'CAUSAC':
                saldo -= ev['capital_aplicado']
                tasa = ev['tasa']
            elif ev['tipo'] == 'AJUINT':
                # Registrar ajuste como periodo de 1 día
                periodos.append({
                    'periodo_inicio': fecha_evento,
                    'periodo_fin': fecha_evento,
                    'dias': 1,
                    'saldo_inicial': float(saldo),
                    'tasa': float(tasa),
                    'interes_causado': 0.0,
                    'ajuste_intrs_causacion': float(ev['monto_transaccion']),
                    'tipo_evento': 'AJUSTE'
                })
                total_ajustes_prestamo += ev['monto_transaccion']
                total_ajustes_general += ev['monto_transaccion']

            fecha_actual = fecha_evento + timedelta(days=1)

        # --- Periodo final ---
        if fecha_actual <= fecha_fin:
            dias = (fecha_fin - fecha_actual).days + 1
            interes_diario = (saldo * tasa / Decimal('100')) / Decimal('360')
            interes = (interes_diario * dias).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            periodos.append({
                'periodo_inicio': fecha_actual,
                'periodo_fin': fecha_fin,
                'dias': dias,
                'saldo_inicial': float(saldo),
                'tasa': float(tasa),
                'interes_causado': float(interes),
                'ajuste_intrs_causacion': 0.0,
                'tipo_evento': 'CAUSACION'})
            total_intereses_prestamo += interes
            total_intereses_general += interes

        if  periodos:
            detalle_por_prestamo[prestamo.prestamo_id] = periodos

    #print(f"[total_intereses_por_periodo] FIN | total_intereses={total_intereses_general}, total_ajustes={total_ajustes_general}")
    #print(f"[total_intereses_por_periodo] Detalle préstamos con datos: {list(detalle_por_prestamo.keys())}")
    return {
        'total_intereses': float(total_intereses_general),
        'total_ajustes': float(total_ajustes_general),
        'detalle_por_prestamo': detalle_por_prestamo
    }
#------------------------------------------------------------------------
# utils.py
# ------ util.py (función aplicar_pago ajustada) ------
#2025-12-21 incluyo manejo de excedentes  y orden de aplicacion
#2025-12-27 Alineado para calcular con funciones factorizadas
#2025_12_30 Con simulacion y aplicacion modulares
# utils.py
from django.utils import timezone
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict
from .models import (
    Pagos, Prestamos, Historia_Prestamos, Conceptos_Transacciones,
    Detalle_Aplicacion_Pago, Fechas_Sistema
)

# === CONFIGURACIÓN (FUTURO: cargar desde BD) ===
ORDEN_CONCEPTOS = ["PLANSEG", "PLANGTO", "INTMORA", "PLANINT", "PLANCAP"]
MODO_APLICACION = "HORIZONTAL"  # o "VERTICAL"


def aplicar_pago(pago_id, usuario_nombre=None, simular=False):
    """
    Aplica (o simula) un pago a un préstamo.
    
    Parámetros:
        pago_id: ID del pago a aplicar
        usuario_nombre: nombre del usuario (solo en modo real)
        simular: si True, devuelve el desglose SIN guardar cambios.
    
    Retorna:
        dict con resultados de la aplicación/simulación
        
    Nota: 
        - Esta función NO maneja transacciones (debe llamarse dentro de transaction.atomic() si es modo real).
        - No usa select_for_update() (la concurrencia la maneja la función llamadora).
    """
    # === OBTENER PAGO ===
    pago = Pagos.objects.get(pago_id=pago_id)
    if not simular:
        if pago.estado_pago == 'APLICADO':
            return {
                'status': 'success',
                'message': f'Pago {pago_id} ya fue aplicado.',
                'pago_id': pago_id
            }
        if pago.estado_pago != 'CONCILIADO':
            raise ValueError("Solo se pueden aplicar pagos en estado 'CONCILIADO'.")

    if pago.prestamo_id_real is None:
        raise ValueError("El pago no tiene un préstamo asignado.")
        
    # === OBTENER PRÉSTAMO Y VALIDAR ===
    prestamo = Prestamos.objects.get(prestamo_id__prestamo_id=pago.prestamo_id_real)

    #  Estado del Desembolso
    desembolso = prestamo.prestamo_id 
    if desembolso.estado != 'DESEMBOLSADO':
        raise ValueError(
            f"Prestamo no desembolsado, si es cuota inicial asigne este pago ID ({pago.pago_id}) al Desembolso"
        )


    if pago.cliente_id_real != prestamo.cliente_id_id:
        raise ValueError("El cliente del pago no coincide con el del préstamo.")

    fecha_aplicacion = pago.fecha_pago
    if fecha_aplicacion < prestamo.fecha_desembolso:
        raise ValueError(
        f"La fecha del pago ({fecha_aplicacion:%d/%m/%Y}) no puede ser anterior "
        f"a la fecha de desembolso ({prestamo.fecha_desembolso:%d/%m/%Y})."
        )


    if pago.cliente_id_real != prestamo.cliente_id_id:
        raise ValueError("El cliente del pago no coincide con el del préstamo.")


    fechas_sistema = Fechas_Sistema.load()
    if not fechas_sistema or not fechas_sistema.fecha_proceso_actual:
        raise ValueError("No hay fecha de sistema configurada.")
    if fecha_aplicacion > fechas_sistema.fecha_proceso_actual:
        raise ValueError("La fecha de aplicación no puede ser posterior a la fecha de sistema.")
        
    # === En aplicar_pago, dentro del bloque principal (después de validar, antes de ejecutar) ===
    if not simular:
        # Obtener número de asiento contable único para esta transacción
        try:
            from .utils import get_next_asientos_id
        except ImportError:
            from utils import get_next_asientos_id  # fallback

        numero_asiento = get_next_asientos_id()
    else:
        numero_asiento = None  # en simulación no se necesita

    # === OBTENER ADEUDO DETALLADO ===
    liquidacion = prestamo.liquidar_prestamo(fecha_liquidacion=fecha_aplicacion)
    monto_debido_base = Decimal(str(liquidacion['total_a_pagar'])).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # === APLICAR DESCUENTO POR AJUINT PENDIENTES ===
    ajuint_pendiente = obtener_ajuint_pendiente(prestamo)
    monto_debido = max(monto_debido_base - ajuint_pendiente, Decimal('0.00'))

    # Registrar para depuración
    print(f"DEBUG: Adeudo base: {monto_debido_base}, AJUINT pendiente: {ajuint_pendiente}, Adeudo ajustado: {monto_debido}")

    monto_pago = Decimal(str(pago.valor_pago)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    diferencia = monto_pago - monto_debido

    if monto_pago > monto_debido:
        monto_ajuste = diferencia.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        ajuste_tipo = 'EXCEDENTE'
    else:
        monto_ajuste = Decimal('0.00')
        ajuste_tipo = None

    # === OBTENER ÍTEMS COBRABLES (INCLUYE INTMORA) ===
    items_cobrables = obtener_items_cobrables(prestamo, fecha_aplicacion)

    # === ORDENAR SEGÚN MODO ===
    if MODO_APLICACION == "HORIZONTAL":
        items_ordenados = aplicar_modo_horizontal(items_cobrables, ORDEN_CONCEPTOS)
    else:
        items_ordenados = aplicar_modo_vertical(items_cobrables, ORDEN_CONCEPTOS)

    print(f"DEBUG: aplicar_pago 1 ejecutar_aplicacion_pago")

    # === EJECUTAR APLICACIÓN ===
    resultado = ejecutar_aplicacion_pago(
        items_ordenados,
        monto_pago,
        pago,
        prestamo,
        fecha_aplicacion,
        simular=simular,
        usuario_nombre=usuario_nombre,
        ajuste_tipo=ajuste_tipo,
        monto_ajuste=monto_ajuste,
        numero_asiento_contable=numero_asiento  
    )

    # === MARCAR AJUINT COMO APLICADOS SI EL PAGO LIQUIDA LA DEUDA ===
    if not simular and monto_debido <= monto_pago and ajuint_pendiente > Decimal('0.00'):
        from .models import Conceptos_Transacciones
        concepto_ajuint = Conceptos_Transacciones.objects.get(concepto_id="AJUINT")
        Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id=concepto_ajuint,
            estado="TRANSACCION"
        ).update(
            estado="PAGADA",
            fecha_efectiva=fecha_aplicacion,
            numero_asiento_contable=numero_asiento
        )
        print(f"DEBUG: {ajuint_pendiente} en AJUINT marcados como aplicados.")

    print(f"DEBUG: aplicar_pago 2 enriquecer")
    # === ENRIQUECER RESULTADO ===
    resultado.update({
        'fecha_aplicacion': fecha_aplicacion,
            'monto_debido': float(monto_debido),  # ← Ya incluye descuento por AJUINT
            'monto_debido_base': float(monto_debido_base),  # ← Adeudo original (sin descuento)
            'ajuint_pendiente': float(ajuint_pendiente),  # ← ¡Nuevo!
            'ajuint_aplicado': float(ajuint_pendiente) if (not simular and monto_debido <= monto_pago and ajuint_pendiente > 0) else 0.0,  # ← ¡Nuevo!
            'monto_pago': float(monto_pago),
            'diferencia': float(diferencia),
            'ajuste_tipo': ajuste_tipo,
            'monto_ajuste': float(monto_ajuste),
            'prestamo_id': prestamo.prestamo_id.prestamo_id, 
            'componentes': {
                'capital_pendiente': liquidacion['capital_pendiente'],
                'intereses_corrientes': liquidacion['intereses_corrientes'],
                'intereses_mora': liquidacion['intereses_mora'],
                'seguros_vencidos': liquidacion['seguros_vencidos'],
                'gastos_vencidos': liquidacion['gastos_vencidos'],
        }
    })
    print(f"DEBUG: aplicar_pago 6 en salgo de aplicar_pago -6-")
    return resultado


def obtener_items_cobrables(prestamo, fecha_aplicacion):
    """
    Obtiene todos los componentes cobrables de las cuotas vencidas,
    incluyendo el interés de mora como ítem virtual 'INTMORA'.
    
    Retorna:
        Lista de dict con:
        - 'concepto': str (ej. 'PLANCAP', 'PLANINT', 'INTMORA', etc.)
        - 'numero_cuota': int
        - 'monto': Decimal (monto pendiente a cobrar)
        - 'fecha_vencimiento': date
    """
    from decimal import Decimal
    cuotas_detalle = prestamo.detalle_cuotas_vencidas_con_mora(fecha_aplicacion)
    items = []

    for cuota in cuotas_detalle:
        if cuota['seguro'] > Decimal('0.00'):
            items.append({
                'concepto': 'PLANSEG',
                'numero_cuota': cuota['numero_cuota'],
                'monto': cuota['seguro'],
                'fecha_vencimiento': cuota['fecha_vencimiento']
            })
        if cuota['gastos'] > Decimal('0.00'):
            items.append({
                'concepto': 'PLANGTO',
                'numero_cuota': cuota['numero_cuota'],
                'monto': cuota['gastos'],
                'fecha_vencimiento': cuota['fecha_vencimiento']
            })
        if cuota['interes_programado'] > Decimal('0.00'):
            items.append({
                'concepto': 'PLANINT',
                'numero_cuota': cuota['numero_cuota'],
                'monto': cuota['interes_programado'],
                'fecha_vencimiento': cuota['fecha_vencimiento']
            })
        if cuota['capital'] > Decimal('0.00'):
            items.append({
                'concepto': 'PLANCAP',
                'numero_cuota': cuota['numero_cuota'],
                'monto': cuota['capital'],
                'fecha_vencimiento': cuota['fecha_vencimiento']
            })
        if cuota['interes_mora'] > Decimal('0.00'):
            items.append({
                'concepto': 'INTMORA',  # concepto virtual
                'numero_cuota': cuota['numero_cuota'],
                'monto': cuota['interes_mora'],
                'fecha_vencimiento': cuota['fecha_vencimiento']
            })
    return items


def aplicar_modo_horizontal(items_cobrables, orden_conceptos):
    """
    Ordena los ítems en modo HORIZONTAL:
    - Primero agrupa por número de cuota (ascendente: más vencida primero).
    - Dentro de cada cuota, ordena los conceptos según `orden_conceptos`.
    
    Retorna:
        Lista de ítems ordenados lista para aplicar pago secuencialmente.
    """
    from collections import defaultdict
    cuotas = defaultdict(list)
    for item in items_cobrables:
        cuotas[item['numero_cuota']].append(item)
    
    items_ordenados = []
    for cuota_num in sorted(cuotas.keys()):  # cuotas más antiguas primero
        # Ordenar componentes de la cuota según prioridad
        cuotas[cuota_num].sort(
            key=lambda x: orden_conceptos.index(x['concepto']) if x['concepto'] in orden_conceptos else 999
        )
        items_ordenados.extend(cuotas[cuota_num])
    return items_ordenados


def aplicar_modo_vertical(items_cobrables, orden_conceptos):
    """
    Ordena los ítems en modo VERTICAL:
    - Primero ordena por posición en `orden_conceptos`.
    - Luego, dentro de cada concepto, ordena por número de cuota (ascendente).
    
    Retorna:
        Lista de ítems ordenados lista para aplicar pago secuencialmente.
    """
    items_ordenados = sorted(
        items_cobrables,
        key=lambda x: (
            orden_conceptos.index(x['concepto']) if x['concepto'] in orden_conceptos else 999,
            x['numero_cuota']  # cuotas más antiguas primero dentro del concepto
        )
    )
    return items_ordenados

#________________________________________________________________________________________
# utils.py
def ejecutar_aplicacion_pago(
    items_ordenados, 
    monto_pago, 
    pago, 
    prestamo,
    fecha_aplicacion,
    simular=False,
    usuario_nombre=None,
    ajuste_tipo=None,
    monto_ajuste=Decimal('0.00'),
    numero_asiento_contable=None 
):
    """
    Aplica el monto del pago a la lista de ítems ordenados.
    Si simular=False, persiste los cambios en la base de datos.
    
    Retorna:
        dict con:
        - 'aplicaciones_realizadas': lista de aplicaciones (componente, monto, cuota, fecha)
        - 'cuotas_detalle': dict con desglose por cuota (para vista)
    """
    from decimal import Decimal
    from django.db import models
    from .models import Historia_Prestamos, Conceptos_Transacciones, Detalle_Aplicacion_Pago

    monto_restante = monto_pago
    aplicaciones_realizadas = []
    total_capital_aplicado = Decimal('0.00')  # ← Para cerrar_periodo_interes y AJUINT
    
    # Mapeo de concepto a nombre de componente para Detalle_Aplicacion_Pago
    CONCEPTO_A_COMPONENTE = {
        "PLANSEG": "SEGURO",
        "PLANGTO": "GASTOS",
        "PLANINT": "INTERES",
        "PLANCAP": "CAPITAL",
        "INTMORA": "INTERES_MORA"
    }
    
    # Para construir cuotas_detalle (útil en simulación y depuración)
    cuotas_detalle = defaultdict(lambda: {
        'numero_cuota': None,
        'fecha_vencimiento': None,
        'capital': Decimal('0.00'),
        'interes': Decimal('0.00'),
        'interes_mora': Decimal('0.00'),
        'seguro': Decimal('0.00'),
        'gastos': Decimal('0.00'),
    })
    
    # --- 0. OBTENER ÚLTIMO ordinal_interno PARA EL PRÉSTAMO (solo en modo real) ---
    ultimo_ordinal = 0
    if not simular:
        ultimo_ordinal = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo
        ).aggregate(models.Max('ordinal_interno'))['ordinal_interno__max'] or 0

    # --- 1. Aplicar pago a ítems ordenados ---
    for item in items_ordenados:
        if monto_restante <= 0:
            break
        aplicado = min(monto_restante, item['monto'])
        monto_restante -= aplicado
        
        # Registrar aplicación
        print(f"DEBUG: aplicar_pago 3 registrar aplic")
        componente = CONCEPTO_A_COMPONENTE.get(item['concepto'], item['concepto'])
        aplicaciones_realizadas.append({
            'componente': componente,
            'monto': float(aplicado),
            'numero_cuota': item['numero_cuota'],
            'fecha_vencimiento': item['fecha_vencimiento']
        })
        
        # Acumular en cuotas_detalle
        cuota = cuotas_detalle[item['numero_cuota']]
        cuota['numero_cuota'] = item['numero_cuota']
        cuota['fecha_vencimiento'] = item['fecha_vencimiento']
        if item['concepto'] == 'PLANCAP':
            cuota['capital'] += aplicado
            total_capital_aplicado += aplicado  # ← Acumular para cerrar_periodo_interes y AJUINT
        elif item['concepto'] == 'PLANINT':
            cuota['interes'] += aplicado
        elif item['concepto'] == 'INTMORA':
            cuota['interes_mora'] += aplicado
        elif item['concepto'] == 'PLANSEG':
            cuota['seguro'] += aplicado
        elif item['concepto'] == 'PLANGTO':
            cuota['gastos'] += aplicado
        
        # --- 2. Persistir en BD (solo si no es simulación) ---
        if not simular:
            if item['concepto'] == 'INTMORA':
                # Interés de mora: solo se registra en Detalle_Aplicacion_Pago (no hay Historia_Prestamos)
                # Buscar el registro de PLANCAP de la misma cuota para evitar not null
                concepto_capital = Conceptos_Transacciones.objects.get(concepto_id="PLANCAP")
                registro_capital = Historia_Prestamos.objects.get(
                    prestamo_id=prestamo,
                    concepto_id=concepto_capital,
                    numero_cuota=item['numero_cuota'],
                    estado__in=['PENDIENTE', 'PAGADA']  # por si ya se actualizó
                )

                # Incrementar ordinal para el registro de mora (aunque no se cree nuevo registro, se actualiza el existente)
                ultimo_ordinal += 1

                Detalle_Aplicacion_Pago.objects.create(
                    pago=pago,
                    historia_prestamo=registro_capital,  # ✅ referencia válida
                    monto_aplicado=aplicado,
                    componente=componente,
                    fecha_aplicacion=fecha_aplicacion
                )
            else:
                # Conceptos reales: actualizar Historia_Prestamos y crear Detalle_Aplicacion_Pago
                print(f"DEBUG: aplicar_pago 4 concep reales")
                concepto_trans = Conceptos_Transacciones.objects.get(concepto_id=item['concepto'])
                registro = Historia_Prestamos.objects.get(
                    prestamo_id=prestamo,
                    concepto_id=concepto_trans,
                    numero_cuota=item['numero_cuota'],
                    estado='PENDIENTE'
                )
                
                # Determinar campo a actualizar
                campo_pago = {
                    "PLANSEG": "seguro",
                    "PLANGTO": "fee",
                    "PLANINT": "intrs_ctes",
                    "PLANCAP": "abono_capital"
                }[item['concepto']]
                
                saldo_actual = getattr(registro, campo_pago, Decimal('0.00'))
                nuevo_saldo = saldo_actual + aplicado
                setattr(registro, campo_pago, nuevo_saldo)
                
                if nuevo_saldo >= registro.monto_transaccion:
                    registro.estado = "PAGADA"
                if registro.fecha_efectiva is None:
                    registro.fecha_efectiva = fecha_aplicacion
                
                fechas_sistema = Fechas_Sistema.load()
                registro.fecha_proceso = fechas_sistema.fecha_proceso_actual

                # Asignar campos comunes
                registro.numero_asiento_contable = numero_asiento_contable
                registro.numero_operacion = pago.pago_id  # ← ¡Importante!
                ultimo_ordinal += 1
                registro.ordinal_interno = ultimo_ordinal  # ← ¡Asignación segura!
                
                registro.save(update_fields=[
                    campo_pago, 
                    'estado', 
                    'fecha_efectiva', 
                    'fecha_proceso', 
                    'numero_asiento_contable',
                    'numero_operacion',
                    'ordinal_interno'
                ])
                
                Detalle_Aplicacion_Pago.objects.create(
                    pago=pago,
                    historia_prestamo=registro,
                    monto_aplicado=aplicado,
                    componente=componente,
                    fecha_aplicacion=fecha_aplicacion
                )
    
    # --- 3. Registrar ajuste por excedente (solo en modo real) ---
    if not simular and ajuste_tipo == "EXCEDENTE" and monto_ajuste > Decimal('0.00'):
        concepto_ajuste = Conceptos_Transacciones.objects.get(concepto_id="AJU_EXC")
        nombre_usuario = (usuario_nombre or 'SYSTEM')[:15]
        print(f"DEBUG: aplicar_pago excdente")
        
        ultimo_ordinal += 1  # ← Incrementar para el ajuste
        Historia_Prestamos.objects.create(
            prestamo_id=prestamo,
            concepto_id=concepto_ajuste,
            fecha_efectiva=fecha_aplicacion,
            fecha_proceso=fechas_sistema.fecha_proceso_actual,
            fecha_vencimiento=fecha_aplicacion,
            monto_transaccion=monto_ajuste,
            abono_capital=monto_ajuste,
            intrs_ctes=Decimal('0.00'),
            seguro=Decimal('0.00'),
            fee=Decimal('0.00'),
            estado="TRANSACCION",
            numero_cuota=0,
            usuario=nombre_usuario,
            numero_operacion=pago.pago_id,  # ← ¡Importante!
            ordinal_interno=ultimo_ordinal,  # ← ¡Corregido!
            numero_asiento_contable=numero_asiento_contable
        )
    
    # --- 4. Registrar AJUINT si es pago anticipado (solo en modo real) ---
    # Este bloque reconoce intereses a favor del deudor por pagos anticipados.
    # Se crea un registro con concepto "AJUINT" para ser compensado al finalizar el préstamo.
    if not simular and total_capital_aplicado > Decimal('0.00'):
        # Obtener la primera cuota pendiente de capital
        concepto_cap = Conceptos_Transacciones.objects.get(concepto_id="PLANCAP")
        cuotas_pendientes = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id=concepto_cap,
            estado="PENDIENTE"
        ).order_by('fecha_vencimiento')
        
        if cuotas_pendientes.exists():
            primera_vencida = cuotas_pendientes.first()
            if fecha_aplicacion < primera_vencida.fecha_vencimiento:
                dias_adelanto = (primera_vencida.fecha_vencimiento - fecha_aplicacion).days
                if dias_adelanto > 0:
                    tasa_diaria = prestamo.tasa / Decimal('100') / Decimal('360')
                    ahorro = (total_capital_aplicado * tasa_diaria * dias_adelanto).quantize(Decimal('0.01'))
                    if ahorro > Decimal('0.00'):
                        # Obtener siguiente ordinal único para este préstamo
                        ultimo_ordinal += 1
                        
                        concepto_ajuint = Conceptos_Transacciones.objects.get(concepto_id="AJUINT")
                        Historia_Prestamos.objects.create(
                            prestamo_id=prestamo,
                            concepto_id=concepto_ajuint,
                            fecha_efectiva=fecha_aplicacion,
                            fecha_proceso=fechas_sistema.fecha_proceso_actual,
                            fecha_vencimiento=fecha_aplicacion,
                            monto_transaccion=ahorro,
                            abono_capital=Decimal('0.00'),
                            intrs_ctes=Decimal('0.00'),
                            seguro=Decimal('0.00'),
                            fee=Decimal('0.00'),
                            estado="TRANSACCION",
                            numero_cuota=0,
                            usuario=usuario_nombre[:15] if usuario_nombre else 'SYSTEM',
                            numero_operacion=pago.pago_id,
                            ordinal_interno=ultimo_ordinal,  # ← ¡Ordinal seguro!
                            numero_asiento_contable=numero_asiento_contable  # ← ¡Mismo asiento del pago!
                        )
                        print(f"DEBUG: AJUINT registrado por adelanto de {dias_adelanto} días. Ahorro: {ahorro}")
    
    # --- 5. CERRAR PERÍODO DE INTERÉS (solo en modo real) ---
    if not simular:
        from .utils import cerrar_periodo_interes
        # Asumimos que prestamo.prestamo_id es el ID real del préstamo (de Desembolsos)
        prestamo_id_real = prestamo.prestamo_id.prestamo_id
        cerrar_periodo_interes(
            prestamo_id_real,
            fecha_aplicacion,
            pago_referencia=f"PAGO_{pago.pago_id}",
            numero_asiento_contable=numero_asiento_contable,
            capital_aplicado=float(total_capital_aplicado)
        )
    
    # --- 6. Actualizar estado del pago (solo en modo real) ---
    if not simular:
        from django.utils import timezone
        pago.estado_pago = 'APLICADO'
        pago.fecha_aplicacion_pago = timezone.now()
        pago.save(update_fields=['estado_pago', 'fecha_aplicacion_pago'])
    
    # Convertir cuotas_detalle a dict standard y a float
    cuotas_detalle_serializable = {}
    for k, v in cuotas_detalle.items():
        cuotas_detalle_serializable[k] = {
        'numero_cuota': v['numero_cuota'],
        'fecha_vencimiento': v['fecha_vencimiento'].isoformat() if v['fecha_vencimiento'] else None,
        'capital': float(v['capital']),
        'interes': float(v['interes']),
        'interes_mora': float(v['interes_mora']),
        'seguro': float(v['seguro']),
        'gastos': float(v['gastos']),
        }
    
    resultado = {
        'aplicaciones_realizadas': aplicaciones_realizadas,
        'cuotas_detalle': cuotas_detalle_serializable,
    }
    
    if not simular:
        resultado.update({
            'status': 'success',
            'message': f'Pago {pago.pago_id} aplicado exitosamente.',
            'numero_asiento_contable': numero_asiento_contable
        })
    
    return resultado 

#____________________________________________________________________________
# utils.py
def obtener_ajuint_pendiente(prestamo):
    """
    Retorna el total de ajustes por intereses a favor (AJUINT) pendientes para un préstamo.
    Solo considera registros con estado='TRANSACCION' (no aplicados aún).
    """
    from .models import Conceptos_Transacciones, Historia_Prestamos
    from django.db.models import Sum
    from decimal import Decimal

    try:
        concepto_ajuint = Conceptos_Transacciones.objects.get(concepto_id="AJUINT")
    except Conceptos_Transacciones.DoesNotExist:
        return Decimal('0.00')

    total_ajuint = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo,
        concepto_id=concepto_ajuint,
        estado="TRANSACCION"
    ).aggregate(total=Sum('monto_transaccion'))['total'] or Decimal('0.00')

    return total_ajuint

#___________________________________________________________________________

#----------------------------------------------------------------2025/11/26
# appfinancia/utils.py
#from .models import Fechas_Sistema
from django.utils import timezone
from datetime import date

def get_fecha_proceso_actual() -> date:
    """Devuelve la fecha de proceso actual (singleton seguro)"""
    from .models import Fechas_Sistema
    return Fechas_Sistema.load().fecha_proceso_actual

def get_fecha_proceso_anterior() -> date:
    from .models import Fechas_Sistema
    return Fechas_Sistema.load().fecha_proceso_anterior

def get_fecha_proximo_proceso() -> date:
    from .models import Fechas_Sistema
    return Fechas_Sistema.load().fecha_proximo_proceso

def get_estado_sistema() -> str:
    from .models import Fechas_Sistema
    return Fechas_Sistema.load().estado_sistema

def get_modo_fecha_sistema() -> str:
    from .models import Fechas_Sistema
    return Fechas_Sistema.load().modo_fecha_sistema

def sistema_esta_abierto() -> bool:
    #from .models import Fechas_Sistema
    return get_estado_sistema() == 'ABIERTO'

def fecha_proceso_es_hoy() -> bool:
    from .models import Fechas_Sistema
    return get_fecha_proceso_actual() == timezone.now().date()
    
if sistema_esta_abierto():
    hoy = get_fecha_proceso_actual()    
    

#------------ ***   para el Historial del prestamo a xls ******
# appfinancia/utils.py

# --- utils.py (AJUSTADO SIN MERGE) ---
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.cell import MergedCell 
from decimal import Decimal
from django.db.models import Sum # Necesario para la agregación
from .models import (
    Historia_Prestamos, Detalle_Aplicacion_Pago, Fechas_Sistema, 
    Desembolsos, Prestamos, Clientes
)
#___________________________________________________________________________
def aplicar_pago_cuota_inicial(desembolso, prestamo, usuario: str):
    """
    Aplica el pago de la cuota inicial (Cuota 0) o la Cuota 1 (si no hay inicial).
    Usa los campos numero_transaccion_cuota_1 y valor_cuota_1 del desembolso.
    Debe ejecutarse dentro de un bloque transaction.atomic() externo.
    """
    from decimal import Decimal, ROUND_HALF_UP
    from .models import Pagos, Fechas_Sistema, Conceptos_Transacciones, Historia_Prestamos, Detalle_Aplicacion_Pago
    from django.utils import timezone

    pago_id = desembolso.numero_transaccion_cuota_1  #PAGO_ID
 
    # 1: Obtener el pago desde Pagos
    try:
        pago = Pagos.objects.select_for_update().get(pago_id=pago_id)
    except Pagos.DoesNotExist:
        raise ValueError(f"El pago con ID {pago_id} no existe en la tabla Pagos.")

    valor_pago = Decimal(str(pago.valor_pago)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    

    # 2. Obtener fecha de proceso del sistema
    fechas_sistema = Fechas_Sistema.objects.first()
    if not fechas_sistema:
        raise ValueError("No se encontró ningún registro en Fechas_Sistema.")
    fecha_proceso_actual = fechas_sistema.fecha_proceso_actual

    monto_restante = valor_pago

    # 3. Conceptos
    concepto_cap = Conceptos_Transacciones.objects.get(concepto_id="PLANCAP")
    concepto_int = Conceptos_Transacciones.objects.get(concepto_id="PLANINT")
    concepto_seg = Conceptos_Transacciones.objects.get(concepto_id="PLANSEG")
    concepto_gto = Conceptos_Transacciones.objects.get(concepto_id="PLANGTO")
    concepto_excedente = Conceptos_Transacciones.objects.get(concepto_id="PAGOEXC")

    # 4 --- LÓGICA DE SELECCIÓN DE CUOTA (CASO 1 y 2) ---
    # Si valor_cuota_inicial > 0, pagamos la cuota 0. Si no, pagamos la cuota 1.
    num_cuota_objetivo = 0 if (desembolso.valor_cuota_inicial or 0) > 0 else 1

    # Filtrar registros PENDIENTES solo de la cuota objetivo
    registros_pendientes = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo,
        numero_cuota=num_cuota_objetivo,
        concepto_id__in=[concepto_cap, concepto_int, concepto_seg, concepto_gto],
        estado="PENDIENTE"
    )

    prioridad = {concepto_seg.concepto_id: 1, concepto_int.concepto_id: 2, concepto_cap.concepto_id: 3, concepto_gto.concepto_id: 4}
    registros_pendientes = sorted(
        registros_pendientes,
        key=lambda r: prioridad.get(r.concepto_id_id, 99)
    )

    print(f"DEBUG:1aplicar_cuota_inicial, cuota = {num_cuota_objetivo}, monto_restante = {monto_restante}, registros_pendientes = {registros_pendientes}")

    # 6. Aplicar pago a la cuota seleccionada
    for reg in registros_pendientes:
        if monto_restante <= 0:
            break

        if reg.concepto_id_id == concepto_cap.concepto_id:
            pagado = reg.abono_capital; campo_pago = 'abono_capital'; comp_nombre = 'CAPITAL'
        elif reg.concepto_id_id == concepto_int.concepto_id:
            pagado = reg.intrs_ctes; campo_pago = 'intrs_ctes'; comp_nombre = 'INTERES'
        elif reg.concepto_id_id == concepto_seg.concepto_id:
            pagado = reg.seguro; campo_pago = 'seguro'; comp_nombre = 'SEGURO'
        elif reg.concepto_id_id == concepto_gto.concepto_id:
            pagado = reg.fee; campo_pago = 'fee'; comp_nombre = 'GASTOS'
        else:
            continue

        print(f"DEBUG:2 ------------->, reg.concepto_id_id = {reg.concepto_id_id}, pagado = {pagado}, comp_nombre = {comp_nombre}")

        saldo_pendiente = reg.monto_transaccion - pagado
        print(f"DEBUG:3aplicar_cuota_inicial, saldo_pendiente = {saldo_pendiente}, , reg.monto_transaccion: { reg.monto_transaccion} -pagado = {pagado}")
        if saldo_pendiente <= 0:
            continue

        aplicado = min(monto_restante, saldo_pendiente)
        monto_restante -= aplicado

        print(f"DEBUG:4aplicar_cuota_inicial,  monto_restante= {monto_restante}, saldo_pendiente = {saldo_pendiente}, aplicado = {aplicado}")

        setattr(reg, campo_pago, getattr(reg, campo_pago) + aplicado)

        print(f"DEBUG:5aplicar_cuota_inicial,  campo_pago= {campo_pago}, reg = {reg} ")

        if getattr(reg, campo_pago) >= reg.monto_transaccion:
            reg.estado = "PAGADA"
        
        if reg.fecha_efectiva is None:
            reg.fecha_efectiva = pago.fecha_pago
        
        reg.fecha_proceso = fecha_proceso_actual
        reg.usuario = usuario
        reg.numero_operacion = pago.pago_id
        reg.save()

        Detalle_Aplicacion_Pago.objects.create(
            pago=pago,
            historia_prestamo=reg,
            monto_aplicado=aplicado,
            componente=comp_nombre
        )

    # 7. Excedente (Si el pago fue mayor a la cuota 0/1)
    if monto_restante > 0:
        hist_excedente = Historia_Prestamos.objects.create(
            prestamo_id=prestamo,
            concepto_id=concepto_excedente,
            fecha_efectiva=pago.fecha_pago,
            fecha_proceso=fecha_proceso_actual,
            fecha_vencimiento=pago.fecha_pago,
            monto_transaccion=monto_restante,
            estado="TRANSACCION",
            numero_cuota=num_cuota_objetivo,
            usuario=usuario,
            numero_operacion = pago.pago_id
        )
        Detalle_Aplicacion_Pago.objects.create(
            pago=pago,
            historia_prestamo=hist_excedente,
            monto_aplicado=monto_restante,
            componente='EXCEDENTE'
        ) 

    # 8. Finalizar estado del pago
    pago.estado_pago = 'aplicado'
    pago.fecha_aplicacion_pago = timezone.now()
    pago.save()

    return {
        'status': 'success',
        'message': f'Pago de cuota {num_cuota_objetivo} aplicado exitosamente.',
        'pago_id': pago_id,
        'monto_aplicado': float(pago.valor_pago - monto_restante),
        'excedente': float(monto_restante),
    }
#_________________________________________________________________________

# utils.py

import os
from django.conf import settings
from decimal import Decimal

def generar_comprobante_pago_en_disco_uso_email_futuro(pago, resultado, cuotas_detalle, output_dir=None):
    """
    Genera comprobante de pago en Excel y PDF.
    Si falla la generación de alguno, lo registra pero no lanza excepción.
    """
    if output_dir is None:
        output_dir = os.path.join(settings.MEDIA_ROOT, 'comprobantes_pagos')
    os.makedirs(output_dir, exist_ok=True)

    filename_base = f"comprobante_pago_{pago.pago_id}_{pago.cliente_id_real}"
    excel_path = None
    pdf_path = None

    # === 1. Generar Excel ===
    try:
        import pandas as pd
        rows = []
        for num_cuota, data in cuotas_detalle.items():
            rows.append({
                'Cuota': num_cuota,
                'Vencimiento': data['fecha_vencimiento'],
                'Capital': float(data['capital']),
                'Interés': float(data['interes']),
                'Interés Mora': float(data['interes_mora']),
                'Seguro': float(data['seguro']),
                'Gastos': float(data['gastos']),
            })
        if resultado.get('monto_ajuste', 0) > 0:
            rows.append({
                'Cuota': 'Excedente',
                'Vencimiento': '',
                'Capital': 0,
                'Interés': 0,
                'Interés Mora': 0,
                'Seguro': 0,
                'Gastos': float(resultado['monto_ajuste']),
            })

        df = pd.DataFrame(rows)
        excel_path = os.path.join(output_dir, f"{filename_base}.xlsx")
        df.to_excel(excel_path, index=False, engine='openpyxl')
    except Exception as e:
        print(f"⚠️ Error generando Excel para pago {pago.pago_id}: {e}")
        excel_path = None

    # === 2. Generar PDF ===
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        pdf_path = os.path.join(output_dir, f"{filename_base}.pdf")
        doc = SimpleDocTemplate(pdf_path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("COMPROBANTE DE APLICACIÓN DE PAGO", styles['Title']))
        story.append(Spacer(1, 12))
        story.append(Paragraph(f"Pago ID: {pago.pago_id}", styles['Normal']))
        story.append(Paragraph(f"Cliente ID: {pago.cliente_id_real}", styles['Normal']))
        story.append(Paragraph(f"Monto: ${pago.valor_pago:,.2f}", styles['Normal']))
        story.append(Paragraph(f"Fecha: {pago.fecha_pago}", styles['Normal']))
        story.append(Spacer(1, 12))

        # Tabla
        data = [['Cuota', 'Vencimiento', 'Capital', 'Interés', 'Mora', 'Seguro', 'Gastos']]
        for row in rows:
            data.append([
                str(row['Cuota']),
                str(row['Vencimiento']) if row['Vencimiento'] else '',
                f"${row['Capital']:,.2f}",
                f"${row['Interés']:,.2f}",
                f"${row['Interés Mora']:,.2f}",
                f"${row['Seguro']:,.2f}",
                f"${row['Gastos']:,.2f}",
            ])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
        doc.build(story)
    except ImportError:
        print(f"⚠️ reportlab no está disponible. PDF no generado para pago {pago.pago_id}.")
        pdf_path = None
    except Exception as e:
        print(f"⚠️ Error generando PDF para pago {pago.pago_id}: {e}")
        pdf_path = None

    return pdf_path, excel_path
#------------------------------------------------------------------
from io import BytesIO
from decimal import Decimal

# utils.py
def _construir_filas_comprobante(cuotas_detalle, resultado):
    """Construye filas para el comprobante (incluye excedente como fila única si aplica)."""
    rows = []
    
    # 1. Cuotas normales
    for num_cuota, data in cuotas_detalle.items():
        if num_cuota == 'Excedente':
            continue  # Ya lo manejamos al final
        rows.append({
            'Cuota': num_cuota,
            'Vencimiento': data['fecha_vencimiento'],
            'Capital': float(data['capital']),
            'Interés': float(data['interes']),
            'Interés Mora': float(data['interes_mora']),
            'Seguro': float(data['seguro']),
            'Gastos': float(data['gastos']),
        })
    
    # 2. Excedente (solo si aplica y es > 0)
    if resultado.get('ajuste_tipo') == 'EXCEDENTE' and resultado.get('monto_ajuste', 0) > 0:
        rows.append({
            'Cuota': 'Excedente a capital',
            'Vencimiento': '',
            'Capital': 0.0,
            'Interés': 0.0,
            'Interés Mora': 0.0,
            'Seguro': 0.0,
            'Gastos': float(resultado['monto_ajuste']),
        })

    return rows
#_____________________________________________________________________________

# utils.py
import io
import pandas as pd
from django.template.loader import render_to_string
from django.utils import timezone
from django.conf import settings

def generar_comprobante_pago_en_memoria(pago, resultado, cuotas_detalle, cliente_nombre=None, prestamo_id=None):
    """
    Genera PDF (basado en HTML) y Excel (Sincronizado) del comprobante en memoria.
    """
    pdf_bytes = None
    excel_bytes = None

    # --- Función auxiliar para asegurar números limpios en Excel ---
    def limpiar_monto(valor):
        try:
            return float(valor or 0)
        except (TypeError, ValueError):
            return 0.0

    # ==========================================
    # === A) Generar EXCEL (Sincronizado) ===
    # ==========================================
    try:
        # 1. Mapeo de conceptos igual al desglose del HTML
        conceptos = [
            ('Gastos', limpiar_monto(resultado.get('gastos_vencidos'))),
            ('Seguros', limpiar_monto(resultado.get('seguros_vencidos'))),
            ('Intereses de Mora', limpiar_monto(resultado.get('intereses_mora'))),
            ('Intereses Corrientes', limpiar_monto(resultado.get('intereses_corrientes'))),
            ('Abono a Capital (Cuota)', limpiar_monto(resultado.get('capital_aplicado'))),
            ('Abono Extraordinario', limpiar_monto(resultado.get('monto_ajuste_total'))),
        ]

        data_rows = []
        for concepto, valor in conceptos:
            if valor > 0:
                data_rows.append({'Concepto': concepto, 'Valor Aplicado': valor})

        # 2. Resumen de la obligación (Post-pago)
        resumen = resultado.get('resumen_obligacion', {})
        if resumen:
            data_rows.append({'Concepto': '', 'Valor Aplicado': None}) # Separador
            data_rows.append({'Concepto': '--- RESUMEN DE OBLIGACIÓN ---', 'Valor Aplicado': None})
            data_rows.append({'Concepto': 'Saldo Anterior', 'Valor Aplicado': limpiar_monto(resumen.get('saldo_anterior'))})
            data_rows.append({'Concepto': '(-) Capital Pagado', 'Valor Aplicado': limpiar_monto(resumen.get('capital_pagado'))})
            data_rows.append({'Concepto': '(=) Nuevo Saldo Capital', 'Valor Aplicado': limpiar_monto(resumen.get('nuevo_saldo_capital'))})
            data_rows.append({'Concepto': 'Cuotas Pendientes', 'Valor Aplicado': resumen.get('cuotas_pendientes')})

        df_final = pd.DataFrame(data_rows)

        output_excel = io.BytesIO()
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            # Encabezado informativo
            df_info = pd.DataFrame([
                ['Financia Seguro S.A.S.'],
                [f'Comprobante: {pago.pago_id}'],
                [f'Cliente: {cliente_nombre or "—"}'],
                [f'Préstamo: {prestamo_id or "—"}'],
                [f'Fecha Pago: {pago.fecha_pago}'],
                ['']
            ])
            df_info.to_excel(writer, index=False, header=False, sheet_name='Comprobante')
            df_final.to_excel(writer, index=False, sheet_name='Comprobante', startrow=6)
            
            # Ajuste de formato numérico (0.00) en la columna B
            worksheet = writer.sheets['Comprobante']
            for row in worksheet.iter_rows(min_row=8, max_col=2):
                if isinstance(row[1].value, (int, float)):
                    row[1].number_format = '0.00'

        excel_bytes = output_excel.getvalue()

    except Exception as e:
        print(f"⚠️ Error generando Excel sincronizado para pago {pago.pago_id}: {e}")
        excel_bytes = None

    # ==========================================
    # === B) Generar PDF (Idéntico al HTML) ===
    # ==========================================
    try:
        from weasyprint import HTML
        from .utils import calcular_totales_aplicados # Ajustar import si es necesario
        
        # Preparamos el contexto exactamente igual al del template
        # Reutilizamos los datos que ya calculaste para el HTML
        context = {
            'pago': pago,
            'cliente': {'nombre': cliente_nombre} if cliente_nombre else None,
            'resultado': resultado,
            'totales_aplicados': {
                'gastos_vencidos': resultado.get('gastos_vencidos', 0),
                'seguros_vencidos': resultado.get('seguros_vencidos', 0),
                'intereses_mora': resultado.get('intereses_mora', 0),
                'intereses_corrientes': resultado.get('intereses_corrientes', 0),
                'capital_aplicado': resultado.get('capital_aplicado', 0),
                'monto_ajuste_total': resultado.get('monto_ajuste_total', 0),
            },
            'resumen_obligacion': resultado.get('resumen_obligacion'),
            'saldos_especiales': resultado.get('saldos_especiales'),
            'fecha_operacion': timezone.now(),
            'es_previsualizacion': False,
        }

        # Renderizamos el HTML a string
        html_string = render_to_string('appfinancia/comprobante_unificado.html', context)
        
        # Convertimos a PDF
        buffer_pdf = io.BytesIO()
        HTML(string=html_string).write_pdf(buffer_pdf)
        pdf_bytes = buffer_pdf.getvalue()

    except Exception as e:
        print(f"⚠️ Error generando PDF vía HTML para pago {pago.pago_id}: {e}")
        pdf_bytes = None

    return pdf_bytes, excel_bytes
#___________________________________________________________________________________
# utils.py
def calcular_resumen_obligacion_obsoleta(prestamo, pago, fecha_aplicacion):
    """Calcula el resumen de la obligación tras el pago."""
    from decimal import Decimal
    from .models import Detalle_Aplicacion_Pago, Conceptos_Transacciones
    
    saldo_anterior = prestamo.saldo_capital_pendiente()
    
    capital_pagado = Detalle_Aplicacion_Pago.objects.filter(
        pago=pago,
        componente='CAPITAL'
    ).aggregate(total=models.Sum('monto_aplicado'))['total'] or Decimal('0.00')
    
    nuevo_saldo_capital = max(saldo_anterior - capital_pagado, Decimal('0.00'))
    
    try:
        concepto_cap = Conceptos_Transacciones.objects.get(concepto_id="PLANCAP")
        cuotas_pendientes = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id=concepto_cap,
            estado="PENDIENTE"
        ).count()
        
        proximo_vencimiento = None
        if cuotas_pendientes > 0:
            proxima_cuota = Historia_Prestamos.objects.filter(
                prestamo_id=prestamo,
                concepto_id=concepto_cap,
                estado="PENDIENTE"
            ).order_by('fecha_vencimiento').first()
            if proxima_cuota:
                proximo_vencimiento = proxima_cuota.fecha_vencimiento
    except:
        cuotas_pendientes = 0
        proximo_vencimiento = None
    
    return {
        'saldo_anterior': float(saldo_anterior),
        'capital_pagado': float(capital_pagado),
        'nuevo_saldo_capital': float(nuevo_saldo_capital),
        'cuotas_pendientes': cuotas_pendientes,
        'proximo_vencimiento': proximo_vencimiento.isoformat() if proximo_vencimiento else None
    }

def calcular_saldos_especiales(prestamo, pago, aplicaciones_realizadas, cuotas_detalle, monto_ajuste=0):
    """Calcula los saldos especiales."""
    cuotas_atrasadas, dias_mora = prestamo.cuotas_atrasadas_info(fecha_corte=pago.fecha_pago)
    estado_mora = f"En mora por {dias_mora} días" if cuotas_atrasadas > 0 else "Al día"
    
    return {
        'saldo_parcial_cuota': 0.0,  # Puedes implementar lógica específica si es necesario
        'saldo_favor_cliente': float(monto_ajuste),
        'estado_mora': estado_mora
    }
#___________________________________________________________________________________________
# utils.py
def calcular_desglose_pago_obsoletea(prestamo, pago, fecha_aplicacion, resultado_aplicacion):
    """
    Calcula el desglose correcto del pago para el comprobante unificado.
    """
    from decimal import Decimal
    from .models import Detalle_Aplicacion_Pago, Conceptos_Transacciones, Historia_Prestamos
    
    # === 1. Capital de cuotas vencidas (correcto para Abono a Capital) ===
    capital_cuotas_vencidas = prestamo.capital_vencido_no_pagado(fecha_corte=fecha_aplicacion)
    
    # === 2. Determinar qué mostrar en "Abono a Capital" ===
    saldo_total = prestamo.saldo_capital_pendiente()
    monto_pago = Decimal(str(pago.valor_pago))
    
    # Verificar si el pago cubre todo el adeudo
    adeudo_total = Decimal(str(resultado_aplicacion.get('monto_debido', 0)))
    pago_cubre_todo = monto_pago >= adeudo_total
    
    if pago_cubre_todo:
        abono_capital = float(saldo_total)
    else:
        abono_capital = float(capital_cuotas_vencidas)
    
    # === 3. Capital realmente aplicado en esta transacción ===
    capital_aplicado = Detalle_Aplicacion_Pago.objects.filter(
        pago=pago,
        componente='CAPITAL'
    ).aggregate(total=models.Sum('monto_aplicado'))['total'] or Decimal('0.00')
    
    # === 4. Nuevo saldo de capital ===
    nuevo_saldo_capital = max(saldo_total - capital_aplicado, Decimal('0.00'))
    
    # === 5. Cuotas pendientes después del pago ===
    try:
        concepto_cap = Conceptos_Transacciones.objects.get(concepto_id="PLANCAP")
        # Cuotas que siguen pendientes después del pago
        cuotas_pendientes = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id=concepto_cap,
            estado="PENDIENTE"
        ).count()
    except:
        cuotas_pendientes = 0
    
    # === 6. Próximo vencimiento ===
    proximo_vencimiento = None
    if cuotas_pendientes > 0:
        try:
            proxima_cuota = Historia_Prestamos.objects.filter(
                prestamo_id=prestamo,
                concepto_id=concepto_cap,
                estado="PENDIENTE"
            ).order_by('fecha_vencimiento').first()
            if proxima_cuota:
                proximo_vencimiento = proxima_cuota.fecha_vencimiento
        except:
            proximo_vencimiento = None
    else:
        # Si no hay cuotas pendientes, verificar si hay cuotas en mora
        cuotas_mora = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id=concepto_cap,
            estado="PENDIENTE",
            fecha_vencimiento__lt=fecha_aplicacion
        ).order_by('fecha_vencimiento').first()
        if cuotas_mora:
            proximo_vencimiento = cuotas_mora.fecha_vencimiento
    
    # === 7. Monto en mora ===
    cuotas_atrasadas, dias_mora = prestamo.cuotas_atrasadas_info(fecha_corte=fecha_aplicacion)
    monto_en_mora = 0.0
    if cuotas_atrasadas > 0:
        # Calcular monto total pendiente de cuotas vencidas
        monto_en_mora = float(prestamo.monto_atrasado(fecha_corte=fecha_aplicacion))
    
    return {
        'abono_capital': abono_capital,
        'capital_aplicado': float(capital_aplicado),
        'nuevo_saldo_capital': float(nuevo_saldo_capital),
        'cuotas_pendientes': cuotas_pendientes,
        'proximo_vencimiento': proximo_vencimiento.isoformat() if proximo_vencimiento else None,
        'monto_en_mora': monto_en_mora,
    }
#____________________________________________________________________________________________________
# utils.py
def formato_numero(valor):
    """
    Formatea un número con coma para miles y punto para decimales.
    Ej: 1250.50 → "1,250.50"
    """
    if valor is None:
        return "0.00"
    try:
        from django.contrib.humanize.templatetags.humanize import intcomma
        from decimal import Decimal
        valor_decimal = Decimal(str(valor))
        return intcomma(f"{valor_decimal:.2f}")
    except:
        return f"{valor:.2f}"
#______________________________________________________________________________________________________
# utils.py
def calcular_totales_aplicados(cuotas_detalle, monto_ajuste=0, ajuint_aplicado=0):
    """
    Calcula los totales aplicados por concepto desde cuotas_detalle.
    Esto es lo que realmente se muestra en la simulación.
    """
    from decimal import Decimal
    
    totales = {
        'gastos': Decimal('0.00'),
        'seguros': Decimal('0.00'),
        'intereses_mora': Decimal('0.00'),
        'intereses_corrientes': Decimal('0.00'),
        'capital': Decimal('0.00'),
        'excedente': Decimal(str(monto_ajuste)),
        'ajuint': Decimal(str(ajuint_aplicado))
    }
    
    for cuota_num, detalle in cuotas_detalle.items():
        if cuota_num == 'Excedente':
            continue  # Ya manejado en excedente
            
        totales['capital'] += Decimal(str(detalle.get('capital', 0)))
        totales['intereses_corrientes'] += Decimal(str(detalle.get('interes', 0)))
        totales['intereses_mora'] += Decimal(str(detalle.get('interes_mora', 0)))
        totales['seguros'] += Decimal(str(detalle.get('seguro', 0)))
        totales['gastos'] += Decimal(str(detalle.get('gastos', 0)))
    
    # Sumar AJUINT al excedente para el display
    totales['excedente'] += totales['ajuint']
    
    # Convertir a float para el template
    return {
        'gastos_vencidos': float(totales['gastos']),
        'seguros_vencidos': float(totales['seguros']),
        'intereses_mora': float(totales['intereses_mora']),
        'intereses_corrientes': float(totales['intereses_corrientes']),
        'capital_aplicado': float(totales['capital']),
        'monto_ajuste_total': float(totales['excedente'])
    }
#____________________________________________________________________________________
# utils.py
def calcular_resumen_correcto(prestamo, pago, fecha_aplicacion, cuotas_detalle):
    """
    Calcula el resumen correcto DESPUÉS de aplicar el pago simulado.
    """
    from decimal import Decimal
    from .models import Historia_Prestamos, Conceptos_Transacciones
    
    # (1) Saldo Anterior (antes del pago actual)
    saldo_anterior = prestamo.saldo_capital_pendiente()
    
    # (2) Capital Pagado en esta Transacción
    capital_aplicado = Decimal('0.00')
    for cuota_num, detalle in cuotas_detalle.items():
        if cuota_num != 'Excedente':
            capital_aplicado += Decimal(str(detalle.get('capital', 0)))
    
    # (3) Nuevo Saldo de Capital
    nuevo_saldo_capital = max(saldo_anterior - capital_aplicado, Decimal('0.00'))
    
    # (4) Cuotas Pendientes (después del pago simulado)
    try:
        concepto_cap = Conceptos_Transacciones.objects.get(concepto_id="PLANCAP")
        # Contar cuotas que siguen pendientes (incluso parcialmente pagadas)
        cuotas_pendientes = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id=concepto_cap,
            estado="PENDIENTE"
        ).count()
    except:
        cuotas_pendientes = 0
    
    # (5) Próximo Vencimiento
    proximo_vencimiento = None
    if cuotas_pendientes > 0:
        try:
            # Primero buscar cuotas VENCIDAS (más antiguas)
            cuota_vencida = Historia_Prestamos.objects.filter(
                prestamo_id=prestamo,
                concepto_id=concepto_cap,
                estado="PENDIENTE",
                fecha_vencimiento__lt=fecha_aplicacion
            ).order_by('fecha_vencimiento').first()
            
            if cuota_vencida:
                proximo_vencimiento = cuota_vencida.fecha_vencimiento
            else:
                # Si no hay vencidas, tomar la próxima futura
                cuota_futura = Historia_Prestamos.objects.filter(
                    prestamo_id=prestamo,
                    concepto_id=concepto_cap,
                    estado="PENDIENTE",
                    fecha_vencimiento__gte=fecha_aplicacion
                ).order_by('fecha_vencimiento').first()
                
                if cuota_futura:
                    proximo_vencimiento = cuota_futura.fecha_vencimiento
        except Exception as e:
            print(f"Error calculando próximo vencimiento: {e}")
            proximo_vencimiento = None
    
    # (7) Monto en Mora (cuotas vencidas NO cubiertas)
    monto_en_mora = Decimal('0.00')
    cuotas_vencidas_pendientes = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo,
        concepto_id__concepto_id__in=["PLANCAP", "PLANINT", "PLANSEG", "PLANGTO"],
        estado="PENDIENTE",
        fecha_vencimiento__lt=fecha_aplicacion
    )
    
    for registro in cuotas_vencidas_pendientes:
        # Calcular monto pendiente real
        if registro.concepto_id.concepto_id == "PLANCAP":
            pendiente = registro.monto_transaccion - registro.abono_capital
        elif registro.concepto_id.concepto_id == "PLANINT":
            pendiente = registro.monto_transaccion - registro.intrs_ctes
        elif registro.concepto_id.concepto_id == "PLANSEG":
            pendiente = registro.monto_transaccion - registro.seguro
        elif registro.concepto_id.concepto_id == "PLANGTO":
            pendiente = registro.monto_transaccion - registro.fee
        else:
            pendiente = registro.monto_transaccion
            
        if pendiente > 0:
            monto_en_mora += pendiente
    
    return {
        'saldo_anterior': float(saldo_anterior),
        'capital_pagado': float(capital_aplicado),
        'nuevo_saldo_capital': float(nuevo_saldo_capital),
        'cuotas_pendientes': cuotas_pendientes,
        'proximo_vencimiento': proximo_vencimiento.isoformat() if proximo_vencimiento else None,
        'monto_en_mora': float(monto_en_mora),
    }
#____________________________________________________________________________________________
# utils.py
def calcular_resumen_simulacion(prestamo, cuotas_detalle, fecha_aplicacion, saldo_anterior):
    """
    Calcula el resumen para SIMULACIÓN usando SOLO datos de la simulación.
    NO consulta la base de datos (porque no hay cambios reales).
    """
    from decimal import Decimal
    from .models import Historia_Prestamos, Conceptos_Transacciones
    
    # === 1. Saldo anterior (ya calculado antes de la simulación) ===
    saldo_anterior = Decimal(str(saldo_anterior))
    
    # === 2. Capital aplicado (de la simulación) ===
    capital_aplicado = Decimal('0.00')
    for cuota_num, detalle in cuotas_detalle.items():
        if cuota_num != 'Excedente':
            capital_aplicado += Decimal(str(detalle.get('capital', 0)))
    
    # === 3. Nuevo saldo después del pago simulado ===
    nuevo_saldo_capital = max(saldo_anterior - capital_aplicado, Decimal('0.00'))
    
    # === 4. Cuotas pendientes DESPUÉS de la simulación ===
    # Simulamos qué cuotas quedarían con saldo pendiente
    try:
        concepto_cap = Conceptos_Transacciones.objects.get(concepto_id="PLANCAP")
        cuotas_originales = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id=concepto_cap,
            estado="PENDIENTE"
        ).order_by('numero_cuota')
        
        cuotas_pendientes = 0
        for cuota_orig in cuotas_originales:
            # Buscar aplicación simulada para esta cuota
            monto_aplicado_sim = Decimal('0.00')
            for cuota_num_sim, detalle_sim in cuotas_detalle.items():
                if cuota_num_sim == cuota_orig.numero_cuota and cuota_num_sim != 'Excedente':
                    monto_aplicado_sim += Decimal(str(detalle_sim.get('capital', 0)))
            
            # Calcular saldo pendiente después de la simulación
            saldo_pendiente = cuota_orig.monto_transaccion - cuota_orig.abono_capital - monto_aplicado_sim
            if saldo_pendiente > 0:
                cuotas_pendientes += 1
                
    except Exception as e:
        print(f"Error en simulación de cuotas pendientes: {e}")
        cuotas_pendientes = 0
    
    # === 5. Próximo vencimiento DESPUÉS de la simulación ===
    proximo_vencimiento = None
    if cuotas_pendientes > 0:
        try:
            cuotas_vencidas_sim = []  # Cuotas vencidas que siguen pendientes
            cuotas_futuras_sim = []   # Cuotas futuras que siguen pendientes
            
            for cuota_orig in cuotas_originales:
                monto_aplicado_sim = Decimal('0.00')
                for cuota_num_sim, detalle_sim in cuotas_detalle.items():
                    if cuota_num_sim == cuota_orig.numero_cuota and cuota_num_sim != 'Excedente':
                        monto_aplicado_sim += Decimal(str(detalle_sim.get('capital', 0)))
                
                saldo_pendiente = cuota_orig.monto_transaccion - cuota_orig.abono_capital - monto_aplicado_sim
                
                if saldo_pendiente > 0:
                    if cuota_orig.fecha_vencimiento < fecha_aplicacion:
                        cuotas_vencidas_sim.append(cuota_orig)
                    else:
                        cuotas_futuras_sim.append(cuota_orig)
            
            # Prioridad: cuotas vencidas más antiguas
            if cuotas_vencidas_sim:
                cuotas_vencidas_sim.sort(key=lambda x: x.fecha_vencimiento)
                proximo_vencimiento = cuotas_vencidas_sim[0].fecha_vencimiento
            elif cuotas_futuras_sim:
                cuotas_futuras_sim.sort(key=lambda x: x.fecha_vencimiento)
                proximo_vencimiento = cuotas_futuras_sim[0].fecha_vencimiento
                
        except Exception as e:
            print(f"Error en simulación de próximo vencimiento: {e}")
            proximo_vencimiento = None
    
    # === 6. Monto en mora DESPUÉS de la simulación ===
    monto_en_mora = Decimal('0.00')
    try:
        conceptos_mora = ["PLANCAP", "PLANINT", "PLANSEG", "PLANGTO"]
        registros_vencidos = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id__concepto_id__in=conceptos_mora,
            estado="PENDIENTE",
            fecha_vencimiento__lt=fecha_aplicacion
        )
        
        for registro in registros_vencidos:
            # Buscar aplicación simulada para este registro
            monto_aplicado_sim = Decimal('0.00')
            for cuota_num_sim, detalle_sim in cuotas_detalle.items():
                if cuota_num_sim == registro.numero_cuota and cuota_num_sim != 'Excedente':
                    if registro.concepto_id.concepto_id == "PLANCAP" and 'capital' in detalle_sim:
                        monto_aplicado_sim += Decimal(str(detalle_sim['capital']))
                    elif registro.concepto_id.concepto_id == "PLANINT" and 'interes' in detalle_sim:
                        monto_aplicado_sim += Decimal(str(detalle_sim['interes']))
                    elif registro.concepto_id.concepto_id == "PLANSEG" and 'seguro' in detalle_sim:
                        monto_aplicado_sim += Decimal(str(detalle_sim['seguro']))
                    elif registro.concepto_id.concepto_id == "PLANGTO" and 'gastos' in detalle_sim:
                        monto_aplicado_sim += Decimal(str(detalle_sim['gastos']))
            
            # Calcular pendiente después de la simulación
            if registro.concepto_id.concepto_id == "PLANCAP":
                pendiente = registro.monto_transaccion - registro.abono_capital - monto_aplicado_sim
            elif registro.concepto_id.concepto_id == "PLANINT":
                pendiente = registro.monto_transaccion - registro.intrs_ctes - monto_aplicado_sim
            elif registro.concepto_id.concepto_id == "PLANSEG":
                pendiente = registro.monto_transaccion - registro.seguro - monto_aplicado_sim
            elif registro.concepto_id.concepto_id == "PLANGTO":
                pendiente = registro.monto_transaccion - registro.fee - monto_aplicado_sim
            else:
                pendiente = registro.monto_transaccion - monto_aplicado_sim
                
            if pendiente > 0:
                monto_en_mora += pendiente
                
    except Exception as e:
        print(f"Error en simulación de monto en mora: {e}")
        monto_en_mora = Decimal('0.00')
    
    return {
        'saldo_anterior': float(saldo_anterior),
        'capital_pagado': float(capital_aplicado),
        'nuevo_saldo_capital': float(nuevo_saldo_capital),
        'cuotas_pendientes': cuotas_pendientes,
        'proximo_vencimiento': proximo_vencimiento.isoformat() if proximo_vencimiento else None,
        'monto_en_mora': float(monto_en_mora),
    }
#___________________________________________________________________________________________________
# utils.py
def calcular_resumen_real(prestamo, pago, fecha_aplicacion):
    """
    Calcula el resumen para el COMPROBANTE REAL usando datos reales de la BD.
    Se ejecuta DENTRO de una transacción, así que los datos ya están actualizados.
    """
    from decimal import Decimal
    from .models import Detalle_Aplicacion_Pago, Historia_Prestamos, Conceptos_Transacciones
    
    # === 1. Saldo anterior (antes del pago actual) ===
    saldo_anterior = prestamo.saldo_capital_pendiente()
    
    # === 2. Capital aplicado (desde Detalle_Aplicacion_Pago real) ===
    capital_aplicado = Detalle_Aplicacion_Pago.objects.filter(
        pago=pago,
        componente='CAPITAL'
    ).aggregate(total=models.Sum('monto_aplicado'))['total'] or Decimal('0.00')
    
    # === 3. Nuevo saldo (después del pago real) ===
    nuevo_saldo_capital = max(saldo_anterior - capital_aplicado, Decimal('0.00'))
    
    # === 4. Cuotas pendientes reales ===
    try:
        concepto_cap = Conceptos_Transacciones.objects.get(concepto_id="PLANCAP")
        cuotas_pendientes = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            concepto_id=concepto_cap,
            estado="PENDIENTE"
        ).count()
        
        # === 5. Próximo vencimiento real ===
        proximo_vencimiento = None
        if cuotas_pendientes > 0:
            # Primero buscar cuotas vencidas
            cuota_vencida = Historia_Prestamos.objects.filter(
                prestamo_id=prestamo,
                concepto_id=concepto_cap,
                estado="PENDIENTE",
                fecha_vencimiento__lt=fecha_aplicacion
            ).order_by('fecha_vencimiento').first()
            
            if cuota_vencida:
                proximo_vencimiento = cuota_vencida.fecha_vencimiento
            else:
                # Si no hay vencidas, buscar la próxima futura
                cuota_futura = Historia_Prestamos.objects.filter(
                    prestamo_id=prestamo,
                    concepto_id=concepto_cap,
                    estado="PENDIENTE",
                    fecha_vencimiento__gte=fecha_aplicacion
                ).order_by('fecha_vencimiento').first()
                if cuota_futura:
                    proximo_vencimiento = cuota_futura.fecha_vencimiento
    except Exception as e:
        print(f"Error en cálculo real: {e}")
        cuotas_pendientes = 0
        proximo_vencimiento = None
    
    # === 6. Monto en mora real ===
    monto_en_mora = prestamo.monto_atrasado(fecha_corte=fecha_aplicacion)
    
    return {
        'saldo_anterior': float(saldo_anterior),
        'capital_pagado': float(capital_aplicado),
        'nuevo_saldo_capital': float(nuevo_saldo_capital),
        'cuotas_pendientes': cuotas_pendientes,
        'proximo_vencimiento': proximo_vencimiento.isoformat() if proximo_vencimiento else None,
        'monto_en_mora': float(monto_en_mora),
    }
#____________________________________________________________________________________________

from django.core.exceptions import ValidationError

def validar_nit(value):
    """
    Valida un NIT numérico donde el último dígito es el DV (Dígito de Verificación).
    Ejemplo: 8600000001 (donde 1 es el DV)
    """
    nit_str = str(value)
    
    if len(nit_str) < 6:
        raise ValidationError("El NIT es demasiado corto.")

    # El último dígito es el DV, el resto es el cuerpo del NIT
    cuerpo_nit = nit_str[:-1]
    dv_ingresado = int(nit_str[-1])
    
    # Pesos oficiales de la DIAN
    pesos = [71, 67, 59, 53, 47, 43, 41, 37, 29, 23, 19, 17, 13, 7, 3]
    
    # Rellenar con ceros a la izquierda hasta 15 dígitos para aplicar los pesos
    nit_completo = cuerpo_nit.zfill(15)
    
    suma = 0
    for i in range(15):
        suma += int(nit_completo[i]) * pesos[i]
    
    residuo = suma % 11
    if residuo > 1:
        dv_calculado = 11 - residuo
    else:
        dv_calculado = residuo
        
    if dv_ingresado != dv_calculado:
        raise ValidationError(
            f"NIT inválido. El dígito de verificación calculado para {cuerpo_nit} es {dv_calculado}, "
            f"por lo que el número completo debería ser {cuerpo_nit}{dv_calculado}."
        )
#_________________________________________________________________________________
def revertir_desembolso(desembolso_id, usuario, motivo):
    from django.db import transaction
    from django.utils import timezone
    from django.forms.models import model_to_dict # Útil para convertir registros a diccionarios
    import json
    from django.core.serializers.json import DjangoJSONEncoder
    from .models import (
        Desembolsos, Prestamos, Historia_Prestamos, 
        Pagos, Detalle_Aplicacion_Pago, BitacoraReversiones, Bitacora
    )

    with transaction.atomic():
        # 1. Obtención y bloqueo de registros principales
        desembolso = Desembolsos.objects.select_for_update().get(pk=desembolso_id)
        prestamo = Prestamos.objects.filter(prestamo_id=desembolso.prestamo_id).first()
        
        if not prestamo:
            raise ValueError("No se encontró un préstamo asociado.")

        # 2. VALIDACIÓN DE SEGURIDAD (Pagos posteriores)
        pagos_extra = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo,
            estado__icontains='PAGADA'
        ).exclude(numero_cuota__in=[0, 1]).count()

        if pagos_extra > 0:
            raise ValueError(f"Acción bloqueada: Existen {pagos_extra} pagos de cuotas posteriores.")

        # 3. CAPTURA TOTAL DE DATOS (El "Snapshot" de transparencia)
        # Convertimos los modelos a diccionarios para que sean serializables en JSON
        historia_qs = Historia_Prestamos.objects.filter(prestamo_id=prestamo)
        pago_qs = Pagos.objects.filter(pago_id=desembolso.numero_transaccion_cuota_1)
        detalles_qs = Detalle_Aplicacion_Pago.objects.filter(historia_prestamo__in=historia_qs)

        dump_datos = {
            "tabla_desembolso": model_to_dict(desembolso),
            "tabla_prestamo": model_to_dict(prestamo) if prestamo else {},
            "tabla_pagos": [model_to_dict(p) for p in pago_qs],
            "tabla_historia_prestamos": [model_to_dict(h) for h in historia_qs],
            "tabla_detalle_aplicacion": [model_to_dict(d) for d in detalles_qs],
        }

        # 4. PROCESO DE REVERSIÓN FÍSICA
        # A. Liberar Pagos y borrar sus aplicaciones
        if desembolso.numero_transaccion_cuota_1:
            try:
                pago_obj = pago_qs.select_for_update().get()
                detalles_qs.delete() # Borramos los vínculos de aplicación
                pago_obj.estado_pago = 'CONCILIADO'
                pago_obj.save()
            except Pagos.DoesNotExist:
                pass

        # B. Limpiar Historia (Borrado físico de cuotas y excedentes)
        historia_qs.delete()

        # C. Resetear Préstamo y Desembolso
        if prestamo:
            #prestamo.estado = 'REVERTIDO'   <----- ACTIVAR ESTA LINEA CUANDO HAGA MIGRATE 2026-01-11
            prestamo.valor = 0
            prestamo.valor_cuota_mensual = 0
            prestamo.save()

        desembolso.estado = 'ELABORACION'
        desembolso.save()

        # 5. REGISTRO EN BITÁCORAS
        bit_rev = BitacoraReversiones.objects.create(
            usuario=usuario,
            proceso="REVERSION_DESEMBOLSO",
            objeto_id=str(desembolso_id),
            motivo=motivo,
            dump_datos_anteriores=json.dumps(dump_datos, cls=DjangoJSONEncoder) 
        )

        Bitacora.objects.create(
            fecha_proceso=timezone.now().date(),
            user_name=usuario,
            evento_realizado='REVERSION_DESEMBOLSO_INDIV',
            proceso='EXITOSO',
            resultado=f"Reversión completa. Snapshot guardado en BitacoraReversiones ID: {bit_rev.id}"
        )

    return True
#_________________________________________________________________________________
def revertir_aplicacion_pago(pago_id, usuario, motivo):
    """
    Revierte la aplicación de un pago específico. 
    Directrices implementadas:
    1. Mora solo se borra de Detalle_Aplicacion.
    2. Cuotas afectadas regresan a PENDIENTE con fecha_efectiva=None.
    3. Pago queda en estado REVERSADO.
    4. Se borran registros CAUSAC, AJU_EXC y AJUINT de ese pago.
    5. Limpieza inteligente de numero_pago_referencia (lista de pagos).
    """
    import json
    from django.db import transaction
    from django.utils import timezone
    from django.core.serializers.json import DjangoJSONEncoder
    from django.forms.models import model_to_dict
    from .models import (
        Pagos, Prestamos, Historia_Prestamos, Detalle_Aplicacion_Pago, 
        BitacoraReversiones, Bitacora, ComprobantePago
    )
    with transaction.atomic():
        # 1. Obtención y bloqueo del pago
        pago = Pagos.objects.select_for_update().get(pk=pago_id)
        
        # 2. Identificación del Asiento Contable
        asiento_id = None
        comprobante = ComprobantePago.objects.filter(pago=pago).first()
        if comprobante:
            asiento_id = comprobante.datos_json.get('resultado', {}).get('numero_asiento_contable')
        
        if not asiento_id:
            primer_detalle = Detalle_Aplicacion_Pago.objects.filter(pago=pago).first()
            if primer_detalle:
                asiento_id = primer_detalle.historia_prestamo.numero_asiento_contable

        # --- VALIDACIÓN CRONOLÓGICA ---
        pago_posterior = Pagos.objects.filter(
            prestamo_id_real=pago.prestamo_id_real,
            fecha_pago__gt=pago.fecha_pago,
            estado_pago='APLICADO'
        ).exists()

        if pago_posterior:
            raise ValueError("Existen pagos posteriores aplicados. Revierta en orden inverso (del más reciente al más antiguo).")

        # --- CAPTURA DE SNAPSHOT (Antes de modificar nada) ---
        historia_total_asiento = Historia_Prestamos.objects.filter(numero_asiento_contable=asiento_id)
        detalles_qs = Detalle_Aplicacion_Pago.objects.filter(pago=pago)
        
        # Snapshot de todas las tablas involucradas
        dump_datos = {
            "pago": model_to_dict(pago),
            "asiento_ancla": asiento_id,
            "detalles_aplicacion": [model_to_dict(d) for d in detalles_qs],
            "historia_afectada": [model_to_dict(h) for h in historia_total_asiento],
            "comprobante": model_to_dict(comprobante) if comprobante else None,
        }

        # --- PROCESO DE REVERSIÓN FÍSICA ---

        # 1. Limpieza de Historia (Cuotas GT 0)
        pago_tag = f"PAGO_{pago_id}"
        for reg in historia_total_asiento.filter(numero_cuota__gt=0):
            # Limpiar referencia en la cadena
            if reg.numero_pago_referencia:
                refs = [r.strip() for r in reg.numero_pago_referencia.split(',') if r.strip()]
                if pago_tag in refs:
                    refs.remove(pago_tag)
                reg.numero_pago_referencia = ", ".join(refs)
            
            reg.estado = 'PENDIENTE'
            reg.fecha_efectiva = None
            reg.numero_asiento_contable = 0
            reg.save()

        # 2. Reversión de Saldos mediante detalles
        for det in detalles_qs:
            hist = det.historia_prestamo
            if det.componente == 'CAPITAL': hist.abono_capital -= det.monto_aplicado
            elif det.componente == 'INTERES': hist.intrs_ctes -= det.monto_aplicado
            elif det.componente == 'SEGURO': hist.seguro -= det.monto_aplicado
            elif det.componente == 'GASTOS': hist.fee -= det.monto_aplicado
            hist.save()

        # 3. ELIMINAR registros nacidos con este asiento (Causaciones o Ajustes)
        historia_total_asiento.filter(concepto_id__concepto_id__in=['CAUSAC', 'AJU_EXC', 'AJUINT']).delete()

        # 4. FINALIZAR ESTADOS Y LIMPIEZA
        pago.estado_pago = 'CONCILIADO' # Vuelve a estar disponible para aplicar
        pago.save()
        
        detalles_qs.delete()
        if comprobante:
            comprobante.delete()

        # 5. REGISTRO EN BITÁCORA DE REVERSIONES (JSON con codificador Django)
        bit_rev = BitacoraReversiones.objects.create(
            usuario=usuario,
            proceso="REVERSION_PAGO",
            objeto_id=str(pago_id),
            motivo=motivo,
            dump_datos_anteriores=json.dumps(dump_datos, cls=DjangoJSONEncoder)
        )

        # 6. REGISTRO EN BITÁCORA GENERAL
        Bitacora.objects.create(
            fecha_proceso=timezone.now().date(),
            user_name=usuario,
            evento_realizado='REVERSION_PAGO_INDIV',
            proceso='EXITOSO',
            resultado=f"Pago {pago_id} revertido (Asiento: {asiento_id}). Snapshot guardado en BitacoraReversiones ID: {bit_rev.id}"
        )

    return True

#_______________________________________________________________________________________________________
def calcular_dv_modulo11(numero):
    numero_str = str(numero)
    # Pesos estándar 2 a 7
    pesos = [2, 3, 4, 5, 6, 7]
    suma = 0
    for i, digito in enumerate(reversed(numero_str)):
        suma += int(digito) * pesos[i % len(pesos)]
    
    resto = suma % 11
    dv = 11 - resto
    if dv >= 10: return 0
    return dv

# ___________________________________________________

from .models import Prestamos   
def obtener_prestamo_unico(cliente_id):
    """
    Busca un préstamo único para un cliente y determina si es cuota inicial
    """
    # Filtramos todos los préstamos que pertenecen a ese cliente
    prestamos = Prestamos.objects.filter(cliente_id=cliente_id)

    # Contamos cuántos resultados existen
    if prestamos.count() == 1:
        prestamo_obj = prestamos.first()
        
        # Accedemos al desembolso relacionado para verificar el estado
        # Asumiendo que la relación en tu modelo es prestamo_obj.prestamo_id
        desembolso = prestamo_obj.prestamo_id 
        
        # (1) Lógica de cuota inicial
        if desembolso and desembolso.estado != 'DESEMBOLSADO':
            es_cuota_inicial = 'SI'
        else:
            es_cuota_inicial = 'NO'
            
        # Retornamos el ID del objeto y el indicador de cuota inicial
        return prestamo_obj.id, es_cuota_inicial

    # En cualquier otro caso (0 o más de 1), retornamos None para ambos
    return None, None


#-----------------------------------------------------------------------------------------
#Funciones públicas:

def get_next_asientos_id() -> int:
    """Obtiene el siguiente ID para asientos contables."""
    return _increment_field('numerador_operacion')

def get_next_transaccion_id() -> int:
    """Obtiene el siguiente ID para transacciones."""
    return _increment_field('numerador_transaccion')

def get_next_prestamo_id() -> int:
    """Obtiene el siguiente ID para préstamos."""
    return _increment_field('numerador_prestamo')

def get_next_conciliacion_id() -> int:
    """Obtiene el siguiente ID para conciliaciones."""
    return _increment_field('numerador_conciliacion')

def get_next_pagos_id() -> int:
    """Obtiene el siguiente ID para pagos."""
    return _increment_field('numerador_pagos')

def get_next_secuencial_1_id() -> int:
    """Obtiene el siguiente ID del contador auxiliar 1."""
    return _increment_field('numerador_aux_1')

def get_next_secuencial_2_id() -> int:
    """Obtiene el siguiente ID del contador auxiliar 2."""
    return _increment_field('numerador_aux_2')

def get_next_secuencial_3_id() -> int:
    """Obtiene el siguiente ID del contador auxiliar 3."""
    return _increment_field('numerador_aux_3')

def get_next_secuencial_4_id() -> int:
    """Obtiene el siguiente ID del contador auxiliar 4."""
    return _increment_field('numerador_aux_4')

def get_next_secuencial_5_id() -> int:
    """Obtiene el siguiente ID del contador auxiliar 5."""
    return _increment_field('numerador_aux_5')


#-----------------------------------------------------------------------------------------

def InBox_Pagos(archivo_pagos, usuario, cabezal):
    #identificar automáticamente el formato del archvo
    formato=f_identificar_formato(cabezal.nombre_archivo_id)
    
    if formato == "1-FORMATO PSE":
        return inbox_formato_pse(
            xls_file=archivo_pagos,
            nombre_archivo_id=cabezal.nombre_archivo_id,
            user=usuario,
            cabezal=cabezal
        )
        return ok, msg

    elif formato == "2-FORMATO ESTANDAR":
        return False, "El Formato Estándar aún no está implementado."

    elif formato == "3-FORMATO EXTRACTO BANCOLOMBIA":
        return inbox_formato_bancolombia(
            pdf_file=archivo_pagos,
            nombre_archivo_id=cabezal.nombre_archivo_id,
            user=usuario,
            cabezal=cabezal
        )

    else:
        return False, "Formato desconocido."

#-----------------------------------------------------------------------------------------
def inbox_formato_bancolombia(pdf_file, nombre_archivo_id, user, cabezal):
    from .models import InBox_PagosDetalle
    cargados = 0
    valor_total = 0
    rechazados = 0

    try:
        pdf_bytes = io.BytesIO(pdf_file.read())
        texto = ""

        with pdfplumber.open(pdf_bytes) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    texto += t + "\n"

        lineas = texto.split("\n")

    except Exception as e:
        return False, f"Error leyendo PDF: {e}"

    for linea in lineas:
        reg = linea.strip()
        if len(reg) < 6:
            continue

        # MOVIMIENTOS (fecha YYYY/MM/DD)
        if reg[:4].isdigit() and reg[4] == "/":
            try:
                fecha_raw=safe(reg[0:10], 10)
                descripcion_raw=safe(clean_descripcion(reg),100)
                valor_raw = clean_money(reg)
                clase_movimiento_raw=clase_movimiento(descripcion_raw,Decimal(valor_raw)) 
                
                estado_conciliacion_raw = f_estado_conciliacion(clase_movimiento_raw)
                
                InBox_PagosDetalle.objects.create(
                    nombre_archivo_id_id=nombre_archivo_id,
                    banco_origen='BANCOLOMBIA',            
                    cuenta_bancaria = "03200001692",
                    tipo_cuenta_bancaria="AHORROS",
                    canal_red_pago="EXTRACTOBANCOLOMBIA",
                    #ref_bancaria=None,
                    #ref_red=None,
                    #ref_cliente_1=v_ref_1[0:11],
                    #ref_cliente_2=v_ref_2[0:11],
                    #ref_cliente_3=safe(v_descripcion, 100),
                    ref_cliente_3=descripcion_raw,
                    #estado_transaccion_reportado=None,
                    #cliente_id_reportado=None,
                    fecha_pago=datetime.strptime(fecha_raw, "%Y/%m/%d").date(),
                    estado_pago=f_estado_pago(clase_movimiento_raw),
                    #cliente_id_real=None, 
                    #prestamo_id_real=None,
                    #poliza_id_real=None,
                    estado_conciliacion=estado_conciliacion_raw,
                    #lote_pse = lote_pse_raw
                    #fecha_conciliacion = fecha_conciliacion_raw
                    
                    clase_movimiento=clase_movimiento_raw, 
                    estado_fragmentacion="NO_REQUIERE",
                    valor_pago=Decimal(valor_raw),
                    creado_por=user,
                )
                if clase_movimiento_raw != "EXCLUIDO":
                    cargados += 1
                    valor_total += float(valor_raw)

            except Exception:
                rechazados += 1
                continue

    # ACTUALIZAR CABEZAL (fuera del for)
    cabezal.formato=f_identificar_formato(nombre_archivo_id)
    cabezal.valor_total = valor_total
    cabezal.registros_cargados = cargados
    cabezal.registros_rechazados = rechazados
    cabezal.estado_proceso_archivo = "RECIBIDO"
    cabezal.save()

    return True, f"{cargados} registros cargados, {rechazados} rechazados."

#-----------------------------------------------------------------------------------------
def inbox_formato_pse(xls_file, nombre_archivo_id, user, cabezal):
    import pandas as pd
    from decimal import Decimal
    from django.db import transaction
    from .models import InBox_PagosDetalle

    # -------------------------------------------------------
    # Cargar archivo Excel
    # -------------------------------------------------------
    try:
        df = pd.read_excel(xls_file, sheet_name="Facturas")
    except Exception as e:
        return False, f"Error leyendo archivo PSE: {e}"

    # -------------------------------------------------------
    # Validar columnas obligatorias
    # -------------------------------------------------------
    columnas = [
        "Modalidad de Pago",
        "Estado",
        "Valor Factura",
        "Valor pago",
        "CUS",
        "Fecha Pago",
        "Estado Transacción/Operación",
        "Nombre",
        "Referencia Externa",
        "Ramo",
        "Cédula"
    ]

    for c in columnas:
        if c not in df.columns:
            return False, f"Columna faltante en archivo PSE: {c}"

    # -------------------------------------------------------
    # Variables de control
    # -------------------------------------------------------
    cargados = 0
    rechazados = 0
    valor_total = Decimal("0.00")

    # -------------------------------------------------------
    # Transacción atómica
    # -------------------------------------------------------
    with transaction.atomic():
        for idx, row in df.iterrows():
            try:
                # -------------------------------------------------------
                # Fecha y hora
                # -------------------------------------------------------
                fecha_hora_str = str(row["Fecha Pago"])  # ej: "2025-01-05 13:24:59"
                fecha_pago = fecha_hora_str[:10]
                
                v_idcliente = row["Cédula"]
                v_idprestamo, v_cuota_0_sn = obtener_prestamo_unico(cliente_id=idcliente)
                 
                if not v_idprestamo: 
                    v_idcliente = None
                    v_idprestamo = None
                
                if v_cuota_0_sn == "SI":
                    v_observaciones = "PAGO CUOTA INICIAL"
                    v_cuota_inicial = "SI"
                else:
                    v_observaciones = None
                    v_cuota_inicial = None
                try:
                    hora_pago = pd.to_datetime(fecha_hora_str).time()
                except:
                    hora_pago = None

                # -------------------------------------------------------
                # Crear registro detalle pagos PSE
                # -------------------------------------------------------
                p = InBox_PagosDetalle(                  
                    nombre_archivo_id_id=nombre_archivo_id,
                    banco_origen="BANCOLOMBIA",            
                    cuenta_bancaria = "03200001692",
                    tipo_cuenta_bancaria="AHORROS",
                    canal_red_pago="PSE",
                    ref_bancaria=row["CUS"],
                    ref_red="",
                    ref_cliente_1=row["Nombre"],
                    ref_cliente_2=row["Referencia Externa"],
                    ref_cliente_3=row["Ramo"],
                    estado_transaccion_reportado=row["Estado"],
                    cliente_id_reportado=row["Cédula"],
                    cliente_id_real = v_idcliente,
                    prestamo_id_real = v_idprestamo,
                    estado_pago="RECIBIDO",
                    clase_movimiento="PAGO_PSE",   #pago pse 
                    estado_fragmentacion="NO_REQUIERE",
                    estado_conciliacion='NO',
                    valor_pago=Decimal(str(row["Valor pago"])),
                    fecha_pago=fecha_pago,
                    observaciones=v_observaciones,
                    cuota_inicial = v_cuota_inicial,
                    creado_por=user,
                    
                )

                p.save()

                cargados += 1
                valor_total += Decimal(str(row["Valor pago"]))

            except Exception:
                rechazados += 1
                continue

        # -------------------------------------------------------
        # Actualizar totales en el cabezal
        # -------------------------------------------------------
        cabezal.formato=f_identificar_formato(nombre_archivo_id)
        cabezal.valor_total = valor_total
        cabezal.registros_cargados = cargados
        cabezal.registros_rechazados = rechazados
        cabezal.estado_proceso_archivo = "RECIBIDO"
        cabezal.save()

    return True, f"Archivo procesado: {cargados} registros cargados, {rechazados} rechazados."


#-----------------------------

def safe(txt, maxlen):
    """Evita errores de overflow truncando cadenas."""
    if txt is None:
        return ""
    # Aseguramos que la entrada sea string antes de truncar y limpiar
    return str(txt).strip()[:maxlen]

#================
#Limpiar el valor
#================

def clean_money(valor_str):
    if not valor_str:
        return None        
    #limpio = str(valor_str).replace('$', '').replace(',', '').strip()
    limpio=valor_str[valor_str.rfind(" "):].replace(" ", "").replace(",", "")
    
    try:
        valor_numerico = float(limpio)
        
        return valor_numerico       
    except ValueError as e:
        # Maneja la excepción si la cadena limpia todavía no es un número válido.
        # print(f"Error de conversión en clean_money: {e} para cadena limpia '{limpio}'")
        return None

#================
#Limpiar descripción
#================
def clean_descripcion(reg: str) -> str:
    """
    Extrae la subcadena desde el inicio (pos 0) hasta justo antes 
    del primer número (0-9) o el signo de resta (-).
    """
    reg=reg[11:70]
    
    # 1. Definir el patrón de búsqueda: [0-9-] busca cualquier dígito o un guion.
    #patron = r'[0-9-]'
    patron = r'\d{2}|-'

    # 2. Buscar la primera coincidencia en la cadena 'reg'.
    match = re.search(patron, reg)

    if match:
        # Si hay una coincidencia, match.start() devuelve el índice de ese primer carácter.
        indice_fin = match.start()
        
        # 3. Recortar la subcadena desde el inicio hasta ese índice.
        # Se usa .rstrip() para quitar cualquier espacio sobrante al final de la descripción.
        return reg[:indice_fin].rstrip()
    else:
        # Si no se encuentra ningún número o signo, se devuelve la cadena completa.
        return reg.rstrip()

#================
#Excluir registros
#================
#from typing import Literal
def f_estado_pago(tipo_mov: str) -> Literal['EXCLUIDO', 'RECIBIDO']:
    if tipo_mov and tipo_mov.upper() == 'EXCLUIDO':
        return 'EXCLUIDO'
    else:
        return 'RECIBIDO'

#===================
#Estado conciliación
#===================
        
def f_estado_conciliacion(p_clas_mov: str) -> Union[str, Literal['NO', 'SI']]:
    clasificacion = p_clas_mov.upper() if p_clas_mov else ""
    if clasificacion == 'LOTE_PSE' or clasificacion == 'PAGO_PSE':
        return 'NO'
    elif clasificacion == 'PAGO_BANCOL':
        return 'SI'
    # 3. Caso por defecto
    else:
        return ""
        

#-----------------------------------------------------------------------------------------
def clase_movimiento(reg: str, valor: Optional[Decimal]) -> str:
    
    # 1.Si el valor no existe (es None) o es negativo.
    if valor is None or valor < 0:
        return "EXCLUIDO"
        # Se sale de la función inmediatamente (Early exit)
    
    # a) PAGO VIRTUAL PSE
    if "PAGO VIRTUAL PSE" in reg:
        return "LOTE_PSE"

    # b) ABONO INTERESES AHORROS
    elif "ABONO INTERESES AHORROS" in reg:
        return "EXCLUIDO"
        # Se sale de la función inmediatamente

    # c) IMPTO GOBIERNO 4X1000
    elif "IMPTO GOBIERNO 4X1000" in reg:
        return "EXCLUIDO"
        # Se sale de la función inmediatamente
    
    # Si pasa todas las comprobaciones anteriores, se asigna la clasificación por defecto.
    return "PAGO_BANCOL"
    
#-----------------------------------------------------------------------------------------

def cerrar_periodo_interes(prestamo_id: int, fecha_corte: date, pago_referencia=None, numero_asiento_contable=None,  capital_aplicado=0):
    """
    Cierra el período de interés para un préstamo en una fecha específica. -v9- 
    Crea un registro de tipo 'CAUSAC' en Historia_Prestamos con el interés causado.
    Nota si modifica esta funcion, analice si debe modificar cerrar_todos_periodos_hasta
    """
    from decimal import Decimal
    from django.db.models import Sum
    from .models import Conceptos_Transacciones, Prestamos, Historia_Prestamos

    # Obtener el concepto de causación
    concepto_causacion = Conceptos_Transacciones.objects.get(concepto_id="CAUSAC")
    
    # Obtener el préstamo (con bloqueo)
    prestamo = Prestamos.objects.select_for_update().get(prestamo_id=prestamo_id)

    # Evitar duplicados de causación en la misma fecha  2025-12-04 lo quito porque puede haber varios pagos el mismo dia...
    #if Historia_Prestamos.objects.filter(
    #    prestamo_id=prestamo_id,
    #    fecha_proceso=fecha_corte,         #
    #    concepto_id=concepto_causacion
    #).exists():
    #    return

    # === 1. Obtener el monto del desembolso (capital inicial) ===
    try:
        valor_desembolso = prestamo.prestamo_id.valor
    except (AttributeError, TypeError):
        valor_desembolso = Decimal('0')

    # === 2. Calcular total de abonos a capital ANTES de la fecha de corte ===
    abonos_result = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo_id,
        fecha_efectiva__isnull=False, 
        fecha_efectiva__lte=fecha_corte,   # <= incluye el mismo día
        abono_capital__gt=0
    ).aggregate(total_abonos=Sum('abono_capital'))
    
    total_abonos = abonos_result['total_abonos'] or Decimal('0')
    saldo_capital = valor_desembolso - total_abonos
    if saldo_capital < 0:
        saldo_capital = Decimal('0')

    # === 3. Obtener intereses acumulados del último registro anterior ===
    ultimo_registro = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo_id,
        fecha_proceso__lt=fecha_corte
    ).order_by('-fecha_proceso').first()
    
    intereses_acumulados = ultimo_registro.intrs_ctes or Decimal('0') if ultimo_registro else Decimal('0')

    # === 4. Calcular interés del día ===
    tasa_diaria = prestamo.tasa / Decimal('100') / Decimal('360')
    interes_diario = saldo_capital * tasa_diaria if saldo_capital > 0 else Decimal('0')
    nuevos_intereses = intereses_acumulados + interes_diario

    # === 5. Crear registro de causación (CAUSAC) ===
    Historia_Prestamos.objects.create(
        prestamo_id=prestamo,                # ✅ FK: usa 'prestamo_id', no 'prestamo'
        fecha_proceso=fecha_corte,           # ✅ obligatorio
        fecha_efectiva=None,                 # ✅ causación no tiene fecha efectiva
        concepto_id=concepto_causacion,      # ✅ FK al concepto
        monto_transaccion=interes_diario,    # ✅ interés causado HOY
        intrs_ctes=nuevos_intereses,         # ✅ total acumulado hasta HOY
        abono_capital=Decimal('0'),          # ✅ no es un abono
        seguro=Decimal('0'),                 # ✅ no causamos seguro aquí
        fee=Decimal('0'),                    # ✅ no causamos fee aquí
        tasa=prestamo.tasa,
        fecha_vencimiento=fecha_corte,       # ✅ opcional: puedes usar None si prefieres
        estado='TRANSACCION',
        usuario='sistema',                   # ✅ ajusta si tu app requiere otro valor
        numero_cuota=None,                   # ✅ no pertenece a una cuota
        ordinal_interno=10,                  #Ordinal interno debe ser 10
        numero_operacion=numero_asiento_contable,                   # ✅ asegúrate que no duplique unique_together
        capital_aplicado_periodo=capital_aplicado or 0,  
        numero_pago_referencia=pago_referencia or "",
        numero_asiento_contable=numero_asiento_contable or 0
    )
#-------------------------------------------------------------------------------------------------------
#   funciones para consulta de causacion - contabilidad
#----------------------------------------------------------------------------------------------------
# utils.py - NUEVA FUNCIÓN PARA CONSULTAS
def cerrar_todos_periodos_hasta(prestamo_id: int, fecha_corte: date):
    """
    Cierra TODOS los períodos de interés faltantes hasta fecha_corte.
    Crea registros CAUSAC para cada fecha de evento (desembolso, pagos) hasta fecha_corte.
    """
    from decimal import Decimal
    from django.db.models import Sum
    from .models import Conceptos_Transacciones, Prestamos, Historia_Prestamos, Pagos
    
    # Obtener el concepto de causación
    concepto_causacion = Conceptos_Transacciones.objects.get(concepto_id="CAUSAC")
    
    # Obtener el préstamo
    prestamo = Prestamos.objects.select_for_update().get(prestamo_id=prestamo_id)
    
    # Obtener todas las fechas de eventos hasta fecha_corte
    fechas_eventos = set()
    
    # Agregar fecha de desembolso
    if prestamo.fecha_desembolso and prestamo.fecha_desembolso <= fecha_corte:
        fechas_eventos.add(prestamo.fecha_desembolso)
    
    # Agregar fechas de pagos aplicados
    pagos = Pagos.objects.filter(
        prestamo_id_real=prestamo_id,
        fecha_pago__lte=fecha_corte,
        estado_pago='aplicado'
    )
    for pago in pagos:
        fechas_eventos.add(pago.fecha_pago)
    
    # Si no hay eventos, al menos agregar la fecha_corte
    if not fechas_eventos:
        fechas_eventos.add(fecha_corte)
    
    # Ordenar fechas
    fechas_ordenadas = sorted(fechas_eventos)
    
    # Cerrar cada período faltante
    for fecha in fechas_ordenadas:
        # Verificar si ya existe registro para esta fecha
        if Historia_Prestamos.objects.filter(
            prestamo_id=prestamo_id,
            fecha_proceso=fecha,
            concepto_id=concepto_causacion
        ).exists():
            continue
        
        # === Reutilizar la lógica de tu función original ===
        try:
            valor_desembolso = prestamo.prestamo_id.valor
        except (AttributeError, TypeError):
            valor_desembolso = Decimal('0')

        abonos_result = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo_id,
            fecha_efectiva__isnull=False, 
            fecha_efectiva__lte=fecha,
            abono_capital__gt=0
        ).aggregate(total_abonos=Sum('abono_capital'))
        
        total_abonos = abonos_result['total_abonos'] or Decimal('0')
        saldo_capital = valor_desembolso - total_abonos
        if saldo_capital < 0:
            saldo_capital = Decimal('0')

        ultimo_registro = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo_id,
            fecha_proceso__lt=fecha
        ).order_by('-fecha_proceso').first()
        
        intereses_acumulados = ultimo_registro.intrs_ctes or Decimal('0') if ultimo_registro else Decimal('0')

        tasa_diaria = prestamo.tasa / Decimal('100') / Decimal('360')
        interes_diario = saldo_capital * tasa_diaria if saldo_capital > 0 else Decimal('0')
        nuevos_intereses = intereses_acumulados + interes_diario

        Historia_Prestamos.objects.create(
            prestamo_id=prestamo,
            fecha_proceso=fecha,
            fecha_efectiva=None,
            concepto_id=concepto_causacion,
            monto_transaccion=interes_diario,
            intrs_ctes=nuevos_intereses,
            abono_capital=Decimal('0'),
            seguro=Decimal('0'),
            fee=Decimal('0'),
            tasa=prestamo.tasa,
            fecha_vencimiento=fecha,
            estado='TRANSACCION',
            usuario='sistema',
            numero_cuota=None,
            ordinal_interno=10,
            numero_operacion=1
        )
#_______________________________________________________
def obtener_tasa_prestamo(prestamo_id):
    """
    Obtiene la tasa anual del préstamo (como Decimal).
    La tasa se almacena en el campo 'tasa' del modelo Prestamos.
    
    Parámetros:
        prestamo_id: ID del préstamo (coincide con el ID del desembolso)
    
    Retorna:
        Decimal: tasa anual en formato decimal (ej: 0.25 para 25%)
    """
    try:
        prestamo = Prestamos.objects.get(prestamo_id=prestamo_id)
        # Asegurarse de que sea Decimal y no float
        return Decimal(str(prestamo.tasa))
    except Prestamos.DoesNotExist:
        raise ValueError(f"Préstamo con ID {prestamo_id} no encontrado.")
# En cualquier parte de tu código
#tasa = obtener_tasa_prestamo(38)  # préstamo ID 38
#print(tasa)  # Ej: Decimal('0.2500') si la tasa es 25%
#------------------------------------------------------------------------
# utils.py - Actualiza SOLO las funciones de consulta

def obtener_intereses_causados_a_fecha(prestamo_id, fecha_corte):
    """Calcula intereses causados hasta fecha_corte RECALCULANDO desde cero."""
    from decimal import Decimal
    from .models import Prestamos, Historia_Prestamos, Conceptos_Transacciones
    
    # ✅ Usar la NUEVA función para consultas
    cerrar_todos_periodos_hasta(prestamo_id, fecha_corte)
    
    concepto_causacion = Conceptos_Transacciones.objects.get(concepto_id="CAUSAC")
    ultimo_registro = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo_id,
        concepto_id=concepto_causacion,
        fecha_proceso__lte=fecha_corte
    ).order_by('-fecha_proceso').first()
    
    return ultimo_registro.intrs_ctes if ultimo_registro else Decimal('0.00')

 
#-----------------------------------------------------------------------------------------
def aplicar_pago_cuota_inicial(desembolso, prestamo, usuario: str,  numero_asiento_contable=None):
    """
    Aplica el pago de la cuota #1 dentro del contexto de procesar_desembolsos_pendientes.
    Usa los campos numero_transaccion_cuota_1 y valor_cuota_1 del desembolso.
    Debe ejecutarse dentro de un bloque transaction.atomic() externo.
    """
    from decimal import Decimal, ROUND_HALF_UP
    from .models import Pagos, Fechas_Sistema, Conceptos_Transacciones, Historia_Prestamos, Detalle_Aplicacion_Pago
    from django.utils import timezone
    from django.db.models import Max

    # Validar campos obligatorios en desembolso
    if not desembolso.numero_transaccion_cuota_1:
        raise ValueError(f"Desembolso {desembolso.prestamo_id}: falta numero_transaccion_cuota_1")
    if desembolso.valor_cuota_1 is None:
        raise ValueError(f"Desembolso {desembolso.prestamo_id}: falta valor_cuota_1")

    pago_id = desembolso.numero_transaccion_cuota_1
    valor_esperado = Decimal(str(desembolso.valor_cuota_1)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # Obtener el pago desde Pagos
    try:
        pago = Pagos.objects.select_for_update().get(ref_bancaria=pago_id)
    except Pagos.DoesNotExist:
        raise ValueError(f"El pago con ID {pago_id} vs {desembolso.numero_transaccion_cuota_1} (cuota inicial) no existe en la tabla Pagos.")

    # Validaciones de coherencia
    valor_pago = Decimal(str(pago.valor_pago)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    if valor_pago != valor_esperado:
        raise ValueError(
            f"Inconsistencia en valor de cuota inicial. "
            f"Desembolso: {valor_esperado}, Pagos: {valor_pago}."
        )
    if pago.prestamo_id_real != desembolso.prestamo_id:
        raise ValueError(
            f"El pago {pago_id} no pertenece al préstamo {desembolso.prestamo_id}."
        )
    if pago.fecha_pago != desembolso.fecha_desembolso:
        raise ValueError(
            f"Fecha de pago ({pago.fecha_pago}) != fecha de desembolso ({desembolso.fecha_desembolso})."
        )

    # Si ya fue aplicado, omitir o fallar
    if pago.estado_pago == 'aplicado':
        raise ValueError(f"El pago de cuota inicial {pago_id} ya fue aplicado.")

    # Obtener fecha de proceso del sistema
    fechas_sistema = Fechas_Sistema.objects.first()
    if not fechas_sistema:
        raise ValueError("No se encontró ningún registro en Fechas_Sistema. Cree uno primero.")
    fecha_proceso_actual = fechas_sistema.fecha_proceso_actual

    # --- A partir de aquí, reutilizamos lógica de aplicar_pago (sin transacción) ---  ###################
    prestamo_id = pago.prestamo_id_real
    #monto_restante = valor_pago
    monto_restante = Decimal(str(pago.valor_pago)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # Conceptos
    concepto_cap = Conceptos_Transacciones.objects.get(concepto_id="PLANCAP")
    concepto_int = Conceptos_Transacciones.objects.get(concepto_id="PLANINT")
    concepto_seg = Conceptos_Transacciones.objects.get(concepto_id="PLANSEG")
    concepto_gto = Conceptos_Transacciones.objects.get(concepto_id="PLANGTO")
    concepto_excedente = Conceptos_Transacciones.objects.get(concepto_id="PAGOEXC")

    # Registros pendientes
    registros_pendientes = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo,
        concepto_id__in=[concepto_cap, concepto_int, concepto_seg, concepto_gto],
        estado="PENDIENTE"
    ).order_by('fecha_vencimiento', 'numero_cuota')

    prioridad = {concepto_cap.concepto_id: 1, concepto_int.concepto_id: 2, concepto_seg.concepto_id: 3, concepto_gto.concepto_id: 4}
    registros_pendientes = sorted(
        registros_pendientes,
        key=lambda r: (r.fecha_vencimiento, r.numero_cuota, prioridad.get(r.concepto_id_id, 99))
    )

    # Número de operación único
    ultimo_numero = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo_id,
        fecha_proceso=fecha_proceso_actual
    ).aggregate(Max('numero_operacion'))['numero_operacion__max'] or 0
	
    numero_siguiente = ultimo_numero + 1

    # Acumulador específico para CAPITAL aplicado (solo PLANCAP)
    total_capital_aplicado = Decimal('0.00')

    # Aplicar pago a cada componente
    registros_creados = []

    for reg in registros_pendientes:
        if monto_restante <= 0:
            break

        if reg.concepto_id_id == concepto_cap.concepto_id:
            pagado = reg.abono_capital; campo_pago = 'abono_capital'; comp_nombre = 'CAPITAL'
        elif reg.concepto_id_id == concepto_int.concepto_id:
            pagado = reg.intrs_ctes; campo_pago = 'intrs_ctes'; comp_nombre = 'INTERES'
        elif reg.concepto_id_id == concepto_seg.concepto_id:
            pagado = reg.seguro; campo_pago = 'seguro'; comp_nombre = 'SEGURO'
        elif reg.concepto_id_id == concepto_gto.concepto_id:
            pagado = reg.fee; campo_pago = 'fee'; comp_nombre = 'GASTOS'
        else:
            continue

        saldo_pendiente = reg.monto_transaccion - pagado
        if saldo_pendiente <= 0:
            continue

        aplicado = min(monto_restante, saldo_pendiente)
        monto_restante -= aplicado

        setattr(reg, campo_pago, getattr(reg, campo_pago) + aplicado)
        if getattr(reg, campo_pago) >= reg.monto_transaccion:
            reg.estado = "PAGADA"
        if reg.fecha_efectiva is None:
            reg.fecha_efectiva = pago.fecha_pago
        reg.fecha_proceso = fecha_proceso_actual
        reg.usuario = usuario
        reg.numero_operacion = numero_siguiente
        reg.numero_asiento_contable = numero_asiento_contable
        reg.save()

        Detalle_Aplicacion_Pago.objects.create(
            pago=pago,
            historia_prestamo=reg,
            monto_aplicado=aplicado,
            componente=comp_nombre
        )
        numero_siguiente += 1

    # === Manejo de excedente ===
    if monto_restante > 0:
        siguiente_capital = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo_id,
            concepto_id=concepto_cap,
            estado="PENDIENTE"
        ).order_by('fecha_vencimiento', 'numero_cuota').first()

        if siguiente_capital:
            monto_aplicar = min(monto_restante, siguiente_capital.monto_transaccion)
            siguiente_capital.abono_capital += monto_aplicar
            if siguiente_capital.abono_capital >= siguiente_capital.monto_transaccion:
                siguiente_capital.estado = "PAGADA"
            siguiente_capital.fecha_efectiva = pago.fecha_pago
            siguiente_capital.usuario = usuario
            siguiente_capital.fecha_proceso = fecha_proceso_actual
            siguiente_capital.numero_asiento_contable = numero_asiento_contable or 0
            siguiente_capital.save()

            registros_creados.append(siguiente_capital)

            Detalle_Aplicacion_Pago.objects.create(
                pago=pago,
                historia_prestamo=siguiente_capital,
                monto_aplicado=monto_aplicar,
                componente='CAPITAL'
            )
            monto_restante -= monto_aplicar
            total_capital_aplicado += monto_aplicar  # ✅ ¡Incluir excedente aplicado a capital!

        if monto_restante > 0:					  
            hist_excedente = Historia_Prestamos.objects.create(
                prestamo_id=prestamo_id,
                concepto_id=concepto_excedente,								   
                fecha_efectiva=pago.fecha_pago,							   
                fecha_proceso=fecha_proceso_actual,		   
                fecha_vencimiento=pago.fecha_pago,
                monto_transaccion=monto_restante,
                abono_capital=monto_restante,
                estado="TRANSACCION",
                numero_cuota=0,
                usuario=usuario,
                numero_operacion=pago.pago_id,
                ordinal_interno=20,												
                numero_asiento_contable=numero_asiento_contable or 0
            )
            registros_creados.append(hist_excedente)
				
            Detalle_Aplicacion_Pago.objects.create(
                pago=pago,
                historia_prestamo=hist_excedente,
                monto_aplicado=monto_restante,
                componente='EXCEDENTE'
            )

    # === ✅ AHORA: Cerrar período de interés UNA SOLA VEZ, con el total de capital aplicado ===
    from .utils import cerrar_periodo_interes
    cerrar_periodo_interes(
        prestamo_id,
        pago.fecha_pago,
        pago_referencia=f"PAGO_{pago.pago_id}",
        numero_asiento_contable=numero_asiento_contable,
        capital_aplicado=total_capital_aplicado  # ✅ PASAR EL TOTAL ACUMULADO
    )

    # Finalizar pago
    pago.estado_pago = 'aplicado'
    pago.fecha_aplicacion_pago = timezone.now()
    pago.save()

    return {
        'status': 'success',
        'message': f'Pago {pago_id} aplicado exitosamente.',
        'pago_id': pago_id,
        'monto_aplicado': float(pago.valor_pago - (monto_restante)),
        'excedente': float(monto_restante),
    }
    
#-----------------------------------------------------------------------------------------
def get_fecha_proceso_actual() -> date:
    """Devuelve la fecha de proceso actual (singleton seguro)"""
    from .models import Fechas_Sistema
    return Fechas_Sistema.load().fecha_proceso_actual

def get_fecha_proceso_anterior() -> date:
    from .models import Fechas_Sistema
    return Fechas_Sistema.load().fecha_proceso_anterior

def get_fecha_proximo_proceso() -> date:
    from .models import Fechas_Sistema
    return Fechas_Sistema.load().fecha_proximo_proceso

def get_estado_sistema() -> str:
    from .models import Fechas_Sistema
    return Fechas_Sistema.load().estado_sistema

def get_modo_fecha_sistema() -> str:
    from .models import Fechas_Sistema
    return Fechas_Sistema.load().modo_fecha_sistema

def sistema_esta_abierto() -> bool:
    #from .models import Fechas_Sistema
    return get_estado_sistema() == 'ABIERTO'

def fecha_proceso_es_hoy() -> bool:
    from .models import Fechas_Sistema
    return get_fecha_proceso_actual() == timezone.now().date()
    
if sistema_esta_abierto():
    hoy = get_fecha_proceso_actual()    
    

#------------ ***   para el Historial del prestamo a xls ******
# --- utils.py (AJUSTADO SIN MERGE) ---
def generar_reporte_excel_en_memoria(prestamo_id):
    """
    Genera un reporte Excel del historial de un préstamo y devuelve un BytesIO buffer.
    Implementa chequeos de MergedCell y salta la fila de Desembolso en el formato numérico.
    """
    # Obtener fecha de corte
    fecha_sistema = Fechas_Sistema.objects.first()
    if not fecha_sistema:
        raise Exception("No hay registro en Fechas_Sistema.")
    fecha_corte = fecha_sistema.fecha_proceso_actual

    # === Obtener información del préstamo y cliente ===
    prestamo = Prestamos.objects.select_related('cliente_id').get(prestamo_id=prestamo_id)
    cliente = prestamo.cliente_id
    nombre_completo = f"{cliente.nombre} {cliente.apellido}" if cliente else "Cliente no encontrado"

    # === Obtener desembolso ===
    try:
        desembolso = Desembolsos.objects.get(prestamo_id=prestamo_id)
        monto_desembolso = float(desembolso.valor)
        fecha_desembolso = desembolso.fecha_desembolso
        tasa_desembolso = float(prestamo.tasa)
    except Desembolsos.DoesNotExist:
        monto_desembolso = 0.0
        fecha_desembolso = "No encontrado"
        tasa_desembolso = 0.0

    # === Contar cuotas pagadas y pendientes ===
    cuotas_vencimiento = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo_id,
        fecha_vencimiento__lte=fecha_corte
    ).exclude(concepto_id__concepto_id="CAUSAC").values_list('fecha_vencimiento', flat=True).distinct()

    cuotas_pagadas = 0
    cuotas_pendientes = 0
    
    for fecha_venc in cuotas_vencimiento:
        componentes_cuota = Historia_Prestamos.objects.filter(
            prestamo_id=prestamo_id,
            fecha_vencimiento=fecha_venc,
            fecha_vencimiento__lte=fecha_corte
        ).exclude(concepto_id__concepto_id="CAUSAC")
        
        total_monto = componentes_cuota.aggregate(total=Sum('monto_transaccion'))['total'] or Decimal('0')
        total_pagado = 0
        
        for comp in componentes_cuota:
            detalles = Detalle_Aplicacion_Pago.objects.filter(historia_prestamo=comp)
            total_pagado += sum(d.monto_aplicado for d in detalles)
        
        if total_pagado >= total_monto - Decimal('0.01'):
            cuotas_pagadas += 1
        else:
            cuotas_pendientes += 1

    # === Agrupar cuotas ===
    from collections import defaultdict
    cuotas = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo_id,
        fecha_vencimiento__lte=fecha_corte
    ).exclude(concepto_id__concepto_id="CAUSAC").order_by('fecha_vencimiento', 'numero_cuota')

    cuotas_agrupadas = defaultdict(lambda: {
        'registro': None, 'capital_prog': Decimal('0'), 'interes_prog': Decimal('0'), 
        'seguro_prog': Decimal('0'), 'fee_prog': Decimal('0'), 'capital_pag': Decimal('0'), 
        'interes_pag': Decimal('0'), 'seguro_pag': Decimal('0'), 'fee_pag': Decimal('0'), 
        'fecha_efectiva': None, 'fecha_proceso': None, 'fecha_vencimiento': None, 
        'tasa': Decimal('0'), 'fecha_pago': None, 'ref_bancaria': set(), 'estado': 'PENDIENTE'
    })

    concepto_a_tipo = {'PLANCAP': 'capital', 'PLANINT': 'interes', 'PLANSEG': 'seguro', 'PLANGTO': 'fee'}
    
    for reg in cuotas:
        num_cuota = reg.numero_cuota or 0
        tipo = concepto_a_tipo.get(reg.concepto_id.concepto_id)
        if tipo:
            if cuotas_agrupadas[num_cuota]['registro'] is None:
                cuotas_agrupadas[num_cuota]['registro'] = reg
                cuotas_agrupadas[num_cuota]['fecha_efectiva'] = reg.fecha_efectiva
                cuotas_agrupadas[num_cuota]['fecha_proceso'] = reg.fecha_proceso
                cuotas_agrupadas[num_cuota]['fecha_vencimiento'] = reg.fecha_vencimiento
                cuotas_agrupadas[num_cuota]['tasa'] = reg.tasa

            cuotas_agrupadas[num_cuota][f'{tipo}_prog'] += reg.monto_transaccion

            detalles = Detalle_Aplicacion_Pago.objects.filter(historia_prestamo=reg)
            total_pagado = sum(d.monto_aplicado for d in detalles)
            cuotas_agrupadas[num_cuota][f'{tipo}_pag'] += total_pagado

            if detalles.exists():
                fecha_pago_max = max(d.fecha_aplicacion.date() for d in detalles)
                if cuotas_agrupadas[num_cuota]['fecha_pago'] is None or fecha_pago_max > cuotas_agrupadas[num_cuota]['fecha_pago']:
                    cuotas_agrupadas[num_cuota]['fecha_pago'] = fecha_pago_max
                refs = {d.pago.ref_bancaria for d in detalles if d.pago.ref_bancaria}
                cuotas_agrupadas[num_cuota]['ref_bancaria'].update(refs)

            if total_pagado < reg.monto_transaccion:
                cuotas_agrupadas[num_cuota]['estado'] = 'PENDIENTE'
            else:
                cuotas_agrupadas[num_cuota]['estado'] = 'PAGADA'

    # === Preparar datos ===
    datos_excel = []
    # Fila 5: Datos de desembolso
    datos_excel.append([
        "Desembolso", "", fecha_desembolso, "", "", "", "", monto_desembolso, "", "", "", "", monto_desembolso, 0, "DESEMBOLSADO", "", "N/A", tasa_desembolso, "", ""
    ])

    total_pendiente_general = Decimal('0')
    for num_cuota in sorted(cuotas_agrupadas.keys()):
        c = cuotas_agrupadas[num_cuota]
        total_prog = c['capital_prog'] + c['interes_prog'] + c['seguro_prog'] + c['fee_prog']
        total_pag = c['capital_pag'] + c['interes_pag'] + c['seguro_pag'] + c['fee_pag']
        total_pend = total_prog - total_pag
        total_pendiente_general += total_pend
        estado_final = "PAGADA" if total_pend <= Decimal('0.01') else c['estado']

        datos_excel.append([
            "Cuota", num_cuota, 
            c['fecha_efectiva'] or c['registro'].fecha_vencimiento if c['registro'] else '',
            float(c['capital_prog']), float(c['interes_prog']), float(c['seguro_prog']), float(c['fee_prog']), float(total_prog),
            float(c['capital_pag']), float(c['interes_pag']), float(c['seguro_pag']), float(c['fee_pag']), float(total_pag), float(total_pend),
            estado_final, 
            c['fecha_pago'] or '',
            "; ".join(c['ref_bancaria']) if c['ref_bancaria'] else "N/A",
            float(c['tasa']) if c['tasa'] else 0.0,
            c['fecha_proceso'] or '',
            c['fecha_vencimiento'] or ''
        ])

    # === Crear Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = f"Préstamo {prestamo_id}"

    # TÍTULOS SIN MERGE (Fila 1 y 2)
    ws['A1'] = f"REPORTE HISTORIAL PRÉSTAMO {prestamo_id} - {nombre_completo}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='left')

    ws['A2'] = f"Cuotas Pagadas: {cuotas_pagadas} | Cuotas Pendientes: {cuotas_pendientes}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='left')

    # Encabezados de columnas (Fila 4)
    headers = [
        "Tipo", "Cuota", "Fecha Efectiva",
        "Capital Programado", "Interés Programado", "Seguro Programado", "Fee Programado", "Total Programado",
        "Capital Pagado", "Interés Pagado", "Seguro Pagado", "Fee Pagado", "Valor Transacción", "Total Pendiente",
        "Estado", "Fecha aplicacion pago", "Ref. Bancaria",
        "Tasa", "Fecha proceso", "Fecha Vencimiento"
    ]
    
    # Insertar headers en la fila 4
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
        ws.cell(row=4, column=col).font = Font(bold=True)
        ws.cell(row=4, column=col).fill = PatternFill("solid", fgColor="D9E2F3")

    # Insertar datos a partir de la fila 5
    for fila_datos in datos_excel:
        ws.append(fila_datos)

    # 🛑 Bloque de formato numérico: Empezar en la fila 6 (min_row=6) para saltar el Desembolso (Fila 5)
    
    # Aplicar formato de números (Columnas 4 a 14)
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=4, max_col=14):
        for cell in row:
            if not isinstance(cell, MergedCell) and isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = '#,##0.00'

    # Formato para la tasa (columna 18)
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=18, max_col=18):
        for cell in row:
            if not isinstance(cell, MergedCell) and isinstance(cell.value, (int, float, Decimal)):
                cell.number_format = '0.0000'

    # Agregar resumen al final
    ws.append([])
    resumen_row = ws.max_row + 1
    ws.cell(row=resumen_row, column=1, value="SALDO TOTAL PENDIENTE:")
    ws.cell(row=resumen_row, column=2, value=float(total_pendiente_general))
    ws.cell(resumen_row, column=2).number_format = '#,##0.00'
    ws.cell(resumen_row, column=1).font = Font(bold=True)
    ws.cell(resumen_row, column=2).font = Font(bold=True, color="FF0000")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

#------------------- fin historial
def cerrar_periodo_interes_migracion(prestamo_id: int, fecha_corte: date, pago_referencia=None, numero_asiento_contable=None, capital_aplicado=0):
    """
    Cierra el período de interés para un préstamo en una fecha específica. 
    Versión especial para migración que recibe prestamo_id como entero y lo convierte a instancia.
    """
    from decimal import Decimal
    from django.db.models import Sum
    from .models import Conceptos_Transacciones, Prestamos, Historia_Prestamos

    concepto_causacion = Conceptos_Transacciones.objects.get(concepto_id="CAUSAC")
    prestamo = Prestamos.objects.select_for_update().get(prestamo_id=prestamo_id)

    try:
        valor_desembolso = prestamo.prestamo_id.valor
    except (AttributeError, TypeError):
        valor_desembolso = Decimal('0')

    abonos_result = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo_id,
        fecha_efectiva__isnull=False, 
        fecha_efectiva__lte=fecha_corte,
        abono_capital__gt=0
    ).aggregate(total_abonos=Sum('abono_capital'))
    
    total_abonos = abonos_result['total_abonos'] or Decimal('0')
    saldo_capital = valor_desembolso - total_abonos
    if saldo_capital < 0:
        saldo_capital = Decimal('0')

    ultimo_registro = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo_id,
        fecha_proceso__lt=fecha_corte
    ).order_by('-fecha_proceso').first()
    
    intereses_acumulados = ultimo_registro.intrs_ctes or Decimal('0') if ultimo_registro else Decimal('0')

    tasa_diaria = prestamo.tasa / Decimal('100') / Decimal('360')
    interes_diario = saldo_capital * tasa_diaria if saldo_capital > 0 else Decimal('0')
    nuevos_intereses = intereses_acumulados + interes_diario

    Historia_Prestamos.objects.create(
        prestamo_id=prestamo,
        fecha_proceso=fecha_corte,
        fecha_efectiva=None,
        concepto_id=concepto_causacion,
        monto_transaccion=interes_diario,
        intrs_ctes=nuevos_intereses,
        abono_capital=Decimal('0'),
        seguro=Decimal('0'),
        fee=Decimal('0'),
        tasa=prestamo.tasa,
        fecha_vencimiento=fecha_corte,
        estado='TRANSACCION',
        usuario='sistema',
        numero_cuota=None,
        ordinal_interno=10,
        numero_operacion=numero_asiento_contable,
        capital_aplicado_periodo=capital_aplicado or 0,
        numero_pago_referencia=pago_referencia or "",
        numero_asiento_contable=numero_asiento_contable or 0)
        

def confirmar_pagos(queryset, usuario):
    from .models import InBox_PagosDetalle, Pagos
     
    confirmados = 0
    errores = []

    with transaction.atomic():        
        # Aplicamos el filtro base y encadenamos las exclusiones solicitadas
        pagos_validos = queryset.filter(
            estado_pago="A_PROCESAR",
            estado_conciliacion="SI",
        ).exclude(
            # 1. Excluir si no tienen cliente o préstamo asignado (IDs nulos)
            cliente_id_real__isnull=True
        ).exclude(
            prestamo_id_real__isnull=True
        ).exclude(
            # 2. Excluir si está pendiente de fragmentación
            estado_fragmentacion="A_FRAGMENTAR"
        ).exclude(
            # 3. Excluir si es un movimiento tipo LOTE
            clase_movimiento="LOTE_PSE"
        )
        
        for pago in pagos_validos:
            try:
                Pagos.objects.create(
                    pago_id=pago.pago_id,
                    nombre_archivo_id=pago.nombre_archivo_id,
                    banco_origen=pago.banco_origen,
                    cuenta_bancaria=pago.cuenta_bancaria,
                    tipo_cuenta_bancaria=pago.tipo_cuenta_bancaria,
                    canal_red_pago=pago.canal_red_pago,
                    ref_bancaria=pago.ref_bancaria,
                    ref_red=pago.ref_red,
                    ref_cliente_1=pago.ref_cliente_1,
                    ref_cliente_2=pago.ref_cliente_2,
                    ref_cliente_3=pago.ref_cliente_3,
                    estado_transaccion_reportado=pago.estado_transaccion_reportado,
                    cliente_id_reportado=pago.cliente_id_reportado,
                    prestamo_id_reportado=pago.prestamo_id_reportado,
                    poliza_id_reportado=pago.poliza_id_reportado,
                    cliente_id_real=pago.cliente_id_real_id,
                    prestamo_id_real=pago.prestamo_id_real_id,
                    poliza_id_real=pago.poliza_id_real,
                    fecha_pago=pago.fecha_pago,
                    hora_pago="00:00:00", #pago.hora_pago,
                    fecha_conciliacion=pago.fecha_conciliacion,
                    estado_pago='CONCILIADO',
                    #estado_conciliacion=pago.estado_conciliacion,
                    valor_pago=pago.valor_pago,
                    observaciones=pago.observaciones,
                    creado_por=usuario,
                    fecha_aplicacion_pago=None,      #timezone.now(),
                )

                pago.estado_pago = "CONFIRMADO"
                pago.save(update_fields=["estado_pago"])

                confirmados += 1

            except Exception as e:
                errores.append(f"Pago {pago.pago_id}: {e}")

    if errores:
        return False, f"Confirmados {confirmados}. Errores: {' | '.join(errores[:5])}"

    return True, f"{confirmados} pagos confirmados correctamente"
    
#-----------------------------------------------------------------------------------------/*
 


