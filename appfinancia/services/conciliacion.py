# appfinancia/services/conciliacion.py

# 1. Librerías del sistema (Standard Library)
import os
import logging
from decimal import Decimal

# 2. Librerías de terceros (Third-party)
from openpyxl import Workbook

# 3. Django Core y utilidades
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from django.http import HttpResponse
from django.db.models import Case, When, IntegerField
from django.utils.timezone import is_aware, make_naive

# 4. Configuración del Logger (Solo una vez)
#logger = logging.getLogger(__name__)

#MAX_CANDIDATES_DP = 40  # límite práctico para DP; si hay más, usamos fallback greedy

# ---------------------------------------------------------------------------------------
# appfinancia/services/conciliacion.py

import os
import logging
from decimal import Decimal
from openpyxl import Workbook
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from django.db.models import Case, When, IntegerField
from django.utils.timezone import is_aware, make_naive

logger = logging.getLogger(__name__)

# Aumentamos a 100. El algoritmo DP con diccionarios en Python 3.10+ es muy eficiente.
MAX_CANDIDATES_DP = 100 

def _to_cents(dec):
    if dec is None: return 0
    return int((Decimal(dec) * 100).quantize(Decimal('1')))

def _find_subset_indices_amounts(amounts_cents, target_cents):
    """
    Algoritmo de Suma de Subconjuntos usando Programación Dinámica.
    Optimizado para encontrar la primera combinación exacta.
    """
    sums = {0: []}
    for idx, amt in enumerate(amounts_cents):
        new_sums = {}
        for s, indices in sums.items():
            new_val = s + amt
            if new_val == target_cents:
                return indices + [idx]
            if new_val < target_cents and new_val not in sums:
                new_sums[new_val] = indices + [idx]
        sums.update(new_sums)
    return None
    
def _greedy_try(target, candidates):
    """
    Intenta encontrar una combinación de pagos que sumen exactamente 'target'.
    """
    # 1. Validación de seguridad
    if isinstance(candidates, int):
        # Si por error llega un entero, no podemos iterar. 
        # Devolvemos None para que la conciliación simplemente no encuentre nada.
        return None

    result = []
    current_sum = 0
    
    # 2. Asegurarnos de que candidates sea una lista para ordenar
    # (Si es un QuerySet de Django, lo convertimos a lista)
    candidate_list = list(candidates)
    
    # 3. Ordenar de mayor a menor valor para el algoritmo voraz
    # Usamos .valor_pago porque es el campo del modelo
    sorted_candidates = sorted(candidate_list, key=lambda x: x.valor_pago, reverse=True)

    for c in sorted_candidates:
        if current_sum + c.valor_pago <= target:
            current_sum += c.valor_pago
            result.append(c)
        
        # Si llegamos al total exacto, éxito
        if current_sum == target:
            return result
            
    return None

#===================
# Función Principal
#===================
def conciliacion_por_movimiento(movimiento, candidatos_qs, num_conciliacion):
    """
    Lógica core de cruce.
    """
    target_cents = _to_cents(movimiento.valor_pago)
    if target_cents <= 0:
        return False, "Valor inválido", {}

    # Filtramos candidatos que tengan valor y materializamos
    candidatos = list(candidatos_qs.exclude(valor_pago__lte=0))
    amounts_cents = [_to_cents(c.valor_pago) for c in candidatos]

    indices = None
    if len(candidatos) <= MAX_CANDIDATES_DP:
        indices = _find_subset_indices_amounts(amounts_cents, target_cents)
    
    if indices is None:
        indices = _greedy_try(amounts_cents, target_cents)

    if indices is None:
        return False, "No se encontró combinación", {"count": len(candidatos)}

    try:
        with transaction.atomic():
            hijos_ids = []
            for idx in indices:
                c = candidatos[idx]
                c.lote_pse = movimiento.pago_id
                c.estado_conciliacion = "SI"
                c.conciliacion_id = num_conciliacion
                c.fecha_conciliacion = timezone.now()
                c.save(update_fields=["lote_pse", "estado_conciliacion", "conciliacion_id", "fecha_conciliacion"])
                hijos_ids.append(c.pago_id)

            movimiento.lote_pse = movimiento.pago_id # Marcamos el padre como su propio lote
            movimiento.estado_conciliacion = "SI"
            movimiento.conciliacion_id = num_conciliacion
            movimiento.fecha_conciliacion = timezone.now()
            movimiento.save(update_fields=["lote_pse", "estado_conciliacion", "conciliacion_id", "fecha_conciliacion"])

        return True, "OK", {"hijos_creados": len(hijos_ids), "hijos_ids": hijos_ids}
    except Exception as e:
        logger.exception("Error en transacción de conciliación")
        return False, str(e), {}


#---------Reporte Resumen de Conciliación-------------------------------------------------

# appfinancia/services/conciliacion.py
from django.utils.timezone import is_aware, make_naive
from openpyxl.styles import Font, PatternFill  # Importamos PatternFill para el color

def _excel_datetime(value):
    if value and is_aware(value):
        return make_naive(value)
    return value

def reporte_resumen_conciliacion(p_conciliacion_id):
    import os
    import logging
    from django.conf import settings
    from openpyxl import Workbook
    from django.db.models import Case, When, IntegerField
    from ..models import InBox_PagosDetalle

    logger = logging.getLogger(__name__)
    wb = Workbook()

    headers = [
        "Conciliación", "Lote PSE", "Pago ID", "Valor Pago",
        "Clase Movimiento", "Fecha Pago", "Estado Pago", "Cliente",
        "Préstamo", "Fragmento de", "Fecha Conciliación",
        "Estado Conciliación", "Ref Cliente 3", "Archivo", "Fecha Carga Archivo",
    ]

    # --- DEFINICIÓN DE ESTILOS ---
    font_bold = Font(bold=True)
    # Color azul muy claro (Hex: D9EAD3 es verde claro, E8F0FE es azul profesional suave)
    fill_padre = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    def llenar_hoja(ws, queryset):
        ws.append(headers)
        # Opcional: Poner cabeceras en negrita también
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for p in queryset:
            fila = [
                p.conciliacion_id,
                p.lote_pse,
                p.pago_id,
                p.valor_pago,
                p.clase_movimiento,
                _excel_datetime(p.fecha_pago),
                p.estado_pago,
                p.cliente_id_real_id,
                p.prestamo_id_real_id,
                p.fragmento_de,
                _excel_datetime(p.fecha_conciliacion),
                p.estado_conciliacion,
                p.ref_cliente_3,
                str(p.nombre_archivo_id),
                _excel_datetime(p.fecha_carga_archivo),
            ]
            ws.append(fila)

            # --- APLICAR ESTILO SI ES PADRE (LOTE_PSE) ---
            if p.clase_movimiento == "LOTE_PSE":
                for cell in ws[ws.max_row]:
                    cell.font = font_bold
                    cell.fill = fill_padre

    # ---------------- CONCILIADOS ----------------
    ws_ok = wb.active
    ws_ok.title = "CONCILIADOS"
    
    qs_ok = (
        InBox_PagosDetalle.objects
        .filter(conciliacion_id=p_conciliacion_id)
        .annotate(
            orden=Case(
                When(clase_movimiento="LOTE_PSE", then=0),
                When(clase_movimiento="PAGO_PSE", then=1),
                default=2,
                output_field=IntegerField(),
            )
        )
        .order_by("lote_pse", "orden", "pago_id")
    )
    llenar_hoja(ws_ok, qs_ok)
    
    # ---------------- NO CONCILIADOS ----------------
    ws_no = wb.create_sheet("NO CONCILIADOS")
    qs_no = (
        InBox_PagosDetalle.objects
        .filter(estado_conciliacion="NO")
        .exclude(clase_movimiento="EXCLUIDO")  # <-- Filtro añadido para ignorar excluidos
        .annotate(
            orden=Case(
                When(clase_movimiento="LOTE_PSE", then=0),
                When(clase_movimiento="PAGO_PSE", then=1),
                default=2,
                output_field=IntegerField(),
            )
        )
        .order_by("lote_pse", "orden", "pago_id")
    )
    llenar_hoja(ws_no, qs_no)

    # ---------------- GUARDAR EN DISCO ----------------
    carpeta = os.path.join(settings.MEDIA_ROOT, "conciliaciones")
    os.makedirs(carpeta, exist_ok=True)

    nombre_archivo = f"resumen_conciliacion_{p_conciliacion_id}.xlsx"
    ruta = os.path.join(carpeta, nombre_archivo)

    wb.save(ruta)
    logger.info(f"Reporte generado con estilos: {ruta}")

    return ruta





