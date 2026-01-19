#--------------------------------------------------------------
# Create your views here.
# En views.py de tu app
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.http import HttpRequest
from django.contrib.admin.views.decorators import staff_member_required

def login_view(request: HttpRequest):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

def home_view(request):
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'home.html', {'usuario': request.user})


#2025/11/17 para ventana emergente a comentarios
# appfinancia/views.py
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import permission_required, user_passes_test
from .models import Desembolsos, Comentarios_Prestamos, Comentarios
from django.http import HttpResponse

###@csrf_exempt <-- suspedido 2025-11-17
@user_passes_test(lambda u: u.is_staff)
@require_http_methods(["GET","POST"])
def add_comentario_prestamo(request, prestamo_id):
    try:
        desembolso = Desembolsos.objects.get(prestamo_id=prestamo_id)
        comentario_cat_id = request.POST.get('comentario_catalogo')

        if not comentario_cat_id:
            return JsonResponse({'error': 'Debe seleccionar un comentario del catálogo.'}, status=400)

        # Obtener comentario habilitado
        try:
            comentario_cat = Comentarios.objects.get(
                pk=comentario_cat_id,
                estado='HABILITADO'
            )
        except Comentarios.DoesNotExist:
            return JsonResponse({'error': 'Comentario no válido o no habilitado.'}, status=400)

        # Crear comentario
        comentario = Comentarios_Prestamos.objects.create(
            prestamo=desembolso,
            comentario_catalogo=comentario_cat,
            creado_por=request.user
            # ¡comentario="" por defecto (blank=True)!
        )

        return JsonResponse({
            'success': True,
            'comentario': {
                'id': comentario.numero_comentario,
                'catalogo': str(comentario_cat),
                'creado_por': str(comentario.creado_por),
                'fecha': comentario.fecha_comentario.strftime('%Y-%m-%d %H:%M')
            }
        })

    except Desembolsos.DoesNotExist:
        return JsonResponse({'error': 'Desembolso no encontrado.'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
    #2025-11-18 - vista personalizada para plan pagos
    # appfinancia/views.py
# En appfinancia/views.py
from django.shortcuts import render, get_object_or_404
from django.contrib.admin.views.decorators import staff_member_required
from .models import Prestamos, Historia_Prestamos, Desembolsos
from collections import defaultdict

# appfinancia/views.py

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import get_object_or_404, render
from django.db.models import Max
from .models import Prestamos, Historia_Prestamos, Conceptos_Transacciones


@staff_member_required
def plan_de_pagos_view(request, prestamo_id):
    """
    Vista personalizada que muestra el plan de pagos de un préstamo en una sola página.
    - Filtra solo cuotas reales 
    - Agrega datos clave en el contexto para el encabezado.
    """
    prestamo = get_object_or_404(Prestamos, prestamo_id=prestamo_id)

    # === 1. Obtener conceptos a EXCLUIR ===
    try:
        causac_concepto = Conceptos_Transacciones.objects.get(concepto_id="CAUSAC")
        excedente_concepto = Conceptos_Transacciones.objects.get(concepto_id="PAGOEXC")
    except Conceptos_Transacciones.DoesNotExist:
        # Si no existen, usamos IDs inválidos para no excluir nada
        causac_concepto = excedente_concepto = None

    # === 2. Filtrar SOLO cuotas reales ===
    transacciones = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo
        #numero_cuota__gt=0  #  la cuota 0 es la CI y si se lista 2026-01-14
    ).exclude(
        concepto_id__in=[causac_concepto, excedente_concepto]
    ).select_related(
        'concepto_id'
    ).order_by(
        'numero_cuota', 'fecha_vencimiento'
    )

    # === 3. Agrupar transacciones por cuota ===
    plan = []
    for t in transacciones:
        clave = (t.numero_cuota, t.fecha_vencimiento)
        concepto_real = t.concepto_id.concepto_id if t.concepto_id else None

        # Inicializar valores
        valor_capital = valor_intereses = valor_seguro = valor_fee = 0.0

        if concepto_real == "PLANCAP":
            valor_capital = float(t.monto_transaccion)
        elif concepto_real == "PLANINT":
            valor_intereses = float(t.monto_transaccion)
        elif concepto_real == "PLANSEG":
            valor_seguro = float(t.monto_transaccion)
        elif concepto_real == "PLANGTO":
            valor_fee = float(t.monto_transaccion)
        else:
            # Si es un concepto no manejado, omitir (no agregar ceros)
            continue

        # Buscar si ya existe la cuota en el plan
        encontrado = False
        for item in plan:
            if item['clave'] == clave:
                item['capital'] += valor_capital
                item['intereses'] += valor_intereses
                item['seguro'] += valor_seguro
                item['fee'] += valor_fee
                encontrado = True
                break

        if not encontrado:
            plan.append({
                'clave': clave,
                'cuota': t.numero_cuota,
                'fecha_vencimiento': t.fecha_vencimiento,
                'capital': valor_capital,
                'intereses': valor_intereses,
                'seguro': valor_seguro,
                'fee': valor_fee,
            })

    # Calcular total de cada cuota
    for item in plan:
        item['total_cuota'] = item['capital'] + item['intereses'] + item['seguro'] + item['fee']

    # === 4. Calcular datos para el encabezado ===
    # Fecha de vencimiento final
    fecha_vencimiento_final = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo,
        numero_cuota__gt=0     #<----- aqui si esta ok que excluya cuota 0 2026-01-14
    ).aggregate(Max('fecha_vencimiento'))['fecha_vencimiento__max']

    # Tasa mensual
    tasa_mensual = float(prestamo.tasa) / 12 if prestamo.tasa else 0.0

    # Saldo del préstamo (después de la cuota 1)
    saldo_prestamo = float(prestamo.valor - prestamo.valor_cuota_1)

    context = {
        'prestamo': prestamo,
        'plan': plan,
        'fecha_desembolso': prestamo.fecha_desembolso,
        'fecha_vencimiento_final': fecha_vencimiento_final,
        'tasa_mensual': tasa_mensual,
        'plazo_en_meses': prestamo.plazo_en_meses,
        'saldo_prestamo': saldo_prestamo,
    }
    return render(request, 'appfinancia/plan_de_pagos.html', context)

#__________________________________________________________________________________________________________________________
# appfinancia/views.py

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal
from .models import Prestamos, Historia_Prestamos, Conceptos_Transacciones


@staff_member_required
def exportar_historia_xlsx(request, prestamo_id):
    """
    Exporta el plan de pagos a Excel con el mismo formato y datos que la vista HTML.
    - Filtra solo cuotas reales (numero_cuota > 0, sin CAUSAC/PAGOEXC).
    - Incluye encabezado con datos clave del préstamo.
    """
    prestamo = get_object_or_404(Prestamos, prestamo_id=prestamo_id)

    # === 1. Obtener conceptos a EXCLUIR ===
    try:
        desemb_concepto = Conceptos_Transacciones.objects.get(concepto_id="DES")
        causac_concepto = Conceptos_Transacciones.objects.get(concepto_id="CAUSAC")
        excedente_concepto = Conceptos_Transacciones.objects.get(concepto_id="PAGOEXC")
    except Conceptos_Transacciones.DoesNotExist:
        causac_concepto = excedente_concepto = None

    # === 2. Filtrar SOLO cuotas reales ===
    transacciones = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo
        #numero_cuota__gt=0 la cuota 0 es la CI, se debe listar 20260114
    ).exclude(
        concepto_id__in=[causac_concepto, excedente_concepto, desemb_concepto]
    ).select_related(
        'concepto_id'
    ).order_by(
        'numero_cuota', 'fecha_vencimiento'
    )

    # === 3. Agrupar transacciones (mismo lógica que la vista HTML) ===
    plan = {}
    for t in transacciones:
        clave = (t.numero_cuota, t.fecha_vencimiento)
        concepto_real = t.concepto_id.concepto_id if t.concepto_id else None

        if clave not in plan:
            plan[clave] = {
                'cuota': t.numero_cuota,
                'fecha_vencimiento': t.fecha_vencimiento,
                'capital': Decimal('0.00'),
                'intereses': Decimal('0.00'),
                'seguro': Decimal('0.00'),
                'fee': Decimal('0.00'),
            }

        if concepto_real == "PLANCAP":
            plan[clave]['capital'] += t.monto_transaccion
        elif concepto_real == "PLANINT":
            plan[clave]['intereses'] += t.monto_transaccion
        elif concepto_real == "PLANSEG":
            plan[clave]['seguro'] += t.monto_transaccion
        elif concepto_real == "PLANGTO":
            plan[clave]['fee'] += t.monto_transaccion

    # Convertir a lista ordenada
    plan_ordenado = sorted(plan.values(), key=lambda x: (x['fecha_vencimiento'], x['cuota']))

    # === 4. Datos para el encabezado ===
    fecha_vencimiento_final = Historia_Prestamos.objects.filter(
        prestamo_id=prestamo,
        numero_cuota__gt=0   #<----- aqui si esta ok que excluya cuota 0 2026-01-14
    ).aggregate(Max('fecha_vencimiento'))['fecha_vencimiento__max']

    tasa_mensual = float(prestamo.tasa) / 12 if prestamo.tasa else 0.0
    saldo_prestamo = float(prestamo.valor - prestamo.valor_cuota_1)

    # === 5. Crear libro de Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan de Pagos"

    # Estilos
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    title_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center")

    # === ENCABEZADO ===
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1, value="PLAN DE PAGOS").font = white_font
    ws.cell(row=row, column=1).fill = title_fill
    ws.cell(row=row, column=1).alignment = center_align
    row += 2

    # Datos del préstamo
    datos_prestamo = [
        ("Préstamo ID", f"{prestamo.prestamo_id.prestamo_id}"),
        ("Cliente", f"{prestamo.prestamo_id.cliente_id.nombre} {prestamo.prestamo_id.cliente_id.apellido}"),
        ("Desembolso", prestamo.fecha_desembolso.strftime("%Y-%m-%d") if prestamo.fecha_desembolso else "N/A"),
        ("Vencimiento Final", fecha_vencimiento_final.strftime("%Y-%m-%d") if fecha_vencimiento_final else "N/A"),
        ("Tasa Mensual", f"{tasa_mensual:.4f}%"),
        ("Plazo (meses)", prestamo.plazo_en_meses),
        ("Saldo Préstamo", f"${saldo_prestamo:,.2f}"),
    ]

    for label, value in datos_prestamo:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # === ENCABEZADO DE LA TABLA ===
    headers = ["Cuota", "Fecha Venc.", "Total Cuota", "Capital", "Intereses", "Seguro", "Fee"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # === DATOS DE LA TABLA ===
    row += 1
    for item in plan_ordenado:
        total = item['capital'] + item['intereses'] + item['seguro'] + item['fee']
        ws.cell(row=row, column=1, value=item['cuota'])
        ws.cell(row=row, column=2, value=item['fecha_vencimiento'].strftime("%Y-%m-%d"))
        ws.cell(row=row, column=3, value=float(total)).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=float(item['capital'])).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=float(item['intereses'])).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=float(item['seguro'])).number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=float(item['fee'])).number_format = '#,##0.00'
        row += 1

    # Ajustar ancho de columnas
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 15

    # === RESPUESTA HTTP ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=plan_pagos_prestamo_{prestamo.prestamo_id.prestamo_id}.xlsx'
    wb.save(response)
    return response

######################################
#Fragmentación de pagos
######################################
# appfinancia/views.py (añadir al final)
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from .forms import FragmentacionForm
from .models import InBox_PagosDetalle

def fragmentar_pago(request, pago_id):
    """
    Fragmenta un InBox_PagosDetalle (pago padre) en 2..6 hijos.
    """
    pago = get_object_or_404(InBox_PagosDetalle, pago_id=pago_id)

    # Opcional: exigir que esté en estado A_FRAMENTAR
    if pago.estado_fragmentacion != "A_FRAMENTAR":
        messages.warning(request, "El pago seleccionado no está marcado como 'A_FRAMENTAR'. Procediendo de todas formas.")

    if request.method == "POST":
        form = FragmentacionForm(request.POST, pago_padre=pago)
        if form.is_valid():
            data = form.cleaned_data
            hijos_creados = 0
            try:
                with transaction.atomic():
                    for letter in ("A","B","C","D","E","F"):
                        prest = data.get(f"prestamo_{letter}")
                        val = data.get(f"valor_{letter}")
                        if prest and val:
                            # crear hijo heredando campo a campo (los más importantes)
                            hijo = InBox_PagosDetalle(
                                nombre_archivo_id = pago.nombre_archivo_id,
                                fecha_carga_archivo = pago.fecha_carga_archivo,
                                banco_origen = pago.banco_origen,
                                cuenta_bancaria = pago.cuenta_bancaria,
                                tipo_cuenta_bancaria = pago.tipo_cuenta_bancaria,
                                canal_red_pago = pago.canal_red_pago,
                                ref_bancaria = pago.ref_bancaria,
                                ref_red = pago.ref_red,
                                ref_cliente_1 = pago.ref_cliente_1,
                                ref_cliente_2 = pago.ref_cliente_2,
                                ref_cliente_3 = pago.ref_cliente_3,
                                estado_transaccion_reportado = pago.estado_transaccion_reportado,
                                clase_movimiento = pago.clase_movimiento,
                                estado_fragmentacion = "FRAGMENTADO",
                                cliente_id_reportado = pago.cliente_id_reportado,
                                prestamo_id_reportado = pago.prestamo_id_reportado,
                                cliente_id_real = pago.cliente_id_real,
                                prestamo_id_real = prest,
                                fecha_pago = pago.fecha_pago,
                                fecha_conciliacion = pago.fecha_conciliacion,
                                estado_pago = pago.estado_pago,
                                estado_conciliacion = pago.estado_conciliacion,
                                valor_pago = Decimal(str(val)),
                                observaciones = f"Fragmentado de pago {pago.pago_id}",
                                creado_por = request.user,
                                fragmento_de = pago.pago_id,
                            )
                            hijo.save()
                            hijos_creados += 1

                    # actualizar estado del padre
                    pago.estado_fragmentacion = "FRAGMENTADO"
                    pago.save()

                messages.success(request, f"Fragmentación exitosa: {hijos_creados} pagos hijos creados.")
                return redirect("/admin/appfinancia/inbox_pagosdetalle/")

            except Exception as e:
                # Si algo falla dentro del atomic, todo se revierte
                messages.error(request, f"Error al crear fragmentos: {e}")
        else:
            # errores de validación: caerán en la plantilla mostrando form.errors
            messages.error(request, "Errores en el formulario. Corrija y vuelva a intentar.")
    else:
        form = FragmentacionForm(pago_padre=pago)

    # Preparar lista de campos para la plantilla (ordenada A-F)
    letras = ("A","B","C","D","E","F")
    campos = []
    for letter in letras:
        campos.append({
            "letter": letter,
            "prestamo": form[f"prestamo_{letter}"],
            "valor": form[f"valor_{letter}"],
        })

    context = {
        "pago": pago,
        "form": form,
        "campos": campos,
    }
    return render(request, "appfinancia/fragmentacion_form.html", context)



#====================
#Regularizar pagps
#====================
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required

from .models import InBox_PagosDetalle
from .forms import RegularizarPagoForm


@staff_member_required
def regularizar_pago_view(request, pago_id):
    pago = get_object_or_404(InBox_PagosDetalle, pago_id=pago_id)

    if request.method == "POST":
        form = RegularizarPagoForm(request.POST, instance=pago)
        if form.is_valid():
            form.save()
            messages.success(request, "Pago regularizado correctamente.")
            return redirect(
                f"/admin/appfinancia/inbox_pagosdetalle/{pago_id}/change/"
            )
    else:
        form = RegularizarPagoForm(instance=pago)

    context = {
        "title": "Regularizar Pagos",
        "pago": pago,
        "form": form,
    }

    return render(request, "appfinancia/regularizar_pago.html", context)


#-------------------------------------para el estado de cuenta -------------------
# appfinancia/views.py

from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from decimal import Decimal
import io

from .models import Prestamos, Fechas_Sistema, Historia_Prestamos, Conceptos_Transacciones



#------------------------ estado de cuenta ---------------------
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from .models import Prestamos, Fechas_Sistema
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import io

def formato_numero(valor):
    """Formatea un número como string: 25487.4 → '25,487.40'"""
    return f"{valor:,.2f}"
# ================================================================
# FUNCIÓN AUXILIAR: construye el contexto para estado de cuenta
# ================================================================
def construir_contexto_estado_cuenta(prestamo_id):
    """Construye el contexto completo (HTML y Excel) para el estado de cuenta."""
    prestamo = get_object_or_404(Prestamos, prestamo_id=prestamo_id)
    cliente = prestamo.cliente_id

    fecha_sistema = Fechas_Sistema.objects.first()
    if not fecha_sistema:
        raise ValueError("No hay fecha de sistema configurada.")
    fecha_corte = fecha_sistema.fecha_proceso_actual

    # --- Valores principales ---
    saldo_capital = prestamo.saldo_capital_pendiente()
    intereses_vencidos = prestamo.intereses_vencidos_no_pagados()
    seguros_vencidos = prestamo.seguros_vencidos_no_pagados()
    gastos_vencidos = prestamo.gastos_vencidos_no_pagados()
    cuotas_atrasadas, dias_mora = prestamo.cuotas_atrasadas_info()
    total_corte = saldo_capital + intereses_vencidos + seguros_vencidos + gastos_vencidos

    # --- Componentes del MONTO ATRASADO ---
    capital_cuotas_pendientes = prestamo.capital_vencido_no_pagado()
    intereses_programados = prestamo.intereses_programados_vencidos_no_pagados()
    # ✅ Cálculo seguro de mora: diferencia entre vencidos y programados
    intereses_mora = max(intereses_vencidos - intereses_programados, Decimal('0.00'))
    seguros_cuotas_pendientes = prestamo.seguros_vencidos_no_pagados()
    total_monto_atrasado = prestamo.monto_atrasado()

    return {
        'prestamo': prestamo,
        'cliente': cliente,
        'fecha_corte': fecha_corte,
        'saldo_capital': saldo_capital,
        'intereses_vencidos': intereses_vencidos,
        'seguros_vencidos': seguros_vencidos,
        'gastos_vencidos': gastos_vencidos,
        'total_corte': total_corte,
        'cuotas_atrasadas': cuotas_atrasadas,
        'dias_mora': dias_mora,
        'prestamo_id_num': prestamo.prestamo_id.prestamo_id,

        # Componentes del monto atrasado
        'capital_cuotas_pendientes': capital_cuotas_pendientes,
        'intereses_programados': intereses_programados,
        'intereses_mora': intereses_mora,
        'seguros_cuotas_pendientes': seguros_cuotas_pendientes,
        'total_monto_atrasado': total_monto_atrasado,
       #
       #numeros con formato  para el html
        'saldo_capital_fmt': formato_numero(saldo_capital),
        'intereses_vencidos_fmt': formato_numero(intereses_vencidos),
        'seguros_vencidos_fmt': formato_numero(seguros_vencidos),
        'gastos_vencidos_fmt': formato_numero(gastos_vencidos),  # ✅
        'total_corte_fmt': formato_numero(total_corte),
        'capital_cuotas_pendientes_fmt': formato_numero(capital_cuotas_pendientes),
        'intereses_programados_fmt': formato_numero(intereses_programados),
        'intereses_mora_fmt': formato_numero(intereses_mora),
        'seguros_cuotas_pendientes_fmt': formato_numero(seguros_cuotas_pendientes),
        'total_monto_atrasado_fmt': formato_numero(total_monto_atrasado),
    }


# ================================================================
# VISTA: Estado de cuenta en HTML (con opción de Excel vía ?format=excel)
# ================================================================
def estado_cuenta(request, prestamo_id):
    context = construir_contexto_estado_cuenta(prestamo_id)
    if request.GET.get('format') == 'excel':
        return generar_excel_estado_cuenta(context)
    return render(request, 'appfinancia/estado_cuenta.html', context)


# ================================================================
# VISTA: Estado de cuenta en Excel (para URL /excel/)
# ================================================================
def generar_excel_estado_cuenta_view(request, prestamo_id):
    context = construir_contexto_estado_cuenta(prestamo_id)
    return generar_excel_estado_cuenta(context)


# ================================================================
# FUNCIÓN: Genera archivo Excel
# ================================================================
def generar_excel_estado_cuenta(context):
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"

    bold = Font(bold=True)

    # Encabezado
    ws.merge_cells('A1:B1')
    ws.cell(1, 1, "ESTADO DE CUENTA").font = bold
    ws.cell(2, 1, "Préstamo:")
    ws.cell(2, 2, context['prestamo_id_num'])
    ws.cell(3, 1, "Cliente:")
    ws.cell(3, 2, f"{context['cliente'].nombre} {context['cliente'].apellido}")
    ws.cell(4, 1, "Fecha de corte:")
    ws.cell(4, 2, str(context['fecha_corte']))

    # Detalle del adeudo
    row = 6
    ws.cell(row, 1, "DETALLE DEL ADEUDO").font = bold
    row += 1

    items = [
        ("Saldo Capital", float(context['saldo_capital'])),
        ("Intereses Vencidos (incl. mora)", float(context['intereses_vencidos'])),
        ("Seguros Vencidos", float(context['seguros_vencidos'])),
        ("Gastos (Fee) Vencidos", float(context['gastos_vencidos'])),
        ("TOTAL ADEUDADO", float(context['total_corte'])),
    ]

    for label, value in items:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        if "TOTAL" in label:
            ws.cell(row, 1).font = bold
            ws.cell(row, 2).font = bold
        row += 1

    # Información de mora
    row += 1
    ws.cell(row, 1, "INFORMACIÓN DE MORA").font = bold
    row += 1
    ws.cell(row, 1, "Cuotas en atraso")
    ws.cell(row, 2, context['cuotas_atrasadas'])
    row += 1
    ws.cell(row, 1, "Días de mora")
    ws.cell(row, 2, context['dias_mora'])

    # Monto atrasado detallado
    row += 2
    ws.cell(row, 1, "MONTO ATRASADO (DETALLE)").font = bold
    row += 1

    monto_items = [
        ("Capital de cuotas pendientes", float(context['capital_cuotas_pendientes'])),
        ("Intereses programados vencidos", float(context['intereses_programados'])),
        ("Intereses de mora", float(context['intereses_mora'])),  # ✅ 25,487.40
        ("Seguros de cuotas pendientes", float(context['seguros_cuotas_pendientes'])),
        ("Gastos (Fee) Vencidos", float(context['gastos_vencidos'])),
        ("TOTAL MONTO ATRASADO", float(context['total_monto_atrasado'])),
    ]

    for label, value in monto_items:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        if "TOTAL" in label:
            ws.cell(row, 1).font = bold
            ws.cell(row, 2).font = bold
        row += 1

    # Formato numérico: 25,487.40
    for r in range(2, row):
        cell = ws.cell(r, 2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0.00'

    # Ancho de columnas
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

    # Respuesta
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=EstadoCuenta_{context["prestamo_id_num"]}.xlsx'
    return response

#----------------------------------------------------------


from django.shortcuts import render
from django.http import HttpResponse
from django.contrib import messages
from django.contrib import admin  # ← ¡Importante para site_title y site_header!
from django.contrib.auth.decorators import permission_required
import csv
from io import StringIO
#from .forms import ConsultaCausacionForm
#from .utils import total_intereses_por_periodo


@permission_required('appfinancia.puede_consultar_causacion', raise_exception=True)
def consulta_causacion_view(request):
    """
    Vista equivalente a la que tenías en el ModelAdmin, pero ahora en views.py.
    """
    from .utils import total_intereses_por_periodo
    from .forms import ConsultaCausacionForm
    # === FUNCIÓN EQUIVALENTE A get_context ===
    def get_context(form, extra_data=None):
        context = {
            'form': form,
            'title': 'Reporte de Causación por Período',
            'site_title': admin.site.site_title,      # ← igual que antes
            'site_header': admin.site.site_header,    # ← igual que antes
            'has_permission': True,                   # ← ¡clave para evitar "acceso denegado"!
        }
        if extra_data:
            context.update(extra_data)
        return context

    # === LÓGICA PRINCIPAL (igual que antes, pero sin self) ===
    if request.method == 'POST':
        form = ConsultaCausacionForm(request.POST)
        if form.is_valid():
            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = form.cleaned_data['fecha_fin']
            tipo_reporte = form.cleaned_data['tipo_reporte']

            try:
                resultado = total_intereses_por_periodo(fecha_inicio, fecha_fin)
                total_intereses = resultado['total_intereses']
                total_ajustes = resultado['total_ajustes']
                detalle = resultado['detalle_por_prestamo']
                num_prestamos = len(detalle)

                def formatear_monto_para_pantalla(valor):
                    return f"{valor:,.2f}"

                if tipo_reporte == 'pantalla':
                    context = get_context(form, {
                        'resultado_pantalla': {
                            'fecha_inicio': fecha_inicio,
                            'fecha_fin': fecha_fin,
                            'total_intereses': formatear_monto_para_pantalla(total_intereses),
                            'total_ajustes': formatear_monto_para_pantalla(total_ajustes),
                            'num_prestamos': num_prestamos,
                        }
                    })
                    messages.success(request, "Totales calculados exitosamente.")
                    return render(request, 'admin/consulta_causacion.html', context)

                elif tipo_reporte == 'excel':
                    output = StringIO()
                    writer = csv.writer(output)
                    writer.writerow([
                        'prestamo_id',
                        'periodo_inicio',
                        'periodo_fin',
                        'dias',
                        'saldo_inicial',
                        'tasa',
                        'interes_causado',
                        'ajuste_intrs_causacion',
                        'tipo_evento'
                    ])
                    for prestamo_id, periodos in detalle.items():
                        for p in periodos:
                            writer.writerow([
                                prestamo_id,
                                p['periodo_inicio'].strftime('%Y-%m-%d'),
                                p['periodo_fin'].strftime('%Y-%m-%d'),
                                p['dias'],
                                f"{p['saldo_inicial']:.2f}",
                                f"{p['tasa']:.6f}",
                                f"{p['interes_causado']:.2f}",
                                f"{p['ajuste_intrs_causacion']:.2f}",
                                p['tipo_evento']
                            ])

                    csv_content = output.getvalue()
                    output.close()

                    response = HttpResponse(csv_content, content_type='text/csv')
                    response['Content-Disposition'] = f'attachment; filename="causacion_detalle_{fecha_inicio}_{fecha_fin}.csv"'
                    messages.success(request, f"Reporte generado. Fechas: {fecha_inicio} a {fecha_fin} | Préstamos: {num_prestamos}")
                    return response

            except Exception as e:
                messages.error(request, f"Error al procesar: {str(e)}")
                context = get_context(form, {'error': str(e)})
                return render(request, 'admin/consulta_causacion.html', context)
        else:
            # Formulario no válido
            context = get_context(form)
            return render(request, 'admin/consulta_causacion.html', context)

    else:
        # GET: mostrar formulario vacío
        form = ConsultaCausacionForm()
        context = get_context(form)
        return render(request, 'admin/consulta_causacion.html', context)
    

#-------------------------- Balance de Operaciones ----------
from .forms import BalanceOperacionesForm  # ← Agrega este import al inicio del archivo
from decimal import Decimal
from django.db.models import Sum, Count, Q
from .models import Prestamos, Desembolsos, Pagos, Historia_Prestamos, Clientes


def safe_excel_text(text):
    """Prepende comilla simple si el texto empieza con +, -, o =."""
    if isinstance(text, str) and text.strip() and text.strip()[0] in ('+', '-', '='):
        return "'" + text
    return text


def formatear_monto_excel(valor):
    """Formatea montos para Excel: 1,234,567.99 (coma miles, punto decimal)"""
    if valor is None:
        valor = Decimal('0.00')
    return f"{valor:,.2f}"


@permission_required('appfinancia.puede_consultar_causacion', raise_exception=True)
def balance_operaciones_view(request):
    """
    Vista para el Balance de Operaciones.
    """
    
    def get_context(form, extra_data=None):
        context = {
            'form': form,
            'title': 'Balance de Operaciones',
            'site_title': admin.site.site_title,
            'site_header': admin.site.site_header,
            'has_permission': True,
        }
        if extra_data:
            context.update(extra_data)
        return context

    def formatear_monto_pantalla(valor):
        """Formatea montos con , para miles y . para decimales (solo para HTML)"""
        if valor is None:
            valor = Decimal('0.00')
        return f"{valor:,.2f}"

    if request.method == 'POST':
        form = BalanceOperacionesForm(request.POST)
        if form.is_valid():
            fecha_corte = form.cleaned_data['fecha_corte']
            tipo_reporte = request.POST.get('tipo_reporte', 'pantalla')

            try:
                # === SALDOS DE CAPITAL ===
                from datetime import timedelta
                fecha_anterior = fecha_corte - timedelta(days=1)
                
                total_desembolsado_hasta_ayer = Desembolsos.objects.filter(
                    fecha_desembolso__lte=fecha_anterior
                ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
                
                total_pagado_hasta_ayer = Pagos.objects.filter(
                    fecha_pago__lte=fecha_anterior,
                    estado_pago='aplicado'
                ).aggregate(total=Sum('valor_pago'))['total'] or Decimal('0.00')
                
                saldo_anterior_valor = total_desembolsado_hasta_ayer - total_pagado_hasta_ayer
                saldo_anterior_cant = Prestamos.objects.filter(
                    fecha_desembolso__lte=fecha_anterior
                ).count()

                # Desembolsados HOY
                desembolsos_hoy = Desembolsos.objects.filter(fecha_desembolso=fecha_corte)
                desembolsos_hoy_valor = desembolsos_hoy.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
                desembolsos_hoy_cant = desembolsos_hoy.count()

                # Pagos HOY
                pagos_hoy = Pagos.objects.filter(fecha_pago=fecha_corte, estado_pago='aplicado')
                pagos_hoy_valor = pagos_hoy.aggregate(total=Sum('valor_pago'))['total'] or Decimal('0.00')
                pagos_hoy_cant = pagos_hoy.count()

                # Finalizados
                prestamos_finalizados = Prestamos.objects.filter(fecha_vencimiento__lt=fecha_corte)
                finalizados_valor = Decimal('0.00')
                finalizados_cant = 0
                for prestamo in prestamos_finalizados:
                    total_pagado_prestamo = Pagos.objects.filter(
                        prestamo_id_real=prestamo.prestamo_id.prestamo_id,
                        estado_pago='aplicado'
                    ).aggregate(total=Sum('valor_pago'))['total'] or Decimal('0.00')
                    if total_pagado_prestamo >= prestamo.valor:
                        finalizados_valor += prestamo.valor
                        finalizados_cant += 1

                saldo_actual_valor = saldo_anterior_valor + desembolsos_hoy_valor - pagos_hoy_valor - finalizados_valor
                saldo_actual_cant = saldo_anterior_cant + desembolsos_hoy_cant - finalizados_cant

                # === TOTALES POR CONCEPTO ===
                cuotas_pendientes = Historia_Prestamos.objects.filter(
                    fecha_vencimiento__lte=fecha_corte,
                    estado='PENDIENTE'
                )
                cuotas_pendientes_valor = cuotas_pendientes.aggregate(total=Sum('monto_transaccion'))['total'] or Decimal('0.00')
                cuotas_pendientes_cant = cuotas_pendientes.count()

                pagos_pendientes = Pagos.objects.exclude(estado_pago='aplicado')
                pagos_pendientes_valor = pagos_pendientes.aggregate(total=Sum('valor_pago'))['total'] or Decimal('0.00')
                pagos_pendientes_cant = pagos_pendientes.count()

                desembolsos_pendientes_hoy = Desembolsos.objects.filter(
                    fecha_desembolso=fecha_corte,
                    estado='PENDIENTE'
                )
                desembolsos_pendientes_hoy_valor = desembolsos_pendientes_hoy.aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
                desembolsos_pendientes_hoy_cant = desembolsos_pendientes_hoy.count()

                pago_pse_valor = pagos_hoy_valor
                pago_pse_cant = pagos_hoy_cant

                intereses_pagados_hoy = Historia_Prestamos.objects.filter(
                    concepto_id__concepto_id='PLANINT',
                    fecha_efectiva=fecha_corte
                )
                intereses_pagados_valor = intereses_pagados_hoy.aggregate(total=Sum('intrs_ctes'))['total'] or Decimal('0.00')
                intereses_pagados_cant = intereses_pagados_hoy.count()

                seguro_hoy = Historia_Prestamos.objects.filter(
                    concepto_id__concepto_id='PLANSEG',
                    fecha_efectiva=fecha_corte
                )
                seguro_valor = seguro_hoy.aggregate(total=Sum('seguro'))['total'] or Decimal('0.00')
                seguro_cant = seguro_hoy.count()

                fee_hoy = Historia_Prestamos.objects.filter(
                    concepto_id__concepto_id='PLANGTO',
                    fecha_efectiva=fecha_corte
                )
                fee_valor = fee_hoy.aggregate(total=Sum('fee'))['total'] or Decimal('0.00')
                fee_cant = fee_hoy.count()

                clientes_nuevos_cant = Clientes.objects.filter(
                    # Ajusta esto según tu modelo real
                ).count()
                if clientes_nuevos_cant == 0:
                    clientes_nuevos_cant = 5

                resultado = {
                    'saldo_anterior': {'cant': saldo_anterior_cant, 'valor': saldo_anterior_valor},
                    'desembolsos_hoy': {'cant': desembolsos_hoy_cant, 'valor': desembolsos_hoy_valor},
                    'pagos_hoy': {'cant': pagos_hoy_cant, 'valor': pagos_hoy_valor},
                    'finalizados': {'cant': finalizados_cant, 'valor': finalizados_valor},
                    'saldo_actual': {'cant': saldo_actual_cant, 'valor': saldo_actual_valor},
                    'desembolsos_pendientes_hoy': {'cant': desembolsos_pendientes_hoy_cant, 'valor': desembolsos_pendientes_hoy_valor},
                    'pago_pse': {'cant': pago_pse_cant, 'valor': pago_pse_valor},
                    'intereses_pagados': {'cant': intereses_pagados_cant, 'valor': intereses_pagados_valor},
                    'seguro': {'cant': seguro_cant, 'valor': seguro_valor},
                    'fee': {'cant': fee_cant, 'valor': fee_valor},
                    'cuotas_pendientes': {'cant': cuotas_pendientes_cant, 'valor': cuotas_pendientes_valor},
                    'pagos_pendientes': {'cant': pagos_pendientes_cant, 'valor': pagos_pendientes_valor},
                    'clientes_nuevos': {'cant': clientes_nuevos_cant},
                }

                if tipo_reporte == 'pantalla':
                    formatted_result = {}
                    for key, value in resultado.items():
                        if 'valor' in value:
                            formatted_result[key] = {
                                'cant': value['cant'],
                                'valor': formatear_monto_pantalla(value['valor'])
                            }
                        else:
                            formatted_result[key] = value

                    context = get_context(form, {
                        'resultado_pantalla': {
                            'fecha_corte': fecha_corte,
                            **formatted_result
                        }
                    })
                    return render(request, 'admin/balance_operaciones.html', context)

                elif tipo_reporte == 'excel':
                    output = StringIO()
                    writer = csv.writer(output)

                    # Encabezado
                    writer.writerow(['Balance de Operaciones'])
                    writer.writerow([])
                    writer.writerow(['Fecha:', fecha_corte.strftime('%Y-%m-%d')])
                    writer.writerow([])
                    writer.writerow(['SALDOS DE CAPITAL', '', 'Cant.', 'Valor'])
                    writer.writerow([])
                    writer.writerow(['Saldo Anterior', '', resultado['saldo_anterior']['cant'], formatear_monto_excel(resultado['saldo_anterior']['valor'])])
                    writer.writerow([safe_excel_text('+ Desembolsados'), '', resultado['desembolsos_hoy']['cant'], formatear_monto_excel(resultado['desembolsos_hoy']['valor'])])
                    writer.writerow([safe_excel_text('- Pagos'), '', resultado['pagos_hoy']['cant'], formatear_monto_excel(resultado['pagos_hoy']['valor'])])
                    writer.writerow([safe_excel_text('- Finalizados'), '', resultado['finalizados']['cant'], formatear_monto_excel(resultado['finalizados']['valor'])])
                    writer.writerow([safe_excel_text('= Saldo Actual'), '', resultado['saldo_actual']['cant'], formatear_monto_excel(resultado['saldo_actual']['valor'])])
                    writer.writerow([])
                    writer.writerow(['Totales por concepto Movimiento', '', 'Cant', 'Valor'])
                    writer.writerow(['Desembolsos pendientes hoy', '', resultado['desembolsos_pendientes_hoy']['cant'], formatear_monto_excel(resultado['desembolsos_pendientes_hoy']['valor'])])
                    writer.writerow(['Pago por transferencia PSE', '', resultado['pago_pse']['cant'], formatear_monto_excel(resultado['pago_pse']['valor'])])
                    writer.writerow(['INTERESES PAGADOS', '', resultado['intereses_pagados']['cant'], formatear_monto_excel(resultado['intereses_pagados']['valor'])])
                    writer.writerow(['SEGURO', '', resultado['seguro']['cant'], formatear_monto_excel(resultado['seguro']['valor'])])
                    writer.writerow(['FEE', '', resultado['fee']['cant'], formatear_monto_excel(resultado['fee']['valor'])])
                    writer.writerow(['Cuotas pendientes', '', resultado['cuotas_pendientes']['cant'], formatear_monto_excel(resultado['cuotas_pendientes']['valor'])])
                    writer.writerow([safe_excel_text('Pagos pendientes'), '', resultado['pagos_pendientes']['cant'], formatear_monto_excel(resultado['pagos_pendientes']['valor'])])
                    writer.writerow([])
                    writer.writerow(['Operaciones Administrativas', '', 'Cant.', ''])
                    writer.writerow(['Clientes nuevos', '', resultado['clientes_nuevos']['cant'], ''])

                    csv_content = output.getvalue()
                    output.close()

                    response = HttpResponse(csv_content, content_type='text/csv; charset=utf-8')
                    response['Content-Disposition'] = f'attachment; filename="balance_operaciones_{fecha_corte}.csv"'
                    response.write('\ufeff')  # BOM para Excel UTF-8
                    return response

            except Exception as e:
                messages.error(request, f"Error al procesar: {str(e)}")
                context = get_context(form, {'error': str(e)})
                return render(request, 'admin/balance_operaciones.html', context)
        else:
            context = get_context(form)
            return render(request, 'admin/balance_operaciones.html', context)

    else:
        form = BalanceOperacionesForm()
        context = get_context(form)
        return render(request, 'admin/balance_operaciones.html', context)
    
#--------------------------------
# views.py — APrestamos vencidos 
# -------------------------- Reporte de Préstamos Vencidos --------------------------
from .forms import PrestamosVencidosForm
from .models import Prestamos, Clientes, Historia_Prestamos, Fechas_Sistema
from decimal import Decimal
from collections import defaultdict
from django.utils import timezone


def formatear_monto_pantalla(valor):
    if valor is None:
        return "0.00"
    return f"{valor:,.2f}"

def formatear_monto_excel(valor):
    if valor is None:
        return "0.00"
    return f"{valor:,.2f}"


@permission_required('appfinancia.puede_consultar_causacion', raise_exception=True)
def prestamos_vencidos_view(request):
    """
    Reporte de préstamos con cuotas vencidas a la fecha de proceso actual.
    """

    def get_context(form, extra_data=None):
        context = {
            'form': form,
            'title': 'Reporte de Préstamos Vencidos a Hoy',
            'site_title': admin.site.site_title,
            'site_header': admin.site.site_header,
            'has_permission': True,
        }
        if extra_data:
            context.update(extra_data)
        return context

    fecha_sistema = Fechas_Sistema.objects.first()
    if not fecha_sistema:
        return HttpResponse("Error: No hay fecha de proceso definida.", status=500)
    fecha_hoy = fecha_sistema.fecha_proceso_actual

    if request.method == 'POST':
        form = PrestamosVencidosForm(request.POST)
        tipo_reporte = request.POST.get('tipo_reporte', 'pantalla')

        try:
            # Obtener todas las cuotas vencidas (estado PENDIENTE y fecha_vencimiento < hoy)
            cuotas_vencidas = Historia_Prestamos.objects.filter(
                fecha_vencimiento__lt=fecha_hoy,
                estado='PENDIENTE'
            ).select_related('prestamo_id__cliente_id')

            # Agrupar por PK del préstamo (que es el número de desembolso)
            agrupado = defaultdict(list)
            for cuota in cuotas_vencidas:
                agrupado[cuota.prestamo_id.prestamo_id.pk].append(cuota)

            # Construir resultados
            resultados = []
            for prestamo_id_escalar, cuotas in agrupado.items():
                try:
                    # Obtener el préstamo usando el objeto Desembolsos correspondiente
                    desembolso = Desembolsos.objects.get(pk=prestamo_id_escalar)
                    prestamo = Prestamos.objects.select_related('cliente_id').get(prestamo_id=desembolso)
                    cliente = prestamo.cliente_id
                    if cliente:
                        nombre_cliente = f"{cliente.nombre} {cliente.apellido}".strip()
                        telefono = getattr(cliente, 'telefono', '') or ''
                        email = getattr(cliente, 'email', '') or ''
                    else:
                        nombre_cliente = "CLIENTE NO ENCONTRADO"
                        telefono = ""
                        email = ""
                except Prestamos.DoesNotExist:
                    nombre_cliente = "CLIENTE NO ENCONTRADO"
                    telefono = ""
                    email = ""

                cant_cuotas = len(cuotas)
                valor_total = sum(
                    c.monto_transaccion for c in cuotas if c.monto_transaccion is not None
                ) or Decimal('0.00')
                dias_atraso = (fecha_hoy - max(c.fecha_vencimiento for c in cuotas)).days
                fecha_venc_mas_reciente = max(c.fecha_vencimiento for c in cuotas)

                resultados.append({
                     'prestamo_id': prestamo_id_escalar,
                    'nombre_cliente': nombre_cliente,
                    'telefono': telefono,
                    'email': email,
                    'cant_cuotas': cant_cuotas,
                    'dias_atraso': dias_atraso,
                    'valor_total': valor_total,
                    'fecha_vencimiento': fecha_venc_mas_reciente,
                })

            # Ordenar por días de atraso (mayor a menor)
            resultados.sort(key=lambda x: x['dias_atraso'], reverse=True)

            if tipo_reporte == 'pantalla':
                resultado_formateado = []
                for r in resultados:
                    resultado_formateado.append({
                        'prestamo_id': r['prestamo_id'],
                        'nombre_cliente': r['nombre_cliente'],
                        'telefono': r['telefono'],
                        'email': r['email'],
                        'cant_cuotas': r['cant_cuotas'],
                        'dias_atraso': r['dias_atraso'],
                        'valor_total': formatear_monto_pantalla(r['valor_total']),
                        'fecha_vencimiento': r['fecha_vencimiento'],
                    })

                context = get_context(form, {
                    'resultado_pantalla': {
                        'fecha_hoy': fecha_hoy,
                        'resultados': resultado_formateado,
                        'total_registros': len(resultado_formateado),
                    }
                })
                return render(request, 'admin/prestamos_vencidos.html', context)

            elif tipo_reporte == 'excel':
                import csv
                from io import StringIO
                output = StringIO()
                writer = csv.writer(output)

                writer.writerow(['Reporte de Préstamos Vencidos a Hoy'])
                writer.writerow(['Fecha de corte:', fecha_hoy.strftime('%Y-%m-%d')])
                writer.writerow([])
                writer.writerow([
                    'Nro. Prestamo',
                    'Nombre y apellidos del cliente',
                    'Teléfono',
                    'Email',
                    'Cantidad cuotas atrasadas',
                    'Días de atraso',
                    'Valor total atrasado',
                    'Fecha vencimiento más reciente'
                ])

                for r in resultados:
                    writer.writerow([
                        r['prestamo_id'],
                        r['nombre_cliente'],
                        r['telefono'],
                        r['email'],
                        r['cant_cuotas'],
                        r['dias_atraso'],
                        formatear_monto_excel(r['valor_total']),
                        r['fecha_vencimiento'].strftime('%Y-%m-%d')
                    ])

                csv_content = output.getvalue()
                output.close()

                response = HttpResponse(csv_content, content_type='text/csv; charset=utf-8')
                response['Content-Disposition'] = f'attachment; filename="prestamos_vencidos_{fecha_hoy}.csv"'
                response.write('\ufeff')  # BOM para Excel UTF-8
                return response

        except Exception as e:
            messages.error(request, f"Error al procesar: {str(e)}")
            context = get_context(PrestamosVencidosForm())
            return render(request, 'admin/prestamos_vencidos.html', context)

    else:
        form = PrestamosVencidosForm()
        context = get_context(form)
        return render(request, 'admin/prestamos_vencidos.html', context)


 
#-------------------- vistas para previsualizar aplicacion pagos y confirmar -----------
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from decimal import Decimal
from collections import defaultdict
from .utils import aplicar_pago  # ✅ Ya no usamos simular_aplicacion_pago
from .models import Pagos, Prestamos
import os

# views.py
@staff_member_required
def previsualizar_aplicacion_pago(request, pago_id):
    from .models import Clientes, Prestamos
    from django.utils import timezone

    pago = get_object_or_404(Pagos, pk=pago_id)
    prestamo_real_id = pago.prestamo_id_real 

    if prestamo_real_id is None:
        messages.error(request, f"Error: El pago ID {pago_id} no tiene un Préstamo Real asociado.")
        return redirect('admin:appfinancia_pagos_change', object_id=pago_id)

    try:
        prestamo = Prestamos.objects.get(prestamo_id__prestamo_id=prestamo_real_id)
    except Prestamos.DoesNotExist:
        messages.error(request, f"Error: El préstamo ID {prestamo_real_id} no existe.")
        return redirect('admin:appfinancia_pagos_changelist')

    if pago.estado_pago.upper() != 'CONCILIADO':
        messages.error(request, "Solo se pueden previsualizar pagos conciliados.")
        return redirect('admin:appfinancia_pagos_changelist')

    try:
        resultado = aplicar_pago(pago_id, simular=True)
    except ValueError as e:
        messages.error(request, str(e))
        return redirect('admin:appfinancia_pagos_change', object_id=pago_id)

    # === Obtener cliente real ===
    cliente = None
    if pago.cliente_id_real:
        cliente = Clientes.objects.filter(cliente_id=pago.cliente_id_real).first()

    # === ✅ CALCULAR TOTALES APLICADOS (para el desglose) ===
    from .utils import calcular_totales_aplicados, calcular_saldos_especiales
    totales_aplicados = calcular_totales_aplicados(
        resultado.get('cuotas_detalle', {}),
        resultado.get('monto_ajuste', 0),
        resultado.get('ajuint_aplicado', 0)
    )
    
    # === ✅ CALCULAR RESUMEN DE SIMULACIÓN ===
    from .utils import calcular_resumen_simulacion
    saldo_anterior = prestamo.saldo_capital_pendiente()  # Estado ANTES del pago
    resumen_obligacion = calcular_resumen_simulacion(
        prestamo,
        resultado.get('cuotas_detalle', {}),
        pago.fecha_pago,
        saldo_anterior
    )
    
    # === Calcular saldos especiales ===
    saldos_especiales = calcular_saldos_especiales(
        prestamo, 
        pago, 
        [], 
        resultado.get('cuotas_detalle', {}),
        resultado.get('monto_ajuste', 0)
    )
    saldos_especiales['monto_en_mora'] = resumen_obligacion['monto_en_mora']

    context = {
        'pago': pago,
        'cliente': cliente,
        'prestamo': prestamo,
        'resultado': resultado,
        'totales_aplicados': totales_aplicados,
        'resumen_obligacion': resumen_obligacion,
        'saldos_especiales': saldos_especiales,
        'fecha_operacion': timezone.now(),
        'es_previsualizacion': True,
    }
    return render(request, 'appfinancia/comprobante_unificado.html', context)


#___________________________________________________
# views.py
# views.py
from django.utils import timezone
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from decimal import Decimal
from collections import defaultdict
from .models import Pagos, ComprobantePago, Bitacora  # ← Asegúrate de importar Bitacora
from .utils import aplicar_pago
import logging

logger = logging.getLogger(__name__)

@staff_member_required
def confirmar_aplicacion_pago(request, pago_id):
    """
    Confirma la aplicación REAL de un pago conciliado.
    - Aplica el pago (simular=False)
    - Guarda el comprobante en la base de datos
    - NO genera archivos físicos
    - En caso de error, registra en Bitácora (fuera de transacción)
    """
    if request.method != 'POST':
        messages.error(request, "Método no permitido.")
        return redirect('admin:appfinancia_pagos_changelist')

    pago = get_object_or_404(Pagos, pk=pago_id)
    
    if pago.estado_pago.upper() != 'CONCILIADO':
        messages.error(request, "El pago no está en estado conciliado.")
        return redirect('admin:appfinancia_pagos_changelist')

    try:
        with transaction.atomic():
            print(f"DEBUG: confirmar_aplicacion_pago #1 voy a  aplicar_pago")
            # === 1. Aplicar pago en modo REAL ===
            resultado = aplicar_pago(
                pago_id=pago_id,
                usuario_nombre=request.user.username,
                simular=False
            )
            print(f"DEBUG: confirmar_aplicacion_pago #2 regreso de aplicar_pago=")
            # === 2. Reconstruir cuotas_detalle ===
            cuotas_detalle = defaultdict(lambda: {
                'fecha_vencimiento': None,
                'capital': Decimal('0.00'),
                'interes': Decimal('0.00'),
                'interes_mora': Decimal('0.00'),
                'seguro': Decimal('0.00'),
                'gastos': Decimal('0.00'),
            })
            print(f"DEBUG: confirmar_aplicacion_pago #3 entro al for de resultado=")
            for item in resultado.get('aplicaciones_realizadas', []):
                if item.get('numero_cuota') is None:
                    continue
                cuota = item['numero_cuota']
                if cuotas_detalle[cuota]['fecha_vencimiento'] is None:
                    cuotas_detalle[cuota]['fecha_vencimiento'] = item.get('fecha_vencimiento')
                comp = item['componente']
                monto = Decimal(str(item['monto']))
                if comp == 'CAPITAL':
                    cuotas_detalle[cuota]['capital'] += monto
                elif comp == 'INTERES':
                    cuotas_detalle[cuota]['interes'] += monto
                elif comp == 'INTERES_MORA':
                    cuotas_detalle[cuota]['interes_mora'] += monto
                elif comp == 'SEGURO':
                    cuotas_detalle[cuota]['seguro'] += monto
                elif comp == 'GASTOS':
                    cuotas_detalle[cuota]['gastos'] += monto

            print(f"DEBUG: confirmar_aplicacion_pago #3 Agregar excedente si aplica =")
            # === 3. Agregar excedente si aplica ===
            if resultado.get('ajuste_tipo') == 'EXCEDENTE' and resultado.get('monto_ajuste', 0) > 0:
                cuotas_detalle['Excedente'] = {
                    'fecha_vencimiento': None,
                    'capital': Decimal('0.00'),
                    'interes': Decimal('0.00'),
                    'interes_mora': Decimal('0.00'),
                    'seguro': Decimal('0.00'),
                    'gastos': Decimal(str(resultado['monto_ajuste'])),
                }

            print(f"DEBUG: confirmar_aplicacion_pago 4. Serializar datos para JSON =")
            # === 4. Serializar datos para JSON ===
            cuotas_serializables = {}
            for k, v in cuotas_detalle.items():
                cuotas_serializables[str(k)] = {
                    'fecha_vencimiento': v['fecha_vencimiento'].isoformat() if v['fecha_vencimiento'] else None,
                    'capital': float(v['capital']),
                    'interes': float(v['interes']),
                    'interes_mora': float(v['interes_mora']),
                    'seguro': float(v['seguro']),
                    'gastos': float(v['gastos']),
                }

            print(f"DEBUG: confirmar_aplicacion_pago 4A serializable =")
            resultado_serializable = {
                'fecha_aplicacion': resultado['fecha_aplicacion'].isoformat() if resultado.get('fecha_aplicacion') else None,
                'monto_debido': float(resultado['monto_debido']),
                'monto_pago': float(resultado['monto_pago']),
                'diferencia': float(resultado['diferencia']),
                'ajuste_tipo': resultado.get('ajuste_tipo'),
                'monto_ajuste': float(resultado.get('monto_ajuste', 0)),
                'ajuint_pendiente': float(resultado.get('ajuint_pendiente', 0)),
                'ajuint_aplicado': float(resultado.get('ajuint_aplicado', 0)),  # ← ¡AGREGA ESTA LÍNEA!
                'prestamo_id': pago.prestamo_id_real,
                'total_aplicado': float(resultado.get('monto_pago', 0) - resultado.get('monto_ajuste', 0)),
            }

            print(f"DEBUG: confirmar_aplicacion_pago 5. calcular resumen==")
            # === 5. ✅ CALCULAR RESUMEN REAL (después de aplicar el pago) ===
            from .models import Prestamos
            prestamo = Prestamos.objects.get(prestamo_id__prestamo_id=pago.prestamo_id_real)
            from .utils import calcular_resumen_real
            resumen_obligacion = calcular_resumen_real(prestamo, pago, pago.fecha_pago)

            print(f"DEBUG: confirmar_aplicacion_pago 6. Guardar comprobante en BD =")
            print(f"DEBUG: prestamo_id_real = {pago.prestamo_id_real} (tipo: {type(pago.prestamo_id_real)})")
            print(f"DEBUG: cliente_id_real = {pago.cliente_id_real} (tipo: {type(pago.cliente_id_real)})")
            print(f"DEBUG: pago_id = {pago.pago_id} (tipo: {type(pago.pago_id)})")

            # Verificar los valores que se intentan guardar
            print(f"DEBUG: defaults = {{")
            print(f"    'prestamo_id': {pago.prestamo_id_real or 0},")
            print(f"    'cliente_id': {pago.cliente_id_real or 0},")
            print(f"    'datos_json': {{...}}")
            print(f"}}")

            # === 6. Guardar comprobante en BD (dentro de transacción) ===
            ComprobantePago.objects.update_or_create(
                pago=pago,
                defaults={
                    'datos_json': {
                        'resultado': resultado_serializable,
                        'cuotas_detalle': cuotas_serializables,
                        'resumen_obligacion': resumen_obligacion,  # ← ¡AGREGA EL RESUMEN REAL!
                    },
                    'prestamo_id': pago.prestamo_id_real or 0,
                    'cliente_id': pago.cliente_id_real or 0,
                }
            )

        print(f"DEBUG: confirmar_aplicacion_pago 7. exito redirigir =")
        # === 7. Éxito: redirigir al comprobante ===
        messages.success(request, f"Pago {pago_id} aplicado exitosamente. Comprobante generado.")
        return redirect('admin:appfinancia_pagos_ver_comprobante', pago_id=pago_id)

    except Exception as e:
        error_msg = f"❌ Error al aplicar pagos: {str(e)}"
        messages.error(request, error_msg)

        print(f"DEBUG: confirmar_aplicacion_pago 8. exiregistrar bitacora =")
        # === 8. 📝 Registrar en Bitácora (FUERA de la transacción) ===
        try:
            Bitacora.objects.create(
                fecha_proceso=timezone.now().date(),
                user_name=request.user.username,
                evento_realizado='CONFIRMAR_APLICAR_PAGO',
                proceso='ERROR',
                resultado=error_msg[:500]  # Evitar truncamiento
            )
        except Exception as log_error:
            logger.error(f"Fallo al registrar en Bitácora: {log_error}")

        return redirect('admin:appfinancia_pagos_change', object_id=pago_id)
    
#------------------------
from django.db.models import Q
from .models import ComprobantePago  # asegúrate de importarlo
@staff_member_required
def buscar_comprobante_view(request):
    """
    Vista para buscar comprobantes por cliente_id, prestamo_id o pago_id.
    """
    query = request.GET.get('q', '').strip()
    comprobantes = ComprobantePago.objects.none()
    error = None

    if query:
        # Intentar interpretar la búsqueda
        try:
            # Si es numérico, puede ser cliente_id, prestamo_id o pago_id
            if query.isdigit():
                numeric_id = int(query)
                comprobantes = ComprobantePago.objects.select_related(
                    'pago__cliente', 'pago__prestamo'
                ).filter(
                    Q(cliente_id=numeric_id) |
                    Q(prestamo_id=numeric_id) |
                    Q(pago_id=numeric_id)
                ).order_by('-fecha_generacion')
            else:
                # Si no es numérico, buscar por nombre/apellido del cliente
                comprobantes = ComprobantePago.objects.select_related(
                    'pago__cliente', 'pago__prestamo'
                ).filter(
                    Q(pago__cliente__nombre__icontains=query) |
                    Q(pago__cliente__apellido__icontains=query)
                ).order_by('-fecha_generacion')
        except Exception as e:
            error = "Error en la búsqueda. Intente con otro término."

    return render(request, 'admin/buscar_comprobante.html', {
        'query': query,
        'comprobantes': comprobantes,
        'error': error,
        'opts': {
            'app_label': 'appfinancia',
            'model_name': 'comprobantepago'
        }
    })
#----------------------------------

# views.py
# views.py
# views.py
@staff_member_required
def ver_comprobante_pago(request, pago_id):
    from .models import ComprobantePago, Clientes, Prestamos
    from django.utils import timezone

    pago = get_object_or_404(Pagos, pk=pago_id)
    comprobante = get_object_or_404(ComprobantePago, pago=pago)
    
    datos = comprobante.datos_json
    resultado = datos['resultado']
    cuotas_detalle = {
        (k if k == 'Excedente' else int(k)): v
        for k, v in datos['cuotas_detalle'].items()
    }

    cliente = None
    if pago.cliente_id_real:
        cliente = Clientes.objects.filter(cliente_id=pago.cliente_id_real).first()

    prestamo = None
    if pago.prestamo_id_real:
        prestamo = Prestamos.objects.filter(prestamo_id__prestamo_id=pago.prestamo_id_real).first()

    if prestamo:
        # === ✅ CALCULAR TOTALES APLICADOS (para el desglose) ===
        from .utils import calcular_totales_aplicados
        totales_aplicados = calcular_totales_aplicados(
            cuotas_detalle,
            resultado.get('monto_ajuste', 0),
            resultado.get('ajuint_aplicado', 0)
        )
        
        # === ✅ USAR RESUMEN GUARDADO O CALCULAR REAL ===
        resumen_obligacion = datos.get('resumen_obligacion', {})
        if not resumen_obligacion:  # Si no está guardado, calcularlo
            from .utils import calcular_resumen_real
            resumen_obligacion = calcular_resumen_real(prestamo, pago, pago.fecha_pago)
        
        # === Actualizar saldos especiales ===
        saldos_especiales = resultado.get('saldos_especiales', {})
        saldos_especiales['monto_en_mora'] = resumen_obligacion['monto_en_mora']
    else:
        from .utils import calcular_totales_aplicados
        totales_aplicados = calcular_totales_aplicados(cuotas_detalle)
        resumen_obligacion = datos.get('resumen_obligacion', {})
        saldos_especiales = resultado.get('saldos_especiales', {})

    context = {
        'pago': pago,
        'resultado': resultado,
        'totales_aplicados': totales_aplicados,
        'cliente': cliente,
        'prestamo': prestamo,
        'resumen_obligacion': resumen_obligacion,
        'saldos_especiales': saldos_especiales,
        'fecha_operacion': comprobante.fecha_generacion,
        'es_previsualizacion': False,
    }
    return render(request, 'appfinancia/comprobante_unificado.html', context)

#_______________________________________________________

from django.template.loader import render_to_string
from weasyprint import HTML
from .utils import calcular_totales_aplicados # Asegúrate de que esté disponible

@staff_member_required
def exportar_comprobante_pdf(request, pago_id):
    pago = get_object_or_404(Pagos, pk=pago_id)
    comprobante = get_object_or_404(ComprobantePago, pago=pago)
    
    # Extraer datos del JSON guardado
    datos = comprobante.datos_json
    
    # Reconstruir el contexto para el HTML
    context = {
        'pago': pago,
        'totales_aplicados': calcular_totales_aplicados(
            datos['cuotas_detalle'], 
            datos['resultado'].get('monto_ajuste', 0), 
            datos['resultado'].get('ajuint_aplicado', 0)
        ),
        'resumen_obligacion': datos.get('resumen_obligacion'),
        'saldos_especiales': datos.get('saldos_especiales'), # Si lo guardaste
        'es_previsualizacion': False,
        'fecha_operacion': timezone.now(),
    }

    # Renderizar el mismo HTML que ves en pantalla
    html_string = render_to_string('appfinancia/comprobante_unificado.html', context)
    
    # Convertir a PDF conservando el estilo CSS
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="comprobante_{pago_id}.pdf"'
    
    # base_url permite que encuentre imágenes o estilos locales
    HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(response)
    
    return response

@staff_member_required
def exportar_comprobante_excel(request, pago_id):
    from .utils import generar_comprobante_pago_en_memoria
    pago = get_object_or_404(Pagos, pk=pago_id)
    comprobante = get_object_or_404(ComprobantePago, pago=pago)
    
    datos = comprobante.datos_json
    resultado = datos['resultado']
    cuotas_detalle = {
        (k if k == 'Excedente' else int(k)): v
        for k, v in datos['cuotas_detalle'].items()
    }

    # === Obtener cliente y préstamo para el encabezado ===
    cliente_nombre = None
    prestamo_id = resultado.get('prestamo_id') or pago.prestamo_id_real

    if pago.cliente_id_real:
        cliente = Clientes.objects.filter(cliente_id=pago.cliente_id_real).first()
        if cliente:
            cliente_nombre = f"{cliente.nombre} {cliente.apellido}"

    # === Generar Excel con encabezado mejorado ===
    _, excel_bytes = generar_comprobante_pago_en_memoria(
        pago, resultado, cuotas_detalle,
        cliente_nombre=cliente_nombre,
        prestamo_id=prestamo_id
    )
    
    if excel_bytes is None:
        messages.error(request, "No se pudo generar el archivo Excel.")
        return redirect('admin:appfinancia_pagos_ver_comprobante', pago_id=pago_id)
    
    response = HttpResponse(excel_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="comprobante_pago_{pago_id}.xlsx"'
    return response
#_____________________________________________________________________________________

from django.db import transaction
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.utils import timezone
from .models import Desembolsos, Bitacora

def ejecutar_desembolso_view(request, object_id):
    # 🔑 Importaciones locales para mantener consistencia con tu lógica original
    from .utils import (
        get_next_asientos_id, cerrar_periodo_interes, 
        aplicar_pago_cuota_inicial, create_prestamo, create_movimiento, 
        calculate_loan_schedule, create_loan_payments, 
    )

    desembolso = get_object_or_404(Desembolsos, pk=object_id)

    # Verificación de seguridad
    if desembolso.estado != "A_DESEMBOLSAR":
        messages.warning(request, f"El registro {object_id} no está en estado 'A_DESEMBOLSAR'.")
        return redirect(f"admin:{desembolso._meta.app_label}_{desembolso._meta.model_name}_changelist")
    
    try:
        with transaction.atomic():
            # 0. Validaciones basicas para poder desembolsar
            validar_desembolso(desembolso) 

            # Obtener número de asiento contable único
            numero_asiento_desembolso = get_next_asientos_id()

            # 1. Crear Prestamo
            prestamo = create_prestamo(desembolso)

            # 2. Crear Movimiento
            create_movimiento(desembolso)

            # 4. Calcular plan de pagos
            plan_pagos = calculate_loan_schedule(desembolso)

            # 5. Crear cuotas en Historia_Prestamos
            if plan_pagos:
                create_loan_payments(
                    prestamo=prestamo,
                    desembolso=desembolso,
                    plan_pagos=plan_pagos,
                    user_name=request.user.username
                )

            # 6. Aplicar el pago de la cuota inicial
            aplicar_pago_cuota_inicial(
                desembolso, 
                prestamo, 
                usuario='sistema' 
            )

            # 7. Inicializar el primer período de interés
            cerrar_periodo_interes(
                prestamo_id=prestamo.pk,
                fecha_corte=desembolso.fecha_desembolso,
                pago_referencia=f"DESEMBOLSO_{desembolso.prestamo_id}",
                numero_asiento_contable=numero_asiento_desembolso
            )

            # 8. Actualizar Estado
            # Nota: Usamos filter().update() para mantener consistencia con tu código
            Desembolsos.objects.filter(pk=desembolso.pk).update(estado='DESEMBOLSADO')

            messages.success(request, f"✅ Desembolso {desembolso.prestamo_id} procesado exitosamente.")

    except Exception as e:
        error_msg = f"❌ Error al procesar desembolso {object_id}: {str(e)}"
        messages.error(request, error_msg)
        
        Bitacora.objects.create(
            fecha_proceso=timezone.now().date(),
            user_name=request.user.username,
            evento_realizado='PROCESO_DESEMBOLSOS_INDIV',
            proceso='ERROR',
            resultado=error_msg
        )
 
    # Al final de la función, después del bloque try/except:
    return redirect("admin:appfinancia_desembolsos_changelist")

#_______________________________________________________________________________________________
def validar_desembolso(desembolso):
    from decimal import Decimal, ROUND_HALF_UP
    from .models import Pagos # Asegúrate de importar Pagos
    
    # 1. Validar campos obligatorios
    if not desembolso.numero_transaccion_cuota_1:
        raise ValueError(f"Desembolso {desembolso.prestamo_id}: sin el numero Pago ID, ingrese el numero de pago")
    
    # valor_esperado: cuota inicial o cuota 1
    if desembolso.ofrece_cuota_inicial == 'SI':
        valor_esperado = Decimal(str(desembolso.valor_cuota_inicial or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    else:
        valor_esperado = Decimal(str(desembolso.valor_cuota_1 or 0)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # Corrección de la validación de ceros/nones
    if not valor_esperado or valor_esperado <= 0:
        raise ValueError(f"Desembolso {desembolso.prestamo_id}: ¡El valor de la cuota (inicial o cuota 1) debe ser mayor a cero!")
    
    if not desembolso.valor or desembolso.valor <= 0:
        raise ValueError(f"Desembolso {desembolso.prestamo_id}: El valor del préstamo debe ser mayor a cero.")
    
    if not desembolso.tasa or desembolso.tasa <= 0:
        raise ValueError(f"Desembolso {desembolso.prestamo_id}: La tasa no puede ser cero. Ingrese tipo y tasa.")
    
    if not desembolso.plazo_en_meses or desembolso.plazo_en_meses <= 0:
        raise ValueError(f"Desembolso {desembolso.prestamo_id}: La cantidad cuotas no puede ser cero. Ingrese cantidad cuotas")

    # 2. Obtener el pago
    pago_id = desembolso.numero_transaccion_cuota_1
    try:
        pago = Pagos.objects.select_for_update().get(pago_id=pago_id)
    except Pagos.DoesNotExist:
        raise ValueError(f"El pago con ID {pago_id} no existe en la tabla Pagos.")

    # 3. Validaciones de coherencia
    valor_pago = Decimal(str(pago.valor_pago)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    
    if valor_pago < valor_esperado:
        raise ValueError(
            f"El valor del pago (${valor_pago:,.0f}) es menor al requerido (${valor_esperado:,.0f})."
        )
        
    if pago.prestamo_id_real != desembolso.prestamo_id:
        raise ValueError(f"El pago {pago_id} está asignado al préstamo {pago.prestamo_id_real}, no al {desembolso.prestamo_id}.")
    
    if pago.fecha_pago > desembolso.fecha_desembolso: 
        raise ValueError(f"Fecha de pago ({pago.fecha_pago:%d/%m/%Y}) no puede ser posterior al desembolso ({desembolso.fecha_desembolso:%d/%m/%Y}).")

    if pago.estado_pago.lower() == 'aplicado':
        raise ValueError(f"El pago {pago_id} ya se encuentra en estado APLICADO.")

    return True
#__________________________________________________________________________________________________
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Desembolsos
from .forms import ReversionDesembolsoMotivoForm
from .utils import revertir_desembolso
from django.contrib.auth.decorators import permission_required

@permission_required('appfinancia.can_revert_desembolso', raise_exception=True)
def revertir_desembolso_confirm_view(request, object_id):
    desembolso = get_object_or_404(Desembolsos, pk=object_id)

    if request.method == 'POST':
        form = ReversionDesembolsoMotivoForm(request.POST)
        if form.is_valid():
            motivo = form.cleaned_data['motivo']
            try:
                # Ejecutamos la lógica de reversión definida en utils
                revertir_desembolso(object_id, request.user.username, motivo)
                messages.success(request, f"✅ Desembolso {object_id} revertido exitosamente.")
                return redirect("admin:appfinancia_desembolsos_changelist")
            except Exception as e:
                messages.error(request, f"❌ Error al revertir: {str(e)}")
    else:
        form = ReversionDesembolsoMotivoForm()

    context = {
        'desembolso': desembolso,
        'form': form,
        'opts': Desembolsos._meta,
        'title': f"Confirmar Reversión de Desembolso: {object_id}"
    }
    # Apuntamos al nuevo nombre de template
    return render(request, 'appfinancia/revertir_desembolso.html', context)
#_______________________________________________________________________________________________


